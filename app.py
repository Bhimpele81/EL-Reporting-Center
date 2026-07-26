"""
EL Reporting Center — Flask Application
-----------------------------------------
Drop-in Excel report converter for Elbow Lane Day Camp.
Shares the same design system as Transport Pro.
"""

import os
import io
import re
import csv
import json
import uuid
import threading
import urllib.request
from functools import wraps
from datetime import datetime, date, timedelta
try:
    from zoneinfo import ZoneInfo
    _EASTERN = ZoneInfo("America/New_York")
except Exception:
    _EASTERN = None
import boto3
from botocore.exceptions import ClientError
from flask import Flask, request, jsonify, send_file, render_template_string, session
from werkzeug.security import generate_password_hash, check_password_hash

from report_processor import process_report, load_bunk_config, save_bunk_config, is_master, parse_master, bunk_snapshot_data

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024  # 32 MB upload limit
# Stable secret for signed-cookie sessions. Set SECRET_KEY in the environment
# (esp. on multi-worker hosts); the fallback keeps sessions working otherwise.
app.secret_key = os.environ.get("SECRET_KEY", "el-reporting-center-default-secret-change-me")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=14)
# Shared code required to register a new account (augments per-user logins)
ACCESS_CODE = os.environ.get("ACCESS_CODE", "trial")

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "bunk_config.json")
UPLOAD_DIR  = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR  = os.path.join(BASE_DIR, "outputs")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# S3 setup (optional — falls back to local if env vars not set)
S3_BUCKET = os.environ.get("AWS_S3_BUCKET")
S3_REGION = os.environ.get("AWS_S3_REGION", "us-east-2")
_s3 = boto3.client(
    "s3",
    region_name=S3_REGION,
    aws_access_key_id=os.environ.get("AWS_ACCESS_KEY_ID"),
    aws_secret_access_key=os.environ.get("AWS_SECRET_ACCESS_KEY"),
) if S3_BUCKET else None

def _s3_upload(local_path: str, filename: str) -> None:
    if _s3:
        _s3.upload_file(local_path, S3_BUCKET, filename)

def _s3_get_file(filename: str):
    """Download file from S3 into a BytesIO buffer and return it."""
    if not _s3:
        return None
    try:
        import io
        buf = io.BytesIO()
        _s3.download_fileobj(S3_BUCKET, filename, buf)
        buf.seek(0)
        return buf
    except ClientError:
        return None

def _s3_save_config(config: dict) -> None:
    """Save bunk config JSON to S3 so it persists across Render restarts."""
    if not _s3:
        return
    try:
        import io
        body = json.dumps(config, indent=2).encode("utf-8")
        _s3.put_object(Bucket=S3_BUCKET, Key="bunk_config.json", Body=body, ContentType="application/json")
    except ClientError as e:
        print(f"S3 config save failed: {e}")

def _s3_load_config() -> dict | None:
    """Load bunk config JSON from S3. Returns None if not found."""
    if not _s3:
        return None
    try:
        import io
        buf = io.BytesIO()
        _s3.download_fileobj(S3_BUCKET, "bunk_config.json", buf)
        buf.seek(0)
        return json.load(buf)
    except ClientError:
        return None


def _save_config_meta(saved_by: str) -> dict:
    """Record who/when last saved the bunk config (for the audit box)."""
    meta = {"saved_at": _now_eastern_stamp(), "saved_by": saved_by or ""}
    try:
        with open(LOCAL_CONFIG_META, "w", encoding="utf-8") as f:
            json.dump(meta, f)
    except Exception:
        pass
    if _s3:
        try:
            _s3.put_object(Bucket=S3_BUCKET, Key=CONFIG_META_KEY,
                           Body=json.dumps(meta).encode("utf-8"),
                           ContentType="application/json")
        except ClientError:
            pass
    return meta


def _load_config_meta() -> dict:
    data = None
    if _s3:
        buf = _s3_get_file(CONFIG_META_KEY)
        if buf:
            try:
                data = json.load(buf)
            except Exception:
                data = None
    if data is None and os.path.exists(LOCAL_CONFIG_META):
        try:
            with open(LOCAL_CONFIG_META, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = None
    return data or {}


def _s3_list_recent(limit: int = 10) -> list:
    if not _s3:
        return []
    resp = _s3.list_objects_v2(Bucket=S3_BUCKET)
    # Only show generated report files — not internal config/master storage
    objects = [o for o in resp.get("Contents", [])
               if o["Key"].lower().endswith((".xlsx", ".docx", ".zip"))]
    objects.sort(key=lambda o: o["LastModified"], reverse=True)
    return objects[:limit]

def _s3_delete_old(keep: int = 10) -> None:
    if not _s3:
        return
    resp = _s3.list_objects_v2(Bucket=S3_BUCKET)
    # Never sweep the saved config or master sheet when pruning old outputs
    objects = sorted(
        [o for o in resp.get("Contents", []) if o["Key"] not in _PROTECTED_KEYS],
        key=lambda o: o["LastModified"], reverse=True,
    )
    for obj in objects[keep:]:
        try:
            _s3.delete_object(Bucket=S3_BUCKET, Key=obj["Key"])
        except ClientError:
            pass


# ---------------------------------------------------------------------------
# Saved master sheet — uploaded once, reused for every report until replaced
# ---------------------------------------------------------------------------

# Reports that take a camp-week selection
WEEK_AWARE_REPORTS = {"driver_totals", "group_attendance",
                      "am_extend", "pm_extend", "pm_grp_extend",
                      "inter_labels", "jr_transport_labels", "upper_labels"}

MASTER_KEY        = "current_master.dat"
MASTER_META_KEY   = "current_master_meta.json"
LOCAL_MASTER      = os.path.join(UPLOAD_DIR, "current_master.dat")
LOCAL_MASTER_META = os.path.join(UPLOAD_DIR, "current_master_meta.json")
PAYROLL_KEY       = "payroll.json"
LOCAL_PAYROLL     = os.path.join(UPLOAD_DIR, "payroll.json")
SEED_PATH         = os.path.join(BASE_DIR, "payroll_seed.json")
WEEK1_MONDAY      = date(2026, 6, 22)   # first Monday of the 2026 season
FAMILIES_KEY      = "families.json"
LOCAL_FAMILIES    = os.path.join(UPLOAD_DIR, "families.json")
USERS_KEY         = "users.json"
LOCAL_USERS       = os.path.join(UPLOAD_DIR, "users.json")
CONFIG_META_KEY   = "bunk_config_meta.json"
LOCAL_CONFIG_META = os.path.join(UPLOAD_DIR, "bunk_config_meta.json")
SEASON_KEY        = "season.json"
LOCAL_SEASON      = os.path.join(UPLOAD_DIR, "season.json")
SCHEDULES_KEY     = "schedules.json"
LOCAL_SCHEDULES   = os.path.join(UPLOAD_DIR, "schedules.json")
PRICING_KEY       = "pricing.json"
LOCAL_PRICING     = os.path.join(UPLOAD_DIR, "pricing.json")
# Default season: Monday of each of the 8 camp weeks (2026)
_DEFAULT_SEASON_MONDAYS = ["2026-06-22", "2026-06-29", "2026-07-06", "2026-07-13",
                           "2026-07-20", "2026-07-27", "2026-08-03", "2026-08-10"]
# Fields stored for each family contact record (matches the contact master export)
FAMILY_FIELDS     = ["last", "first", "family", "bunk",
                     "primary_first", "primary_last", "primary_phone", "primary_email",
                     "secondary_first", "secondary_last", "secondary_phone", "secondary_email",
                     "address", "address2", "city", "state", "zip",
                     "pu1_name", "pu1_auth", "pu2_name", "pu2_auth",
                     "pu3_name", "pu3_auth", "pu4_name", "pu4_auth"]
_PROTECTED_KEYS   = {"bunk_config.json", CONFIG_META_KEY, MASTER_KEY, MASTER_META_KEY,
                     PAYROLL_KEY, FAMILIES_KEY, USERS_KEY, SEASON_KEY, SCHEDULES_KEY, PRICING_KEY}


def _camper_key(name, bunk):
    """Stable key for a camper (name + bunk), used by schedule overrides."""
    return f"{(name or '').strip().lower()}||{(bunk or '').strip().lower()}"


def _canon_days(s):
    """Canonicalize a day string to M/T/W/R/F in order."""
    up = (s or "").upper()
    return "".join(L for L in "MTWRF" if L in up)


def _now_eastern_stamp() -> str:
    """Formatted Eastern-time timestamp, e.g. '6/19/2026 4:00 PM EDT'."""
    now = datetime.now(_EASTERN) if _EASTERN else datetime.now()
    fmt = "%#m/%#d/%Y %#I:%M %p %Z" if os.name == "nt" else "%-m/%-d/%Y %-I:%M %p %Z"
    return now.strftime(fmt).strip()


def _save_master(file_bytes: bytes, filename: str, uploaded_by: str = "") -> dict:
    """Persist the uploaded master sheet (S3 if configured, plus local copy)."""
    meta = {
        "filename":    filename or "master",
        "uploaded_at": _now_eastern_stamp(),
        "uploaded_by": uploaded_by or "",
        "size":        len(file_bytes),
    }
    try:
        with open(LOCAL_MASTER, "wb") as f:
            f.write(file_bytes)
        with open(LOCAL_MASTER_META, "w") as f:
            json.dump(meta, f)
    except Exception:
        pass
    if _s3:
        try:
            _s3.put_object(Bucket=S3_BUCKET, Key=MASTER_KEY, Body=file_bytes)
            _s3.put_object(Bucket=S3_BUCKET, Key=MASTER_META_KEY,
                           Body=json.dumps(meta).encode("utf-8"),
                           ContentType="application/json")
        except ClientError:
            pass
    return meta


def _load_master() -> bytes | None:
    """Return the saved master sheet bytes, or None if none stored."""
    if _s3:
        buf = _s3_get_file(MASTER_KEY)
        if buf:
            return buf.read()
    if os.path.exists(LOCAL_MASTER):
        try:
            with open(LOCAL_MASTER, "rb") as f:
                return f.read()
        except Exception:
            pass
    return None


def _load_master_meta() -> dict | None:
    """Return metadata about the saved master (filename, uploaded_at)."""
    if _s3:
        buf = _s3_get_file(MASTER_META_KEY)
        if buf:
            try:
                return json.load(buf)
            except Exception:
                pass
    if os.path.exists(LOCAL_MASTER_META):
        try:
            with open(LOCAL_MASTER_META) as f:
                return json.load(f)
        except Exception:
            pass
    return None

# ---------------------------------------------------------------------------
# Payroll attendance (staff roster + daily check-ins) — persisted like config
# ---------------------------------------------------------------------------

# --- Season calendar (editable 8-week Monday dates) ---

def _season_save(data: dict) -> None:
    try:
        with open(LOCAL_SEASON, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass
    if _s3:
        try:
            _s3.put_object(Bucket=S3_BUCKET, Key=SEASON_KEY,
                           Body=json.dumps(data).encode("utf-8"),
                           ContentType="application/json")
        except ClientError:
            pass


def _season_load() -> dict:
    data = None
    if _s3:
        buf = _s3_get_file(SEASON_KEY)
        if buf:
            try:
                data = json.load(buf)
            except Exception:
                data = None
    if data is None and os.path.exists(LOCAL_SEASON):
        try:
            with open(LOCAL_SEASON, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = None
    if data is None:
        data = {}
    mondays = data.get("mondays") or []
    # Pad/trim to exactly 8, falling back to defaults for any blanks
    mondays = [(mondays[i] if i < len(mondays) and mondays[i] else _DEFAULT_SEASON_MONDAYS[i])
               for i in range(8)]
    return {"mondays": mondays}


def _season_mondays() -> list:
    out = []
    for iso in _season_load()["mondays"]:
        try:
            out.append(date.fromisoformat(iso))
        except (ValueError, TypeError):
            out.append(None)
    return out


def _week_range_str(monday: date) -> str:
    """'June 22 – 26' (same month) or 'June 29 – July 3' (spanning months)."""
    if monday is None:
        return ""
    fri = monday + timedelta(days=4)
    if monday.month == fri.month:
        return f"{monday.strftime('%B')} {monday.day} – {fri.day}"
    return f"{monday.strftime('%B')} {monday.day} – {fri.strftime('%B')} {fri.day}"


def _season_week_strings() -> list:
    """The 8 date-range strings used in report headers."""
    return [_week_range_str(m) for m in _season_mondays()]


def _current_week_day():
    """(week 1-8, day 0-4) for today in the season, or (None, None) if outside
    the camp weeks. day is 0=Mon..4=Fri (None on weekends within a week)."""
    today = (datetime.now(_EASTERN) if _EASTERN else datetime.now()).date()
    mondays = _season_mondays()
    for wk in range(8):
        base = mondays[wk] or date.fromisoformat(_DEFAULT_SEASON_MONDAYS[wk])
        delta = (today - base).days
        if 0 <= delta <= 6:
            return wk + 1, (delta if delta <= 4 else None)
    return None, None


# --- Per-camper, per-week day-schedule overrides ---

def _schedules_save(data: dict) -> None:
    try:
        with open(LOCAL_SCHEDULES, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass
    if _s3:
        try:
            _s3.put_object(Bucket=S3_BUCKET, Key=SCHEDULES_KEY,
                           Body=json.dumps(data).encode("utf-8"),
                           ContentType="application/json")
        except ClientError:
            pass


def _schedules_load() -> dict:
    data = None
    if _s3:
        buf = _s3_get_file(SCHEDULES_KEY)
        if buf:
            try:
                data = json.load(buf)
            except Exception:
                data = None
    if data is None and os.path.exists(LOCAL_SCHEDULES):
        try:
            with open(LOCAL_SCHEDULES, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = None
    if data is None:
        data = {}
    data.setdefault("overrides", {})   # { camper_key: { "1": "MWF", ... } }
    return data


# --- Pricing calculator / explorer (admin) ---

# Seeded from the 2026 worksheets; all values are editable in the app.
_DEFAULT_PRICING = {
    "season_label": "2026",
    "camp": {
        "week_order": ["8", "7", "6", "5", "4", "Mini"],
        # Two rate tiers by week-count: ES = early-signup (fall), Final = regular
        "tiers": {
            "ES":    {"8": 5580, "7": 5415, "6": 5080, "5": 4575, "4": 3905, "Mini": 550},
            "Final": {"8": 6000, "7": 5820, "6": 5460, "5": 4920, "4": 4200, "Mini": 550},
        },
        # Multiplier applied to the (weeks) tuition for 5/4/3 days per week.
        # Flat (1.0) for now — camp may provide day-based rates later.
        "day_mult": {"5": 1.0, "4": 1.0, "3": 1.0},
    },
    # Weekly transportation add-on, by days per week
    "transport": {
        "2way": {"5": 160, "4": 140, "3": 120},
        "1way": {"5": 120, "4": 120, "3": 120},
    },
    # Childcare / school (separate from camp): weekly rate by days per week.
    #   base     = one child
    #   sibling2 = combined weekly rate for two siblings (~130% of base)
    "childcare": {
        "5": {"base": 485, "sibling2": 595},
        "4": {"base": 435, "sibling2": 530},
        "3": {"base": 395, "sibling2": 485},
    },
}


def _deep_fill(target, defaults):
    """Recursively add any keys present in defaults but missing from target."""
    for k, v in defaults.items():
        if isinstance(v, dict):
            target[k] = target.get(k) if isinstance(target.get(k), dict) else {}
            _deep_fill(target[k], v)
        else:
            target.setdefault(k, v)
    return target


def _pricing_save(data: dict) -> None:
    try:
        with open(LOCAL_PRICING, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass
    if _s3:
        try:
            _s3.put_object(Bucket=S3_BUCKET, Key=PRICING_KEY,
                           Body=json.dumps(data).encode("utf-8"),
                           ContentType="application/json")
        except ClientError:
            pass


def _pricing_load() -> dict:
    data = None
    if _s3:
        buf = _s3_get_file(PRICING_KEY)
        if buf:
            try:
                data = json.load(buf)
            except Exception:
                data = None
    if data is None and os.path.exists(LOCAL_PRICING):
        try:
            with open(LOCAL_PRICING, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = None
    if not isinstance(data, dict):
        data = {}
    return _deep_fill(data, _DEFAULT_PRICING)


def _payroll_days() -> list:
    """The 40 camp days (8 weeks x Mon-Fri) with day-of-week + m/d labels,
    derived from each week's Monday in the season calendar."""
    dows = ["MON", "TUES", "WED", "TH", "FRI"]
    mondays = _season_mondays()
    out = []
    for wk in range(8):
        base = mondays[wk] or (date.fromisoformat(_DEFAULT_SEASON_MONDAYS[wk]))
        for dow in range(5):
            d = base + timedelta(days=dow)
            out.append({"iso": d.isoformat(), "dow": dows[dow],
                        "md": f"{d.month}/{d.day}", "week": wk + 1})
    return out


def _payroll_save(data: dict) -> None:
    try:
        with open(LOCAL_PAYROLL, "w") as f:
            json.dump(data, f)
    except Exception:
        pass
    if _s3:
        try:
            _s3.put_object(Bucket=S3_BUCKET, Key=PAYROLL_KEY,
                           Body=json.dumps(data).encode("utf-8"),
                           ContentType="application/json")
        except ClientError:
            pass


def _payroll_load() -> dict:
    data = None
    if _s3:
        buf = _s3_get_file(PAYROLL_KEY)
        if buf:
            try:
                data = json.load(buf)
            except Exception:
                data = None
    if data is None and os.path.exists(LOCAL_PAYROLL):
        try:
            with open(LOCAL_PAYROLL) as f:
                data = json.load(f)
        except Exception:
            data = None
    if data is None:
        # First run — seed the roster from the bundled seed file
        try:
            with open(SEED_PATH) as f:
                seed = json.load(f)
            data = {"staff": seed.get("staff", []), "checks": {}}
        except Exception:
            data = {"staff": [], "checks": {}}
        _payroll_save(data)
    data.setdefault("staff", [])
    data.setdefault("checks", {})
    data.setdefault("locked", False)
    # Backfill 'bunk'/'title'/'ext' (added later) from the seed for staff missing them
    if any(("bunk" not in s or "title" not in s or "ext" not in s) for s in data["staff"]):
        try:
            with open(SEED_PATH) as f:
                seed_map = {(x["last"].lower(), x["first"].lower()): x
                            for x in json.load(f).get("staff", [])}
        except Exception:
            seed_map = {}
        for s in data["staff"]:
            sm = seed_map.get((s.get("last", "").lower(), s.get("first", "").lower()), {})
            if "bunk" not in s:
                s["bunk"] = sm.get("bunk", "")
            if "title" not in s:
                s["title"] = sm.get("title", "")
            if "ext" not in s:
                s["ext"] = sm.get("ext", "")
        _payroll_save(data)
    return data


# ---------------------------------------------------------------------------
# Family contacts — uploaded/imported once, then editable in the Utilities tab
# ---------------------------------------------------------------------------

def _families_save(data: dict) -> None:
    try:
        with open(LOCAL_FAMILIES, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass
    if _s3:
        try:
            _s3.put_object(Bucket=S3_BUCKET, Key=FAMILIES_KEY,
                           Body=json.dumps(data).encode("utf-8"),
                           ContentType="application/json")
        except ClientError:
            pass


def _families_load() -> dict:
    data = None
    if _s3:
        buf = _s3_get_file(FAMILIES_KEY)
        if buf:
            try:
                data = json.load(buf)
            except Exception:
                data = None
    if data is None and os.path.exists(LOCAL_FAMILIES):
        try:
            with open(LOCAL_FAMILIES, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = None
    if data is None:
        data = {"families": []}
    data.setdefault("families", [])
    return data


def _family_next_id(families: list) -> str:
    nums = [int(x["id"][1:]) for x in families
            if str(x.get("id", "")).startswith("f") and str(x["id"])[1:].isdigit()]
    return "f" + str((max(nums) if nums else 0) + 1)


def _norm_header(h) -> str:
    """Lowercase, collapse whitespace, drop a leading '2026 >' prefix."""
    s = re.sub(r"\s+", " ", str(h or "").strip().lower())
    s = re.sub(r"^20\d\d\s*>\s*", "", s)   # strip '2026 >' / '2026 > ' prefixes
    return s.strip()


# Normalized header → canonical field (best-effort import mapping)
_FAMILY_ALIASES = {
    "last":     ["last name", "last", "camper last name", "camper last"],
    "first":    ["first name", "first", "camper first name", "camper first"],
    "bunk":     ["bunk name", "bunk"],
    "family":   ["family last name", "family name", "family"],
    "primary_first":   ["p1 first name", "parent 1 first name", "guardian 1 first name", "primary first name"],
    "primary_last":    ["p1 last name", "parent 1 last name", "guardian 1 last name", "primary last name"],
    "primary_phone":   ["p1 cell phone", "p1 phone", "parent 1 phone", "parent 1 cell phone", "primary phone"],
    "primary_email":   ["p1 email", "p1 email address", "parent 1 email", "guardian 1 email", "primary email", "primary email address"],
    "secondary_first": ["p2 first name", "parent 2 first name", "guardian 2 first name", "secondary first name"],
    "secondary_last":  ["p2 last name", "parent 2 last name", "guardian 2 last name", "secondary last name"],
    "secondary_phone": ["p2 cell phone", "p2 phone", "parent 2 phone", "parent 2 cell phone", "secondary phone"],
    "secondary_email": ["p2 email", "p2 email address", "parent 2 email", "guardian 2 email", "secondary email", "secondary email address"],
    "address":  ["primary family address 1", "primary family address", "address", "home address", "street"],
    "address2": ["primary family address 2", "address 2", "address line 2", "apt", "unit", "suite"],
    "city":     ["primary family city", "city"],
    "state":    ["primary family state", "state"],
    "zip":      ["primary family zip", "zip", "zip code", "postal code"],
    "pu1_name": ["authorized pick-up/emergency contact: 1 - first & last name", "pickup 1 name"],
    "pu1_auth": ["authorized pick-up/emergency contact: 1 - authorization", "pickup 1 authorization"],
    "pu2_name": ["authorized pick-up/emergency contact: 2 - first & last name", "pickup 2 name"],
    "pu2_auth": ["authorized pick-up/emergency contact: 2 - authorization", "pickup 2 authorization"],
    "pu3_name": ["authorized pick-up/emergency contact: 3 - first & last name", "pickup 3 name"],
    "pu3_auth": ["authorized pick-up/emergency contact: 3 - authorization", "pickup 3 authorization"],
    "pu4_name": ["authorized pick-up/emergency contact: 4 - first & last name", "pickup 4 name"],
    "pu4_auth": ["authorized pick-up/emergency contact: 4 - authorization", "pickup 4 authorization"],
}


def _field_slug(header: str) -> str:
    """Stable key for an unrecognized column, e.g. 'Family' -> 'family'."""
    return re.sub(r"[^a-z0-9]+", "_", _norm_header(header)).strip("_")


def _families_from_rows(rows: list) -> list:
    """Map a header row + data rows into family records. Known columns map to
    their canonical field; any UNRECOGNIZED column is auto-captured under a slug
    of its header so new fields are kept (not dropped)."""
    if not rows:
        return []
    hl = [_norm_header(h) for h in rows[0]]
    col_for = {}
    for field, aliases in _FAMILY_ALIASES.items():
        for a in aliases:
            if a in hl:
                col_for[field] = hl.index(a)
                break
    mapped_cols = set(col_for.values())
    # Auto-capture extra columns (skip blank headers like the leading index col)
    extra, used = [], set(FAMILY_FIELDS)
    for ci, h in enumerate(hl):
        if ci in mapped_cols or not h:
            continue
        slug = _field_slug(h)
        if not slug:
            continue
        base, k = slug, 2
        while slug in used:
            slug = f"{base}_{k}"; k += 1
        used.add(slug)
        extra.append((ci, slug))

    def _val(r, ci):
        return str(r[ci]).strip() if (ci is not None and ci < len(r) and r[ci] is not None) else ""

    out = []
    for r in rows[1:]:
        rec = {f: _val(r, col_for.get(f)) for f in FAMILY_FIELDS}
        for ci, slug in extra:
            rec[slug] = _val(r, ci)
        if any(rec.values()):
            out.append(rec)
    return out


def _read_spreadsheet_rows(file_bytes: bytes, filename: str) -> list:
    """Return a list of rows (each a list of cell values); first row = headers."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        return [[c for c in row] for row in ws.iter_rows(values_only=True)]
    # CSV / TSV
    text = file_bytes.decode("utf-8-sig", errors="replace")
    delim = "\t" if (name.endswith(".tsv") or "\t" in text.split("\n", 1)[0]) else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delim)]


# ---------------------------------------------------------------------------
# User accounts (login) — username + hashed password, server-side sessions
# ---------------------------------------------------------------------------

def _users_save(data: dict) -> None:
    try:
        with open(LOCAL_USERS, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass
    if _s3:
        try:
            _s3.put_object(Bucket=S3_BUCKET, Key=USERS_KEY,
                           Body=json.dumps(data).encode("utf-8"),
                           ContentType="application/json")
        except ClientError:
            pass


def _users_load() -> dict:
    data = None
    if _s3:
        buf = _s3_get_file(USERS_KEY)
        if buf:
            try:
                data = json.load(buf)
            except Exception:
                data = None
    if data is None and os.path.exists(LOCAL_USERS):
        try:
            with open(LOCAL_USERS, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = None
    if data is None:
        data = {"users": []}
    data.setdefault("users", [])
    return data


def _find_user(username: str) -> dict | None:
    u = (username or "").strip().lower()
    return next((x for x in _users_load()["users"] if x.get("username", "").lower() == u), None)


def _current_user() -> dict | None:
    uname = session.get("user")
    return _find_user(uname) if uname else None


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if _current_user() is None:
            return jsonify({"error": "auth", "message": "Please sign in."}), 401
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = _current_user()
        if u is None:
            return jsonify({"error": "auth", "message": "Please sign in."}), 401
        if not u.get("is_admin"):
            return jsonify({"error": "forbidden", "message": "Admins only."}), 403
        return f(*args, **kwargs)
    return wrapper


# Paths reachable without a session (the page shell + auth endpoints)
_PUBLIC_PATHS = {"/", "/logo.png", "/health", "/healthz",
                 "/api/me", "/api/login", "/api/register", "/api/logout"}


@app.before_request
def _require_login():
    """Gate every /api/* route behind a valid session, except the auth ones."""
    p = request.path
    if p in _PUBLIC_PATHS or not p.startswith("/api/"):
        return None
    if _current_user() is None:
        return jsonify({"error": "auth", "message": "Please sign in."}), 401
    return None


# In-memory job store  {job_id: {status, progress, result}}
jobs: dict = {}
jobs_lock = threading.Lock()


# ---------------------------------------------------------------------------
# Background job runner
# ---------------------------------------------------------------------------

def run_job(job_id: str, file_bytes: bytes, report_type: str, week_num: int = None) -> None:
    def log(msg: str, level: str = "info") -> None:
        with jobs_lock:
            jobs[job_id]["progress"].append({"msg": msg, "level": level})

    try:
        with jobs_lock:
            jobs[job_id]["status"] = "running"

        log("Loading bunk configuration…")
        config = _s3_load_config() or load_bunk_config(CONFIG_PATH)

        log(f"Processing report type: {report_type}…")
        result = process_report(file_bytes, report_type, config, job_id, OUTPUT_DIR,
                                week_num=week_num, week_dates=_season_week_strings(),
                                families=_families_load()["families"],
                                schedule_overrides=_schedules_load()["overrides"])

        if result["success"]:
            log(result["message"], "ok")
            with jobs_lock:
                jobs[job_id]["status"]   = "done"
                jobs[job_id]["filename"] = result["filename"]
                jobs[job_id]["rows"]     = result.get("rows", 0)
            # Upload to S3 and clean up old files
            try:
                local_path = os.path.join(OUTPUT_DIR, result["filename"])
                _s3_upload(local_path, result["filename"])
                _s3_delete_old(keep=10)
            except Exception:
                pass
            # Keep only 10 most recent local files
            try:
                all_files = sorted(
                    [f for f in os.listdir(OUTPUT_DIR) if f.endswith(".xlsx")],
                    key=lambda f: os.path.getmtime(os.path.join(OUTPUT_DIR, f)),
                    reverse=True
                )
                for old in all_files[10:]:
                    os.remove(os.path.join(OUTPUT_DIR, old))
            except Exception:
                pass
        else:
            log(result["message"], "err")
            with jobs_lock:
                jobs[job_id]["status"] = "error"
                jobs[job_id]["error"]  = result["message"]

    except Exception as exc:
        with jobs_lock:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"]  = str(exc)


# ---------------------------------------------------------------------------
# API routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/logo.png")
def logo():
    path = os.path.join(BASE_DIR, "logo.png")
    if os.path.exists(path):
        return send_file(path, mimetype="image/png")
    return "", 404


# --- Bunk / Camp config ---

@app.route("/api/config", methods=["GET"])
def get_config():
    try:
        # Try S3 first (persists across Render restarts), fall back to local file
        config = _s3_load_config() or load_bunk_config(CONFIG_PATH)
        return jsonify(config)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/config/meta", methods=["GET"])
def get_config_meta():
    """Audit info for the camp/bunk config: who/when last saved + whether
    persistent (S3) storage is active."""
    m = _load_config_meta()
    return jsonify({"saved_at": m.get("saved_at", ""), "saved_by": m.get("saved_by", ""),
                    "persistent": bool(_s3)})


@app.route("/api/config", methods=["POST"])
def save_config():
    try:
        data = request.get_json(force=True)
        save_bunk_config(CONFIG_PATH, data)  # save locally as backup
        _s3_save_config(data)                # save to S3 for persistence
        u = _current_user() or {}
        meta = _save_config_meta(u.get("name") or u.get("username") or "")
        return jsonify({"ok": True, **meta})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _season_summary() -> dict:
    mondays = _season_mondays()
    start = _season_load()["mondays"][0]
    last_fri = (mondays[-1] + timedelta(days=4)) if mondays[-1] else None
    end_str = last_fri.strftime("%B %#d, %Y" if os.name == "nt" else "%B %-d, %Y") if last_fri else ""
    return {"start": start, "end": end_str,
            "weeks": [{"week": i + 1, "range": _week_range_str(m)} for i, m in enumerate(mondays)]}


@app.route("/api/season", methods=["GET"])
def api_season():
    return jsonify(_season_summary())


@app.route("/api/season", methods=["POST"])
def api_season_save():
    body = request.get_json(force=True, silent=True) or {}
    start = (body.get("start") or "").strip()
    try:
        d0 = date.fromisoformat(start)
    except (ValueError, TypeError):
        return jsonify({"error": "Enter a valid camp-start date."}), 400
    # Derive 8 consecutive Mondays from the start date
    clean = [(d0 + timedelta(days=7 * i)).isoformat() for i in range(8)]
    _season_save({"mondays": clean})
    return jsonify({"ok": True, **_season_summary()})


@app.route("/api/schedules", methods=["GET"])
def api_schedules():
    """Camper list (from the saved master) + per-week day overrides, for the editor."""
    overrides = _schedules_load()["overrides"]
    weeks = [{"n": i + 1, "range": r} for i, r in enumerate(_season_week_strings())]
    fb = _load_master()
    campers = []
    if fb:
        try:
            for r in parse_master(fb) or []:
                campers.append({
                    "key":   _camper_key(r.get("name"), r.get("bunk")),
                    "name":  r.get("name", ""),
                    "bunk":  r.get("bunk", ""),
                    "weeks": r.get("weeks", []),
                    "days":  r.get("days_sched", "MTWRF"),
                })
        except Exception:
            campers = []
        # de-dupe by key (master may repeat), keep first
        seen, uniq = set(), []
        for c in campers:
            if c["key"] in seen:
                continue
            seen.add(c["key"]); uniq.append(c)
        uniq.sort(key=lambda c: c["name"].lower())
        campers = uniq
    return jsonify({"campers": campers, "overrides": overrides, "weeks": weeks,
                    "has_master": bool(fb)})


@app.route("/api/schedules", methods=["POST"])
def api_schedules_save():
    body = request.get_json(force=True, silent=True) or {}
    key = (body.get("key") or "").strip()
    if not key:
        return jsonify({"error": "missing key"}), 400
    data = _schedules_load()
    # Replace the camper's entire week map in one shot (used by the Save button)
    if isinstance(body.get("replace"), dict):
        clean = {}
        for w, dd in body["replace"].items():
            try:
                n = int(w)
            except (TypeError, ValueError):
                continue
            if 1 <= n <= 8:
                clean[str(n)] = _canon_days(dd)   # "" is valid (attends no days that week)
        if clean:
            data["overrides"][key] = clean
        else:
            data["overrides"].pop(key, None)
        _schedules_save(data)
        return jsonify({"ok": True})
    # Single-week set/clear
    try:
        wk = int(body.get("week", 0))
    except (TypeError, ValueError):
        wk = 0
    if not (1 <= wk <= 8):
        return jsonify({"error": "missing week"}), 400
    ov = data["overrides"].setdefault(key, {})
    if body.get("clear"):
        ov.pop(str(wk), None)
    else:
        ov[str(wk)] = _canon_days(body.get("days"))
    if not ov:
        data["overrides"].pop(key, None)
    _schedules_save(data)
    return jsonify({"ok": True})


@app.route("/api/pricing", methods=["GET"])
@admin_required
def api_pricing():
    """Editable pricing config (camp tuition tiers, transport, childcare)."""
    return jsonify(_pricing_load())


@app.route("/api/pricing", methods=["POST"])
@admin_required
def api_pricing_save():
    """Replace the pricing config with the posted body (validated-filled)."""
    body = request.get_json(force=True, silent=True) or {}
    if not isinstance(body, dict):
        return jsonify({"error": "bad payload"}), 400
    _pricing_save(_deep_fill(body, _DEFAULT_PRICING))
    return jsonify({"ok": True})


@app.route("/api/bunk-snapshot", methods=["GET"])
def api_bunk_snapshot():
    """On-screen Bunk Snapshot (Totals + Bunks) computed from the saved master."""
    meta = _load_master_meta() or {}
    fb = _load_master()
    if not fb:
        return jsonify({"has_master": False, "meta": meta})
    try:
        campers = parse_master(fb) or []
        config = _s3_load_config() or load_bunk_config(CONFIG_PATH)
        data = bunk_snapshot_data(campers, config)
    except Exception as e:
        return jsonify({"has_master": True, "error": str(e), "meta": meta}), 200
    cur_week, cur_day = _current_week_day()
    return jsonify({
        "has_master": True,
        "meta":       {"filename": meta.get("filename", ""),
                       "uploaded_at": meta.get("uploaded_at", ""),
                       "uploaded_by": meta.get("uploaded_by", "")},
        "report":     data["report"],
        "totals":     data["totals"],
        "current_week": cur_week,   # 1-8 or null
        "current_day":  cur_day,    # 0=Mon..4=Fri or null
    })


@app.route("/api/families/full", methods=["GET"])
def api_families_full():
    """Grouped family records joined with master (age/grade/enrollment) and
    per-week schedule overrides, for the on-screen Families directory."""
    fams = _families_load()["families"]
    week_ranges = _season_week_strings()
    weeks = [{"n": i + 1, "range": r} for i, r in enumerate(week_ranges)]
    overrides = _schedules_load()["overrides"]

    # Index the saved master by camper key and by name (for enrichment)
    master_by_key, master_by_name = {}, {}
    fb = _load_master()
    if fb:
        try:
            for c in parse_master(fb) or []:
                master_by_key.setdefault(_camper_key(c.get("name"), c.get("bunk")), c)
                master_by_name.setdefault((c.get("name") or "").strip().lower(), c)
        except Exception:
            pass

    def _camper_schedule(name, bunk):
        c = master_by_key.get(_camper_key(name, bunk)) or master_by_name.get((name or "").strip().lower())
        if not c:
            return {"found": False, "age": "", "grade": "", "bunk": bunk, "weeks_detail": [], "sched_key": ""}
        skey = _camper_key(c.get("name"), c.get("bunk"))
        ov = overrides.get(skey, {})
        default_days = c.get("days_sched") or "MTWRF"
        detail = []
        for i, enrolled in enumerate(c.get("weeks", [])):
            if not enrolled:
                continue
            days = ov.get(str(i + 1), default_days)
            detail.append({"n": i + 1,
                           "range": week_ranges[i] if i < len(week_ranges) else "",
                           "days": _canon_days(days),
                           "default": default_days == _canon_days(days) and str(i + 1) not in ov})
        return {"found": True, "age": c.get("age") or "", "grade": c.get("grade") or "",
                "bunk": c.get("bunk") or bunk, "weeks_detail": detail, "sched_key": skey}

    def _full(*parts):
        return " ".join(p.strip() for p in parts if (p or "").strip()).strip()

    # Emails aren't a mapped field, but the importer auto-captures unknown columns
    # under a slug — so pull any email out of the record by key hint or value shape.
    email_re = re.compile(r"[^@\s,;]+@[^@\s,;]+\.[^@\s,;]+")

    def _emails_from_record(rec):
        primary, secondary, others = "", "", []
        for k, v in rec.items():
            if not isinstance(v, str) or not v.strip():
                continue
            kl = k.lower()
            m = email_re.search(v)
            if not ("email" in kl or "e_mail" in kl or kl.endswith("_mail") or m):
                continue
            val = m.group(0) if m else v.strip()
            if any(t in kl for t in ("p1", "parent_1", "parent1", "primary", "guardian_1", "guardian1", "mother", "mom")):
                primary = primary or val
            elif any(t in kl for t in ("p2", "parent_2", "parent2", "secondary", "guardian_2", "guardian2", "father", "dad")):
                secondary = secondary or val
            elif val not in (primary, secondary) and val not in others:
                others.append(val)
        return primary, secondary, others

    groups = {}
    for r in fams:
        fam = (r.get("family") or r.get("last") or "").strip()
        gkey = "|".join([fam.lower(), (r.get("address") or "").strip().lower(),
                         (r.get("zip") or "").strip().lower(),
                         _full(r.get("primary_first"), r.get("primary_last")).lower()])
        g = groups.get(gkey)
        if not g:
            pe, se, oe = _emails_from_record(r)
            # Canonical, editable contact values (prefer the mapped email field,
            # fall back to a detected one so existing imports are editable too).
            g = groups[gkey] = {
                "key": gkey,
                "name": fam or (r.get("last") or "Family"),
                "ids": [],
                "address": {"address": r.get("address", ""), "address2": r.get("address2", ""),
                            "city": r.get("city", ""), "state": r.get("state", ""), "zip": r.get("zip", "")},
                "contacts": {
                    "primary":   {"name": _full(r.get("primary_first"), r.get("primary_last")),
                                  "phone": r.get("primary_phone", ""), "email": (r.get("primary_email") or pe)},
                    "secondary": {"name": _full(r.get("secondary_first"), r.get("secondary_last")),
                                  "phone": r.get("secondary_phone", ""), "email": (r.get("secondary_email") or se)},
                    "emails_other": oe,
                    "pickups":   [{"name": r.get(f"pu{i}_name", ""), "auth": r.get(f"pu{i}_auth", "")}
                                  for i in range(1, 5) if (r.get(f"pu{i}_name") or "").strip()],
                },
                "fields": {
                    "address": r.get("address", ""), "address2": r.get("address2", ""),
                    "city": r.get("city", ""), "state": r.get("state", ""), "zip": r.get("zip", ""),
                    "primary_first": r.get("primary_first", ""), "primary_last": r.get("primary_last", ""),
                    "primary_phone": r.get("primary_phone", ""), "primary_email": (r.get("primary_email") or pe),
                    "secondary_first": r.get("secondary_first", ""), "secondary_last": r.get("secondary_last", ""),
                    "secondary_phone": r.get("secondary_phone", ""), "secondary_email": (r.get("secondary_email") or se),
                    "pu1_name": r.get("pu1_name", ""), "pu1_auth": r.get("pu1_auth", ""),
                    "pu2_name": r.get("pu2_name", ""), "pu2_auth": r.get("pu2_auth", ""),
                    "pu3_name": r.get("pu3_name", ""), "pu3_auth": r.get("pu3_auth", ""),
                    "pu4_name": r.get("pu4_name", ""), "pu4_auth": r.get("pu4_auth", ""),
                },
                "campers": [],
                "_search": set(),
            }
        if r.get("id"):
            g["ids"].append(r.get("id"))
        last, first = (r.get("last") or "").strip(), (r.get("first") or "").strip()
        disp = _full(f"{last}," if last else "", first) or last or first
        sched = _camper_schedule(f"{last}, {first}", r.get("bunk", ""))
        g["campers"].append({
            "name": disp, "first": first, "last": last, "id": r.get("id", ""),
            "bunk": sched["bunk"] or r.get("bunk", ""),
            "age": sched["age"], "grade": sched["grade"],
            "weeks_detail": sched["weeks_detail"], "in_master": sched["found"],
            "sched_key": sched["sched_key"],
        })
        for t in (disp, first, last):
            if t:
                g["_search"].add(t.lower())

    out = []
    for g in groups.values():
        s = g.pop("_search")
        s.add((g["name"] or "").lower())
        s.add(g["contacts"]["primary"]["name"].lower())
        s.add(g["contacts"]["secondary"]["name"].lower())
        g["search"] = " ".join(x for x in s if x)
        out.append(g)
    out.sort(key=lambda g: (g["name"] or "").lower())
    return jsonify({"families": out, "weeks": weeks,
                    "has_families": bool(fams), "has_master": bool(fb)})


# --- Report processing ---

@app.route("/api/master", methods=["GET"])
def api_master():
    """Report whether a master sheet is currently saved (for the UI)."""
    cur_week, cur_day = _current_week_day()   # for the Camp Snapshot column highlight
    meta = _load_master_meta()
    if meta and _load_master() is not None:
        return jsonify({"loaded": True, "current_week": cur_week, "current_day": cur_day, **meta})
    return jsonify({"loaded": False, "current_week": cur_week, "current_day": cur_day})


@app.route("/api/master", methods=["POST"])
def api_master_save():
    """Upload + save a master sheet directly (from the Utilities tab)."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "No file uploaded."}), 400
    file_bytes = f.read()
    if not is_master(file_bytes):
        return jsonify({"error": "That file doesn't look like a master sheet. "
                                 "Check that it has the expected camper columns."}), 400
    u = _current_user() or {}
    meta = _save_master(file_bytes, f.filename, uploaded_by=u.get("name") or u.get("username") or "")
    return jsonify({"loaded": True, **meta})


@app.route("/api/master/download")
def api_master_download():
    """Download the currently-saved master sheet (to edit and re-upload)."""
    data = _load_master()
    if data is None:
        return jsonify({"error": "No saved master sheet."}), 404
    meta = _load_master_meta() or {}
    fname = meta.get("filename") or "master.csv"
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname,
                     mimetype=_mime_for(fname))


@app.route("/api/master", methods=["DELETE"])
def api_master_clear():
    """Forget the saved master sheet."""
    for p in (LOCAL_MASTER, LOCAL_MASTER_META):
        try:
            os.remove(p)
        except OSError:
            pass
    if _s3:
        for key in (MASTER_KEY, MASTER_META_KEY):
            try:
                _s3.delete_object(Bucket=S3_BUCKET, Key=key)
            except ClientError:
                pass
    return jsonify({"loaded": False})


@app.route("/api/payroll", methods=["GET"])
def api_payroll():
    data = _payroll_load()
    return jsonify({"staff": data["staff"], "checks": data["checks"],
                    "days": _payroll_days(), "locked": data.get("locked", False)})


@app.route("/api/payroll/lock", methods=["POST"])
def api_payroll_lock():
    body = request.get_json(force=True, silent=True) or {}
    data = _payroll_load()
    data["locked"] = bool(body.get("locked"))
    _payroll_save(data)
    return jsonify({"locked": data["locked"]})


@app.route("/api/payroll/export")
def api_payroll_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    view   = request.args.get("view", "weeks")
    sort   = request.args.get("sort", "last")
    extp   = request.args.get("extp", "ALL")
    sel_areas = [a for a in (request.args.get("areas", "") or "").split(",") if a]
    q = (request.args.get("q", "") or "").strip().lower()
    try:
        period = int(request.args.get("period", "0"))
    except (TypeError, ValueError):
        period = 0

    data = _payroll_load()
    checks = data["checks"]
    days_all = _payroll_days()

    def _q_ok(s):
        if not q:
            return True
        hay = " ".join(str(s.get(k, "")) for k in ("last", "first", "area", "bunk", "title")).lower()
        return q in hay
    staff = [s for s in data["staff"]
             if (not sel_areas or s.get("area") in sel_areas) and _q_ok(s)]

    SYM = {"check": "✓", "x": "✗", "half": "½", "na": "N/A", True: "✓"}
    def namekey(s): return (s.get("last", "") + s.get("first", "")).lower()
    def cnt(sid, days):
        c = checks.get(sid, {})
        total = 0.0
        for d in days:
            v = c.get(d["iso"])
            if v in ("check", True):
                total += 1
            elif v == "half":
                total += 0.5
        return int(total) if total == int(total) else total
    def area_txt(s):
        return s.get("title") if (s.get("area") == "Support" and s.get("title")) else s.get("area", "")

    wb = Workbook(); ws = wb.active; ws.title = "Payroll"
    HDR = Font(bold=True, color="FFFFFF"); FILL = PatternFill("solid", fgColor="6D1F2F")
    CTR = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")
    _t = Side(style="thin", color="CCCCCC"); BORD = Border(left=_t, right=_t, top=_t, bottom=_t)

    def header(cols):
        ws.append(cols)
        for c in ws[ws.max_row]:
            c.font = HDR; c.fill = FILL; c.alignment = CTR; c.border = BORD

    if view == "totals":
        staff.sort(key=(lambda s: (-cnt(s["id"], days_all), namekey(s))) if sort == "total"
                   else (lambda s: (s.get("area", "").lower(), namekey(s))) if sort == "area"
                   else namekey)
        header(["Staff", "Area", "Total Days (all 8 weeks)"])
        for s in staff:
            jc = " (JC)" if "junior" in (s.get("title", "").lower()) else ""
            a = area_txt(s) + (f" — {s['bunk']}" if s.get("bunk") else "")
            ws.append([f"{s.get('last','')}, {s.get('first','')}{jc}", a, cnt(s["id"], days_all)])
            for c in ws[ws.max_row]:
                c.border = BORD; c.alignment = CTR if c.column == 3 else LEFT
        widths = {"A": 30, "B": 22, "C": 14}; fname = "Payroll_Totals.xlsx"
    elif view == "ext":
        def _shift_ok(e):
            e = e or ""
            if extp == "AM": return "AM" in e.upper()
            if extp == "PM": return "PM" in e.upper()
            return bool(e)
        staff = [s for s in staff if s.get("ext") and _shift_ok(s.get("ext"))]
        staff.sort(key=namekey)
        header(["Staff", "MON", "TUES", "WED", "THURS", "FRI"])
        for s in staff:
            ws.append([f"{s.get('last','')}, {s.get('first','')}", "", "", "", "", ""])
            ws.row_dimensions[ws.max_row].height = 26
            for c in ws[ws.max_row]:
                c.border = BORD; c.alignment = LEFT if c.column == 1 else CTR
        widths = {"A": 30, "B": 11, "C": 11, "D": 11, "E": 11, "F": 11}
        _sfx = {"AM": "_AM", "PM": "_PM"}.get(extp, "")
        fname = f"Extended_Staff{_sfx}.xlsx"
    elif view == "holiday":
        # All staff with BS / SP\\MTC plus the holiday-week days (Th 7/2, Mon 7/6, Fri 7/3)
        staff_h = sorted(data["staff"], key=namekey)
        hol = [d for md in ("7/2", "7/6", "7/3") for d in days_all if d["md"] == md]
        header(["Staff", "Area", "BS", "SP\\MTC"] + [f"{d['dow']} {d['md']}" for d in hol])
        for s in staff_h:
            c = checks.get(s["id"], {})
            row = [f"{s.get('last','')}, {s.get('first','')}",
                   area_txt(s) + (f" / {s['bunk']}" if s.get("bunk") else ""),
                   SYM.get(c.get("xtra:0:1"), ""), SYM.get(c.get("xtra:0:2"), "")]
            row += [SYM.get(c.get(d["iso"]), "") for d in hol]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = BORD
                cell.alignment = LEFT if cell.column <= 2 else CTR
        widths = {"A": 26, "B": 18, "C": 8, "D": 10}
        fname = "Payroll_Holiday.xlsx"
    else:  # weeks
        days = days_all[period * 10:period * 10 + 10]
        staff.sort(key=(lambda s: (-cnt(s["id"], days), namekey(s))) if sort == "total"
                   else (lambda s: (s.get("area", "").lower(), namekey(s))) if sort == "area"
                   else namekey)
        cols = ["#", "Staff", "Area"] + [f"{d['dow']} {d['md']}" for d in days]
        if period == 0:
            cols += ["BS", "SP\\MTC"]
        header(cols)
        for s in staff:
            c = checks.get(s["id"], {})
            row = [cnt(s["id"], days), f"{s.get('last','')}, {s.get('first','')}",
                   area_txt(s) + (f" / {s['bunk']}" if s.get("bunk") else "")]
            row += [SYM.get(c.get(d["iso"]), "") for d in days]
            if period == 0:
                row += [SYM.get(c.get(f"xtra:0:{cc}"), "") for cc in (1, 2)]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = BORD; cell.alignment = LEFT if cell.column == 2 else CTR
        widths = {"A": 5, "B": 26, "C": 18}; fname = f"Payroll_Weeks_{period*2+1}_{period*2+2}.xlsx"

    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = "landscape"
    ws.print_options.horizontalCentered = True

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname, mimetype=_XLSX_MIME)


@app.route("/api/payroll/check", methods=["POST"])
def api_payroll_check():
    body = request.get_json(force=True, silent=True) or {}
    sid = str(body.get("id", "")); dt = str(body.get("date", "")); val = body.get("value")
    if not sid or not dt:
        return jsonify({"error": "missing id/date"}), 400
    data = _payroll_load()
    if data.get("locked"):
        return jsonify({"error": "locked"}), 403
    checks = data["checks"].setdefault(sid, {})
    if val in ("check", "x", "half", "na"):
        checks[dt] = val
    else:
        checks.pop(dt, None)   # blank / cleared
    _payroll_save(data)
    return jsonify({"ok": True})


def _timecard_iso(v):
    """Parse a time-card Date cell into an iso date string, or None."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v or "").strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


@app.route("/api/payroll/import-timecard", methods=["POST"])
def api_payroll_import_timecard():
    """Import clock-in rows (Last, First, Date) and mark ✓ for matched staff."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "No file uploaded."}), 400
    data = _payroll_load()
    if data.get("locked"):
        return jsonify({"error": "locked"}), 403
    try:
        rows = _read_spreadsheet_rows(f.read(), f.filename)
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400
    if not rows or len(rows) < 2:
        return jsonify({"error": "No rows found in the file."}), 400

    # Locate Last / First / Date columns by header (fallback to B / C / D)
    hl = [str(h or "").strip().lower() for h in rows[0]]
    def _col(keys, fb):
        for i, h in enumerate(hl):
            if any(k in h for k in keys):
                return i
        return fb
    li, fi, di = _col(["last"], 1), _col(["first"], 2), _col(["date"], 3)

    valid_days = {d["iso"] for d in _payroll_days()}

    # Build name matchers from the staff roster
    by_lf, by_last, by_li = {}, {}, {}
    for s in data["staff"]:
        ln = (s.get("last") or "").strip().lower()
        fn = (s.get("first") or "").strip().lower()
        by_lf[(ln, fn)] = s["id"]
        by_last.setdefault(ln, []).append(s["id"])
        if fn:
            by_li.setdefault((ln, fn[0]), []).append(s["id"])

    def _match(ln, fn):
        ln, fn = ln.strip().lower(), fn.strip().lower()
        if (ln, fn) in by_lf:
            return by_lf[(ln, fn)]                       # exact last + first
        if len(by_last.get(ln, [])) == 1:
            return by_last[ln][0]                        # only one staffer with that last name
        cand = by_li.get((ln, fn[:1]), [])
        if len(cand) == 1:
            return cand[0]                               # unique last + first-initial (nicknames)
        return None

    cells, dates, unmatched = set(), set(), set()
    for r in rows[1:]:
        ln = str(r[li]).strip() if li < len(r) and r[li] is not None else ""
        fn = str(r[fi]).strip() if fi < len(r) and r[fi] is not None else ""
        if not ln and not fn:
            continue
        iso = _timecard_iso(r[di] if di < len(r) else None)
        if not iso or iso not in valid_days:
            continue                                     # date outside the season → skip
        sid = _match(ln, fn)
        if sid is None:
            unmatched.add(f"{ln}, {fn}".strip(", "))
            continue
        data["checks"].setdefault(sid, {})[iso] = "check"
        cells.add((sid, iso)); dates.add(iso)
    _payroll_save(data)
    return jsonify({"ok": True, "checks_set": len(cells),
                    "staff_matched": len({c[0] for c in cells}),
                    "dates": sorted(dates), "unmatched": sorted(unmatched)})


@app.route("/api/payroll/clearday", methods=["POST"])
def api_payroll_clearday():
    """Remove every staff member's mark for one date."""
    body = request.get_json(force=True, silent=True) or {}
    dt = str(body.get("date", ""))
    if not dt:
        return jsonify({"error": "missing date"}), 400
    data = _payroll_load()
    if data.get("locked"):
        return jsonify({"error": "locked"}), 403
    for c in data["checks"].values():
        if isinstance(c, dict):
            c.pop(dt, None)
    _payroll_save(data)
    return jsonify({"ok": True})


@app.route("/api/payroll/staff", methods=["POST"])
def api_payroll_add():
    body = request.get_json(force=True, silent=True) or {}
    last = (body.get("last") or "").strip()
    first = (body.get("first") or "").strip()
    area = (body.get("area") or "").strip()
    if not last and not first:
        return jsonify({"error": "Name required."}), 400
    data = _payroll_load()
    if data.get("locked"):
        return jsonify({"error": "locked"}), 403
    nums = [int(s["id"][1:]) for s in data["staff"]
            if str(s.get("id", "")).startswith("s") and str(s["id"])[1:].isdigit()]
    entry = {"id": "s" + str((max(nums) if nums else 0) + 1),
             "last": last, "first": first, "area": area}
    data["staff"].append(entry)
    _payroll_save(data)
    return jsonify(entry)


@app.route("/api/payroll/staff/<sid>", methods=["PATCH"])
def api_payroll_edit(sid):
    body = request.get_json(force=True, silent=True) or {}
    data = _payroll_load()
    if data.get("locked"):
        return jsonify({"error": "locked"}), 403
    s = next((x for x in data["staff"] if x.get("id") == sid), None)
    if s is None:
        return jsonify({"error": "not found"}), 404
    if "area" in body:
        s["area"] = (body.get("area") or "").strip()
    if "last" in body:
        s["last"] = (body.get("last") or "").strip()
    if "first" in body:
        s["first"] = (body.get("first") or "").strip()
    _payroll_save(data)
    return jsonify(s)


@app.route("/api/payroll/staff/<sid>", methods=["DELETE"])
def api_payroll_del(sid):
    data = _payroll_load()
    if data.get("locked"):
        return jsonify({"error": "locked"}), 403
    data["staff"] = [s for s in data["staff"] if s.get("id") != sid]
    data["checks"].pop(sid, None)
    _payroll_save(data)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.route("/api/me", methods=["GET"])
def api_me():
    u = _current_user()
    any_users = bool(_users_load()["users"])
    if u is None:
        return jsonify({"authenticated": False, "has_users": any_users})
    return jsonify({"authenticated": True, "username": u["username"],
                    "name": u.get("name") or u["username"], "is_admin": bool(u.get("is_admin"))})


@app.route("/api/login", methods=["POST"])
def api_login():
    body = request.get_json(force=True, silent=True) or {}
    u = _find_user(body.get("username", ""))
    if u is None or not check_password_hash(u.get("pw_hash", ""), body.get("password", "")):
        return jsonify({"error": "Invalid username or password."}), 401
    session.permanent = True
    session["user"] = u["username"]
    return jsonify({"username": u["username"], "name": u.get("name") or u["username"],
                    "is_admin": bool(u.get("is_admin"))})


@app.route("/api/register", methods=["POST"])
def api_register():
    body = request.get_json(force=True, silent=True) or {}
    if (body.get("code") or "").strip() != ACCESS_CODE:
        return jsonify({"error": "Incorrect access code."}), 403
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    if not username or not password:
        return jsonify({"error": "Username and password are required."}), 400
    if len(password) < 4:
        return jsonify({"error": "Password must be at least 4 characters."}), 400
    data = _users_load()
    if any(x.get("username", "").lower() == username.lower() for x in data["users"]):
        return jsonify({"error": "That username is taken."}), 409
    entry = {"username": username, "name": username,
             "pw_hash": generate_password_hash(password),
             "is_admin": len(data["users"]) == 0}   # first account is the admin
    data["users"].append(entry)
    _users_save(data)
    session.permanent = True
    session["user"] = username
    return jsonify({"username": username, "name": entry["name"], "is_admin": entry["is_admin"]})


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/account/password", methods=["POST"])
def api_account_password():
    """Let the signed-in user change their own password."""
    u = _current_user()
    if u is None:
        return jsonify({"error": "auth"}), 401
    body = request.get_json(force=True, silent=True) or {}
    current = body.get("current") or ""
    new = body.get("new") or ""
    if not check_password_hash(u.get("pw_hash", ""), current):
        return jsonify({"error": "Current password is incorrect."}), 403
    if len(new) < 4:
        return jsonify({"error": "New password must be at least 4 characters."}), 400
    data = _users_load()
    rec = next((x for x in data["users"] if x.get("username", "").lower() == u["username"].lower()), None)
    if rec is None:
        return jsonify({"error": "not found"}), 404
    rec["pw_hash"] = generate_password_hash(new)
    _users_save(data)
    return jsonify({"ok": True})


@app.route("/api/users", methods=["GET"])
@admin_required
def api_users():
    users = [{"username": x["username"], "name": x.get("name") or x["username"],
              "is_admin": bool(x.get("is_admin"))} for x in _users_load()["users"]]
    return jsonify({"users": users})


@app.route("/api/users", methods=["POST"])
@admin_required
def api_users_add():
    body = request.get_json(force=True, silent=True) or {}
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    if not username or not password:
        return jsonify({"error": "Username and password are required."}), 400
    if len(password) < 4:
        return jsonify({"error": "Password must be at least 4 characters."}), 400
    data = _users_load()
    if any(x.get("username", "").lower() == username.lower() for x in data["users"]):
        return jsonify({"error": "That username is taken."}), 409
    entry = {"username": username, "name": username,
             "pw_hash": generate_password_hash(password),
             "is_admin": bool(body.get("is_admin"))}
    data["users"].append(entry)
    _users_save(data)
    return jsonify({"username": username, "name": entry["name"], "is_admin": entry["is_admin"]})


@app.route("/api/users/<username>", methods=["PATCH"])
@admin_required
def api_users_edit(username):
    body = request.get_json(force=True, silent=True) or {}
    data = _users_load()
    u = next((x for x in data["users"] if x.get("username", "").lower() == username.lower()), None)
    if u is None:
        return jsonify({"error": "not found"}), 404
    new_u = (body.get("username") or "").strip()
    if new_u and new_u != u["username"]:
        if new_u.lower() != u["username"].lower() and \
           any(x.get("username", "").lower() == new_u.lower() for x in data["users"]):
            return jsonify({"error": "That username is taken."}), 409
        old = u["username"]
        if u.get("name", "") == old:   # name mirrors username in our model
            u["name"] = new_u
        u["username"] = new_u
        if session.get("user", "") == old:
            session["user"] = new_u
    if "name" in body:
        u["name"] = (body.get("name") or "").strip() or u["username"]
    if body.get("password"):
        if len(body["password"]) < 4:
            return jsonify({"error": "Password must be at least 4 characters."}), 400
        u["pw_hash"] = generate_password_hash(body["password"])
    if "is_admin" in body:
        me = _current_user()
        # Don't let an admin strip their own admin rights (avoid lockout)
        if not (me and me["username"].lower() == username.lower() and not body["is_admin"]):
            u["is_admin"] = bool(body["is_admin"])
    _users_save(data)
    return jsonify({"username": u["username"], "name": u.get("name"), "is_admin": bool(u.get("is_admin"))})


@app.route("/api/users/<username>", methods=["DELETE"])
@admin_required
def api_users_del(username):
    me = _current_user()
    if me and me["username"].lower() == username.lower():
        return jsonify({"error": "You can't delete your own account."}), 400
    data = _users_load()
    data["users"] = [x for x in data["users"] if x.get("username", "").lower() != username.lower()]
    _users_save(data)
    return jsonify({"ok": True})


@app.route("/api/families", methods=["GET"])
def api_families():
    return jsonify(_families_load())


@app.route("/api/families", methods=["POST"])
def api_families_add():
    body = request.get_json(force=True, silent=True) or {}
    data = _families_load()
    entry = {"id": _family_next_id(data["families"])}
    for f in FAMILY_FIELDS:
        entry[f] = (body.get(f) or "").strip()
    if not any(entry[f] for f in FAMILY_FIELDS):
        return jsonify({"error": "Enter at least one field."}), 400
    data["families"].append(entry)
    _families_save(data)
    return jsonify(entry)


@app.route("/api/families/<fid>", methods=["PATCH"])
def api_families_edit(fid):
    body = request.get_json(force=True, silent=True) or {}
    data = _families_load()
    fam = next((x for x in data["families"] if x.get("id") == fid), None)
    if fam is None:
        return jsonify({"error": "not found"}), 404
    for f in FAMILY_FIELDS:
        if f in body:
            fam[f] = (body.get(f) or "").strip()
    _families_save(data)
    return jsonify(fam)


@app.route("/api/families/<fid>", methods=["DELETE"])
def api_families_del(fid):
    data = _families_load()
    data["families"] = [x for x in data["families"] if x.get("id") != fid]
    _families_save(data)
    return jsonify({"ok": True})


@app.route("/api/families/import", methods=["POST"])
def api_families_import():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "No file uploaded."}), 400
    mode = request.form.get("mode", "replace")   # 'replace' or 'append'
    try:
        rows = _read_spreadsheet_rows(f.read(), f.filename)
        parsed = _families_from_rows(rows)
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400
    if not parsed:
        return jsonify({"error": "No family rows found. Check that the sheet has a header row."}), 400
    data = _families_load()
    existing = [] if mode == "replace" else data["families"]
    out = list(existing)
    for rec in parsed:
        rec = {"id": _family_next_id(out), **rec}
        out.append(rec)
    data["families"] = out
    u = _current_user() or {}
    data["uploaded_at"] = _now_eastern_stamp()
    data["uploaded_by"] = u.get("name") or u.get("username") or ""
    data["filename"] = f.filename
    _families_save(data)
    return jsonify({"ok": True, "count": len(parsed), "total": len(out), "mode": mode})


@app.route("/api/process", methods=["POST"])
def api_process():
    excel_file  = request.files.get("excel_file")
    report_type = request.form.get("report_type", "").strip()

    if not report_type:
        return jsonify({"error": "No report type selected."}), 400

    # A new upload is used directly; if it's a master, save it for reuse.
    # With no upload, fall back to the previously saved master sheet.
    if excel_file and excel_file.filename:
        file_bytes = excel_file.read()
        if is_master(file_bytes):
            _save_master(file_bytes, excel_file.filename)
    else:
        file_bytes = _load_master()
        if not file_bytes:
            return jsonify({"error": "No file uploaded and no saved master sheet found. "
                                     "Upload a master sheet first."}), 400

    job_id     = uuid.uuid4().hex[:8]

    # Week-specific reports: Driver Totals highlights the week; the others
    # filter campers to those enrolled in the selected week.
    week_num = None
    if report_type in WEEK_AWARE_REPORTS:
        try:
            week_num = int(request.form.get("week_num", 0))
            if week_num < 1 or week_num > 8:
                week_num = None
        except (TypeError, ValueError):
            week_num = None

    with jobs_lock:
        jobs[job_id] = {"status": "queued", "progress": []}

    thread = threading.Thread(target=run_job,
                              args=(job_id, file_bytes, report_type),
                              kwargs={"week_num": week_num},
                              daemon=True)
    thread.start()

    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
def api_status(job_id: str):
    with jobs_lock:
        job = jobs.get(job_id)
    if job is None:
        return jsonify({"error": "Job not found."}), 404
    return jsonify(job)


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

def _mime_for(filename: str) -> str:
    fn = filename.lower()
    if fn.endswith(".docx"):
        return _DOCX_MIME
    if fn.endswith(".zip"):
        return "application/zip"
    if fn.endswith(".csv"):
        return "text/csv"
    if fn.endswith(".xls"):
        return "application/vnd.ms-excel"
    return _XLSX_MIME


@app.route("/api/download/<job_id>")
def api_download(job_id: str):
    with jobs_lock:
        job = jobs.get(job_id)
    if job is None or job.get("status") != "done":
        return jsonify({"error": "File not ready."}), 404
    filename = job["filename"]
    buf = _s3_get_file(filename)
    if buf:
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype=_mime_for(filename))
    path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(path):
        return jsonify({"error": "Output file missing."}), 500
    return send_file(path, as_attachment=True, download_name=filename,
                     mimetype=_mime_for(filename))


@app.route("/api/files/<path:filename>")
def api_download_file(filename: str):
    safe = os.path.basename(filename)
    buf = _s3_get_file(safe)
    if buf:
        return send_file(buf, as_attachment=True, download_name=safe,
                         mimetype=_mime_for(safe))
    path = os.path.join(OUTPUT_DIR, safe)
    if not os.path.exists(path):
        return jsonify({"error": "File not found."}), 404
    return send_file(path, as_attachment=True, download_name=safe,
                     mimetype=_mime_for(safe))


@app.route("/api/recent")
def api_recent():
    try:
        if _s3:
            objects = _s3_list_recent(10)
            return jsonify([{
                "name":  o["Key"],
                "mtime": o["LastModified"].timestamp(),
                "url":   f"/api/files/{o['Key']}",
            } for o in objects])
        files = []
        for f in os.listdir(OUTPUT_DIR):
            if f.endswith(".xlsx"):
                fpath = os.path.join(OUTPUT_DIR, f)
                files.append({"name": f, "mtime": os.path.getmtime(fpath)})
        files.sort(key=lambda x: x["mtime"], reverse=True)
        return jsonify([{
            "name":  f["name"],
            "mtime": f["mtime"],
            "url":   f"/api/files/{f['name']}",
        } for f in files[:10]])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/weather")
def api_weather():
    """5-day forecast for Warrington, PA via Open-Meteo (no API key required)."""
    try:
        url = (
            "https://api.open-meteo.com/v1/forecast"
            "?latitude=40.2479&longitude=-75.1330"
            "&daily=temperature_2m_max,temperature_2m_min,weathercode,"
            "precipitation_probability_max,windspeed_10m_max"
            "&current_weather=true"
            "&temperature_unit=fahrenheit&windspeed_unit=mph"
            "&timezone=America%2FNew_York"
            "&forecast_days=5"
        )
        with urllib.request.urlopen(url, timeout=5) as resp:
            data = json.loads(resp.read())
        daily = data["daily"]
        def _num(arr, i):
            try:
                return round(arr[i]) if arr[i] is not None else None
            except (KeyError, IndexError, TypeError):
                return None
        days = []
        for i in range(5):
            days.append({
                "date":    daily["time"][i],
                "high":    round(daily["temperature_2m_max"][i]),
                "low":     round(daily["temperature_2m_min"][i]),
                "code":    daily["weathercode"][i],
                "pop":     _num(daily.get("precipitation_probability_max", []), i),
                "wind":    _num(daily.get("windspeed_10m_max", []), i),
            })
        cw = data.get("current_weather") or {}
        current = {"temp": round(cw["temperature"]), "code": cw.get("weathercode")} if cw.get("temperature") is not None else None
        return jsonify({"days": days, "current": current})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/health")
@app.route("/healthz")
def health():
    return "OK"


# ---------------------------------------------------------------------------
# Embedded HTML / CSS / JS
# ---------------------------------------------------------------------------

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<link rel="icon" type="image/png" href="/logo.png">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Elbow Lane — Reporting Center</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Roboto+Slab:wght@600;700;800&family=DM+Sans:ital,wght@0,300;0,400;0,500;0,600;1,400&display=swap" rel="stylesheet">
<style>
:root {
--brand: #6D1F2F;
--brand-dark: #4a1520;
--brand-mid: #9e3347;
--brand-light: #f5e6e9;
--gold: #c9a84c;
--gold-lt: #f0d98a;
--ink: #1a1018;
--mist: #f8f4f5;
--border: #e8dde0;
--success: #2d6a4f;
--warn: #b36a00;
--r: 12px;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;background:var(--mist);color:var(--ink);min-height:100vh}
header{background:var(--brand);color:#fff;padding:0 2rem;display:flex;align-items:center;gap:1.25rem;height:80px;box-shadow:0 2px 16px rgba(109,31,47,.35);position:sticky;top:0;z-index:200}
.h-nav{margin-left:auto;display:flex;align-items:center;gap:.6rem}
.h-support,.h-pricing{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);color:#fff;font-size:.78rem;font-weight:600;letter-spacing:.05em;padding:.45rem 1rem;border-radius:6px;cursor:pointer;text-decoration:none;display:flex;align-items:center;gap:.4rem;line-height:1.1;height:34px;box-sizing:border-box;transition:background .18s}
.h-support:hover,.h-pricing:hover{background:rgba(255,255,255,.28)}
/* ---- First-time notice modal ---- */
#notice-overlay{position:fixed;inset:0;background:rgba(20,6,9,.72);backdrop-filter:blur(4px);z-index:10000;display:flex;align-items:center;justify-content:center;padding:1.5rem}
#notice-overlay.hidden{display:none}
#notice-box{background:#fff;border-radius:16px;padding:2rem 2rem 1.6rem;max-width:460px;width:94%;box-shadow:0 20px 60px rgba(0,0,0,.35);text-align:center}
#notice-box .notice-icon{font-size:2.2rem;margin-bottom:.5rem}
#notice-box h2{font-family:'Roboto Slab',serif;font-size:1.2rem;color:var(--brand);margin:0 0 .7rem}
#notice-box p{font-size:.9rem;color:#444;line-height:1.5;margin:0 0 .8rem}
#notice-box .notice-btn{margin-top:.4rem;padding:.6rem 1.6rem;background:var(--brand);color:#fff;border:none;border-radius:8px;font-family:'Roboto Slab',serif;font-weight:700;font-size:.85rem;letter-spacing:.03em;text-transform:uppercase;cursor:pointer}
#notice-box .notice-btn:hover{background:var(--brand-dark)}
/* ---- Change-password modal ---- */
#cpw-overlay{position:fixed;inset:0;background:rgba(20,6,9,.72);backdrop-filter:blur(4px);z-index:10001;display:flex;align-items:center;justify-content:center;padding:1.5rem}
#cpw-overlay.hidden{display:none}
#cpw-box{background:#fff;border-radius:14px;padding:1.8rem 1.8rem 1.5rem;max-width:380px;width:94%;box-shadow:0 20px 60px rgba(0,0,0,.35)}
#cpw-box h2{font-family:'Roboto Slab',serif;font-size:1.15rem;color:var(--brand);margin:0 0 .5rem}
/* ---- Pricing modal ---- */
#pricing-overlay{position:fixed;inset:0;background:rgba(20,6,9,.72);backdrop-filter:blur(4px);z-index:9999;display:flex;align-items:flex-start;justify-content:center;overflow-y:auto;padding:2rem 0}
#pricing-overlay.hidden{display:none}
#pricing-box{background:#fff;border-radius:16px;padding:2.4rem 2.2rem 2rem;max-width:820px;width:94%;box-shadow:0 20px 60px rgba(0,0,0,.35);position:relative;margin:auto}
#pricing-box .px-close{position:absolute;top:1rem;right:1.2rem;background:none;border:none;font-size:1.4rem;color:#bbb;cursor:pointer;line-height:1;transition:color .15s}
#pricing-box .px-close:hover{color:var(--brand)}
#pricing-box h2{font-family:'Roboto Slab',serif;font-size:1.4rem;color:var(--brand-dark);text-align:center;margin-bottom:.3rem}
#pricing-box .px-sub{text-align:center;font-size:.85rem;color:#888;margin-bottom:2rem}
.px-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:1.1rem}
.px-card{border:2px solid var(--border);border-radius:12px;padding:1.6rem 1.4rem;display:flex;flex-direction:column;gap:.6rem;position:relative;transition:border-color .2s,box-shadow .2s}
.px-card:hover{border-color:var(--brand-mid);box-shadow:0 6px 24px rgba(109,31,47,.1)}
.px-card.featured{border-color:var(--brand);box-shadow:0 6px 24px rgba(109,31,47,.15)}
.px-badge{position:absolute;top:-13px;left:50%;transform:translateX(-50%);background:var(--brand);color:#fff;font-size:.65rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;padding:.25rem .85rem;border-radius:20px;white-space:nowrap}
.px-tier{font-family:'Roboto Slab',serif;font-size:1rem;font-weight:700;color:var(--brand-dark);text-transform:uppercase;letter-spacing:.04em}
.px-price{font-family:'Roboto Slab',serif;font-size:2rem;font-weight:700;color:var(--brand)}
.px-price span{font-size:.85rem;font-weight:400;color:#999}
.px-desc{font-size:.82rem;color:#666;line-height:1.55;flex:1}
.px-features{list-style:none;display:flex;flex-direction:column;gap:.45rem;margin-top:.4rem}
.px-features li{font-size:.8rem;color:#555;display:flex;align-items:flex-start;gap:.5rem}
.px-features li::before{content:"✓";color:var(--brand);font-weight:700;flex-shrink:0}
.px-features li.px-feat-head{font-weight:700;color:var(--brand-dark);margin-top:.2rem}
.px-features li.px-feat-head::before{content:""}
.px-cta{margin-top:1rem;padding:.65rem 1rem;background:var(--brand);color:#fff;border:none;border-radius:8px;font-family:'Roboto Slab',serif;font-size:.82rem;font-weight:700;letter-spacing:.04em;text-transform:uppercase;cursor:pointer;transition:background .18s;text-align:center}
.px-cta:hover{background:var(--brand-dark)}
.px-card.featured .px-cta{background:var(--brand-dark)}
.px-note{text-align:center;font-size:.75rem;color:#aaa;margin-top:1.4rem}
@media(max-width:640px){.px-grid{grid-template-columns:1fr}.h-nav{gap:.4rem}}
.h-logo{width:60px;height:60px;flex-shrink:0;border-radius:50%;background-image:url("/logo.png");background-size:90%;background-position:center;background-repeat:no-repeat;background-color:var(--brand-dark)}
.h-titlewrap{min-width:0;overflow:hidden}
.h-title{font-family:'Roboto Slab',serif;font-size:1.25rem;font-weight:700;letter-spacing:.02em;text-transform:uppercase;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.h-sub{font-size:.72rem;opacity:.75;font-weight:400;margin-top:2px;letter-spacing:.08em;text-transform:uppercase}
.h-badge{margin-left:auto;background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);color:#fff;font-size:.68rem;font-family:'Roboto Slab',serif;font-weight:500;letter-spacing:.12em;text-transform:uppercase;padding:.35rem .9rem;border-radius:20px;white-space:nowrap}
/* Left sidebar navigation */
.layout{display:flex;align-items:flex-start}
.sidebar{position:sticky;top:80px;align-self:flex-start;width:210px;flex-shrink:0;background:#fff;border-right:2px solid var(--border);min-height:calc(100vh - 80px);padding:.9rem .6rem;display:flex;flex-direction:column;gap:.3rem;z-index:100}
.tab{display:flex;align-items:center;gap:.7rem;padding:.7rem .85rem;font-size:.82rem;font-weight:500;font-family:'Roboto Slab',serif;letter-spacing:.04em;text-transform:uppercase;color:#777;cursor:pointer;border-radius:8px;border-left:3px solid transparent;transition:background .15s,color .15s;white-space:nowrap}
.tab:hover{background:var(--brand-light);color:var(--brand)}
.tab.active{background:var(--brand-light);color:var(--brand);border-left-color:var(--brand);font-weight:700}
.tab-badge{background:var(--brand);color:#fff;font-size:.65rem;font-weight:700;padding:.15rem .45rem;border-radius:10px;min-width:18px;text-align:center;margin-left:auto}
.container{flex:1;min-width:0;max-width:1400px;padding:2rem 2rem 4rem;box-sizing:border-box}
.tab-panel{display:none}.tab-panel.active{display:block}
/* NEW badge next to a nav tab */
.nav-new{background:#2e7d32;color:#fff;font-size:.6rem;font-weight:800;letter-spacing:.5px;padding:.1rem .35rem;border-radius:8px;margin-left:.4rem}
/* Bunk Snapshot viewer */
.snap-meta{font-size:.83rem;color:#1A79BF;background:#eef4fb;border:1px solid #b9d2ec;border-radius:8px;padding:.55rem .85rem;margin-bottom:1rem;display:flex;align-items:center;gap:.5rem}
.snap-subtabs{display:flex;gap:.4rem;margin-bottom:1rem}
.snap-subtab{padding:.45rem .95rem;border:1px solid var(--border);border-radius:8px;background:#fff;color:#555;font-weight:600;font-size:.9rem;cursor:pointer}
.snap-subtab.on{background:var(--brand);border-color:var(--brand);color:#fff}
.snap-view{display:none}.snap-view.on{display:block}
.snap-search{width:100%;max-width:380px;margin-bottom:1rem;font-size:.92rem}
.snap-tbl{border-collapse:collapse;font-size:.85rem;width:100%}
.snap-tbl th,.snap-tbl td{border:1px solid #d8d8d8;padding:.3rem .5rem;text-align:center;white-space:nowrap}
.snap-tbl thead th{background:var(--brand);color:#fff;font-weight:700}
.snap-tbl td.snap-l,.snap-tbl th.snap-l{text-align:left}
.snap-tbl tr.snap-alt td{background:#f6f6f8}
.snap-tbl tr.snap-total td{background:#fde9cf;font-weight:700}
.snap-tbl th.snap-hl,.snap-tbl td.snap-hl{background:#dcecf9 !important}
.snap-tbl th.snap-hl{background:#1A79BF !important}
.snap-tbl .snap-sep{border-left:2px solid #9a9a9a}
.snap-bunk-block{margin-bottom:1.6rem}
.snap-bunk-name{font-weight:800;color:#000;font-size:1.15rem;margin:.2rem 0 .35rem}
.snap-grids{display:flex;flex-wrap:wrap;gap:1.6rem;align-items:flex-start}
.snap-grids>div{min-width:280px}
.snap-sec-title{font-weight:700;color:var(--brand-dark);margin:.2rem 0 .4rem}
/* Families directory */
.fam-cards{display:flex;flex-wrap:wrap;gap:1.2rem;align-items:flex-start}
.fam-card{border:1px solid var(--border);border-radius:12px;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.06);width:100%;max-width:520px;overflow:hidden}
.fam-card-hd{background:var(--brand);color:#fff;padding:.7rem 1rem;font-weight:700;font-size:1.05rem}
.fam-card-bd{padding:.9rem 1rem}
.fam-sec{margin-bottom:1rem}
.fam-sec:last-child{margin-bottom:0}
.fam-sec-h{font-weight:700;color:var(--brand-dark);font-size:.8rem;text-transform:uppercase;letter-spacing:.5px;border-bottom:1px solid #eee;padding-bottom:.25rem;margin-bottom:.5rem}
.fam-camper{padding:.5rem .65rem;border:1px solid #eee;border-radius:8px;margin-bottom:.6rem;background:#fafafa}
.fam-camper:last-child{margin-bottom:0}
.fam-camper-name{font-weight:700;color:#222}
.fam-camper-meta{font-size:.82rem;color:#666;margin:.1rem 0 .4rem}
.fam-wk-row{display:flex;align-items:center;gap:.5rem;margin:.18rem 0}
.fam-wk-lbl{font-size:.75rem;color:#777;width:190px;flex-shrink:0;white-space:nowrap}
.fam-day{display:inline-block;width:24px;height:22px;line-height:22px;text-align:center;border-radius:5px;font-size:.7rem;font-weight:700;background:#c0392b;color:#fff}
.fam-day.on{background:#2e7d32}
.fam-row{font-size:.88rem;color:#333;margin:.15rem 0}
.fam-row .fam-lbl,.fam-pickup .fam-lbl{color:#888;font-size:.78rem;margin-right:.35rem}
.fam-note{font-size:.8rem;color:#aaa;font-style:italic}
.fam-pickup{font-size:.85rem;color:#333;margin:.12rem 0}
.fam-card-hd{display:flex;align-items:center;justify-content:space-between;gap:.5rem}
.fam-edit-btn{background:rgba(255,255,255,.18);color:#fff;border:1px solid rgba(255,255,255,.5);border-radius:6px;font-size:.78rem;font-weight:600;padding:.25rem .6rem;cursor:pointer}
.fam-edit-btn:hover{background:rgba(255,255,255,.3)}
.fam-eday{cursor:pointer}
.fam-inp{width:100%;box-sizing:border-box;padding:.32rem .45rem;border:1px solid var(--border);border-radius:6px;font-size:.85rem;margin-bottom:.35rem}
.fam-inp-row{display:flex;gap:.4rem}
.fam-inp-row .fam-inp{flex:1}
.fam-flbl{font-size:.72rem;color:#888;margin:.3rem 0 .1rem;text-transform:uppercase;letter-spacing:.04em}
.fam-edit-actions{display:flex;align-items:center;gap:.6rem;margin-top:1rem;padding-top:.8rem;border-top:1px solid #eee}
.fam-edit-msg{font-size:.82rem;color:#777}
/* Pricing module */
.px-view{display:none}.px-view.on{display:block}
.px-sec{margin-bottom:1.6rem}
.px-sec-title{font-weight:700;color:var(--brand-dark);margin:.2rem 0 .5rem;font-size:1rem}
.px-tbl{border-collapse:collapse;font-size:.88rem;margin-bottom:.6rem}
.px-tbl th,.px-tbl td{border:1px solid #d8d8d8;padding:.35rem .55rem;text-align:center;white-space:nowrap}
.px-tbl thead th{background:var(--brand);color:#fff;font-weight:700}
.px-tbl td.px-l,.px-tbl th.px-l{text-align:left;font-weight:600}
.px-tbl tr.px-total td{background:#fde9cf;font-weight:700}
.px-rate-inp{width:78px;border:1px solid var(--border);border-radius:5px;padding:.25rem .35rem;text-align:right;font-size:.85rem}
.px-field{display:inline-flex;flex-direction:column;gap:.2rem;margin:0 .9rem .8rem 0;font-size:.8rem;color:#555}
.px-field select,.px-field input{padding:.4rem .5rem;border:1px solid var(--border);border-radius:6px;font-size:.9rem;min-width:120px}
.px-controls{display:flex;flex-wrap:wrap;align-items:flex-end;gap:.4rem;margin-bottom:1rem}
.px-camper-row{display:flex;flex-wrap:wrap;align-items:flex-end;gap:.6rem;padding:.6rem .7rem;border:1px solid #eee;border-radius:8px;margin-bottom:.6rem;background:#fafafa}
.px-total-box{margin-top:1rem;padding:1rem 1.2rem;border:2px solid var(--brand);border-radius:10px;background:#faf3f4;max-width:520px}
.px-total-box .px-grand{font-size:1.6rem;font-weight:800;color:var(--brand-dark)}
.px-line{display:flex;justify-content:space-between;gap:1.5rem;font-size:.88rem;padding:.15rem 0;border-bottom:1px dashed #e5d9dc}
.px-line.px-sub{font-weight:700;border-bottom:none;padding-top:.4rem}
.px-btn{background:var(--brand);color:#fff;border:none;border-radius:7px;padding:.5rem 1rem;font-weight:600;font-size:.85rem;cursor:pointer}
.px-btn.ghost{background:#fff;color:var(--brand);border:1px solid var(--brand)}
.px-diff-up{color:#c0392b}.px-diff-flat{color:#888}
.px-msg{font-size:.82rem;color:#777;margin-left:.5rem}
/* Rate sheet (print one-pager) */
#px-sheet .px-sheet-head{text-align:center;margin-bottom:1rem}
#px-sheet .px-sheet-head h2{font-family:'Roboto Slab',serif;color:var(--brand-dark);margin:.2rem 0}
.px-sheet-grid{display:flex;flex-wrap:wrap;gap:1.6rem;align-items:flex-start}
.payroll-table{border-collapse:collapse;width:100%;font-size:.85rem}
.payroll-table th,.payroll-table td{border:1px solid #cfcfcf;padding:.35rem .4rem;text-align:center;vertical-align:middle}
.payroll-table td{height:42px}
.payroll-table thead th{background:var(--brand);color:#fff;font-weight:700;white-space:nowrap}
.payroll-table td.pr-name{text-align:left;font-weight:600;min-width:160px;white-space:normal;line-height:1.15}
.payroll-table .pr-delcol{width:36px;min-width:36px}
.payroll-table td.pr-area{color:#555;min-width:92px;white-space:normal;line-height:1.15}
.payroll-table td.pr-area-edit,.payroll-table td.pr-name-edit{cursor:pointer}
.payroll-table td.pr-area-edit:hover,.payroll-table td.pr-name-edit:hover{background:#f4eef0;outline:1px dashed var(--brand)}
.pr-area-input{width:80px;font-size:.8rem;padding:2px 3px;border:1px solid var(--brand);border-radius:4px;text-align:center}
.payroll-table th.pr-extday{width:74px;min-width:74px}
.fam-table{border-collapse:collapse;width:100%;font-size:.8rem}
.fam-table th,.fam-table td{border:1px solid #e2e2e2;padding:.35rem .5rem;text-align:left;vertical-align:top}
.fam-table thead th{background:var(--brand);color:#fff;font-weight:600;white-space:nowrap;font-size:.72rem;text-transform:uppercase;letter-spacing:.03em}
.fam-table td.fam-cell{cursor:pointer;min-width:70px}
.fam-table td.fam-cell:hover{background:#f4eef0;outline:1px dashed var(--brand)}
.fam-table .fam-del{cursor:pointer;border:none;background:none;color:#c0392b;font-size:.9rem;padding:0}
.usr-ic{cursor:pointer;border:none;background:none;color:#1A79BF;font-size:1rem;padding:0 .3rem;line-height:1}
.usr-ic:hover{color:var(--brand)}
/* Camper Schedules */
.sched-hit{padding:.5rem .7rem;border:1px solid var(--border);border-radius:8px;margin-bottom:.4rem;cursor:pointer;display:flex;justify-content:space-between;align-items:center;font-size:.88rem}
.sched-hit:hover{background:var(--brand-light);border-color:var(--brand-mid)}
.sched-hit .sh-bunk{font-size:.78rem;color:#888}
.sched-wk{display:flex;align-items:center;gap:.6rem;padding:.5rem 0;border-bottom:1px solid #eee;flex-wrap:wrap}
.sched-wk .sw-label{width:170px;font-size:.85rem;color:var(--brand-dark);font-weight:600}
.sched-day{width:42px;height:36px;border:1.5px solid #c0392b;border-radius:6px;background:#c0392b;color:#fff;font-weight:700;font-size:.8rem;cursor:pointer;transition:all .12s}
.sched-day.on{background:#2e7d32;border-color:#2e7d32;color:#fff}
.sched-day:hover{filter:brightness(1.08)}
.sched-back{background:none;border:none;color:var(--brand);cursor:pointer;font-size:.85rem;font-weight:600;padding:0;margin-bottom:.6rem}
.sched-ov{font-size:.72rem;color:#9a5b00;margin-left:.3rem}
.season-row{display:flex;align-items:center;gap:.7rem;padding:.3rem 0}
.season-row .sr-wk{font-weight:700;color:var(--brand);width:64px;font-size:.85rem}
.season-row input[type=date]{padding:.4rem .5rem;border:1px solid var(--border);border-radius:6px;font-size:.85rem}
.season-row .sr-range{font-size:.82rem;color:#666}
.payroll-table td.pr-count{font-weight:700;color:var(--brand);width:34px}
.payroll-table tbody tr:nth-child(even){background:#f4eef0}
.payroll-table td.pr-cell,.payroll-table td.pr-xcell{cursor:pointer;font-weight:800;font-size:1.6rem;user-select:none;line-height:1}
.payroll-table td.st-check{color:#2e7d32}
.payroll-table td.st-x{color:#c0392b}
.payroll-table td.st-half{color:#1A79BF}
.payroll-table td.st-na{color:#888;font-size:1.05rem;font-weight:700}
.payroll-table th.pr-day,.payroll-table td.pr-cell{min-width:48px}
.payroll-table th.pr-extra{width:42px;min-width:42px;max-width:42px;background:#3f1119;color:#fff;white-space:nowrap;font-size:.6rem;letter-spacing:-.02em;padding-left:.12rem;padding-right:.12rem;line-height:1.1}
.payroll-table td.pr-xcell{width:42px;min-width:42px;max-width:42px}
.pr-xsep{border-left:2px solid #6d1f2f !important}
.payroll-table.pr-locked td.pr-cell,.payroll-table.pr-locked td.pr-xcell{cursor:not-allowed}
.payroll-table .pr-del{cursor:pointer;border:none;background:none;color:#c0392b;font-size:.95rem;padding:0}
.pr-week-sep{border-left:3px solid #6d1f2f !important}
.payroll-table th.pr-day-click{cursor:pointer}
.payroll-table th.pr-day-click:hover{background:#7a2236}
.payroll-table th.pr-day-active{background:var(--gold) !important;color:#1a1018 !important}
.payroll-table td.pr-missed-hl{background:#fff3d6}
.payroll-table tfoot .pr-dayclear{cursor:pointer;border:none;background:none;color:#c0392b;font-size:.95rem;padding:0;line-height:1}
.payroll-table tfoot .pr-dayclear:hover{color:#7a1420}
.payroll-table.pr-locked tfoot .pr-dayclear{opacity:.3;cursor:not-allowed}
/* Weeks grid: fixed column widths so the layout doesn't shift when filtering */
.payroll-table.pr-weeks{table-layout:fixed}
.pr-weeks .pr-hnum{width:38px}
.pr-weeks .pr-harea{width:104px}
.pr-weeks .pr-day{width:54px}
.pr-weeks .pr-extra{width:46px}
.pr-weeks .pr-delcol{width:34px}
/* (Staff column has no fixed width → it absorbs the remaining space) */
.pr-period-btn{padding:.4rem .8rem;border:1px solid var(--brand);background:#fff;color:var(--brand);border-radius:8px;cursor:pointer;font-weight:600;font-size:.85rem}
.pr-period-btn.active{background:var(--brand);color:#fff}
.pr-period-btn.pr-sm{padding:.28rem .55rem;font-size:.72rem;font-weight:600}
.pr-period-btn.pr-lock-on{background:#2e7d32;border-color:#2e7d32;color:#fff}
.pr-period-btn.pr-lock-on:hover{background:#27682b}
.pr-period-btn.pr-lock-off{background:#c0392b;border-color:#c0392b;color:#fff}
.pr-period-btn.pr-lock-off:hover{background:#a5311f}
.pr-input{padding:.45rem .6rem;border:1px solid var(--border);border-radius:8px;font-size:.85rem}
.pr-multi{position:relative;display:inline-block}
.pr-multi-btn{cursor:pointer;background:#fff;display:flex;align-items:center;gap:.5rem;min-width:120px;justify-content:space-between}
.pr-multi-btn .caret{font-size:.65rem;color:#888}
.pr-multi-menu{position:absolute;top:100%;left:0;margin-top:.25rem;background:#fff;border:1px solid var(--border);border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,.15);min-width:170px;max-height:460px;overflow-y:auto;z-index:50;padding:.3rem}
.pr-multi-menu.hidden{display:none}
.pr-multi-menu label{display:flex;align-items:center;gap:.5rem;padding:.35rem .5rem;border-radius:6px;cursor:pointer;font-size:.85rem;color:#333;white-space:nowrap}
.pr-multi-menu label:hover{background:#f4eef0}
.pr-multi-menu .pr-multi-sep{border-top:1px solid #eee;margin:.25rem 0}
.payroll-table caption{display:none}   /* on screen the title shows in #pr-title; caption is for print */
@media print {
  body * { visibility:hidden; }
  #payroll-table, #payroll-table * { visibility:visible; }
  #payroll-table { position:absolute; left:0; top:0; width:100%; font-size:9pt; }
  #payroll-table caption { display:table-caption; caption-side:top; text-align:left; font-weight:700; font-size:12pt; color:#6D1F2F; padding-bottom:.3rem; }
  #payroll-table th, #payroll-table td { min-width:0 }   /* shrink to fit the page */
  /* uniform day + extra columns so they print the same width */
  #payroll-table th.pr-day, #payroll-table td.pr-cell,
  #payroll-table th.pr-extra, #payroll-table td.pr-xcell { width:42px; min-width:42px; max-width:42px }
  #payroll-table .pr-del { display:none; }
  /* keep maroon headers, row shading and symbol colors when printing */
  #payroll-table, #payroll-table * { -webkit-print-color-adjust:exact; print-color-adjust:exact; }
  /* Extended Staff sheet fills the page width (Staff ~30%, 5 day cols split the rest) */
  #payroll-table.pr-ext { width:100%; }
  #payroll-table.pr-ext .pr-extday { width:14%; }
}
.card{background:#fff;border:1px solid var(--border);border-radius:var(--r);padding:1.5rem 1.75rem;margin-bottom:1.1rem;box-shadow:0 1px 4px rgba(0,0,0,.04);transition:box-shadow .2s}
.card:hover{box-shadow:0 3px 12px rgba(109,31,47,.07)}
/* Two tiles side by side (collapses to one column on narrow screens) */
.card-grid{display:grid;grid-template-columns:1fr 1fr;gap:1.3rem;align-items:stretch;margin-bottom:1.6rem}
.card-grid > .card{margin-bottom:0}
@media(max-width:900px){.card-grid{grid-template-columns:1fr}}
.util-col{display:flex;flex-direction:column;gap:1.1rem}
.util-col > .card{margin-bottom:0}
/* Help / FAQ accordion */
.faq{border:1px solid var(--border);border-radius:10px;margin-bottom:.7rem;overflow:hidden;background:#fff}
.faq>summary{cursor:pointer;list-style:none;padding:.85rem 1.1rem;font-family:'Roboto Slab',serif;font-weight:700;font-size:.92rem;color:var(--brand-dark);display:flex;align-items:center;gap:.6rem}
.faq>summary::-webkit-details-marker{display:none}
.faq>summary::before{content:"▸";color:var(--brand);font-size:.8rem;transition:transform .15s}
.faq[open]>summary::before{transform:rotate(90deg)}
.faq[open]>summary{border-bottom:1px solid var(--border);background:var(--brand-light)}
.faq-body{padding:.9rem 1.2rem 1.1rem;font-size:.88rem;color:#444;line-height:1.6}
.faq-body p{margin:0 0 .7rem}
.faq-body ul{margin:.2rem 0 .7rem;padding-left:1.2rem}
.faq-body li{margin-bottom:.35rem}
.faq-body strong{color:var(--brand-dark)}
.faq-q{font-weight:700;color:var(--brand-dark);margin-top:.6rem}
.help-sub{font-size:.8rem;color:var(--brand);font-weight:700;text-transform:uppercase;letter-spacing:.05em;margin:1.4rem 0 .5rem}
.help-sub:first-of-type{margin-top:.3rem}
.card-hd{display:flex;align-items:center;gap:.7rem;margin-bottom:1.1rem}
.card-num{width:26px;height:26px;background:var(--brand);color:#fff;border-radius:50%;font-size:.75rem;font-weight:700;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.card-title{font-family:'Roboto Slab',serif;font-size:1.05rem;font-weight:700;color:var(--brand-dark);letter-spacing:.01em;text-transform:uppercase}
.card-hint{font-size:.75rem;color:#999;margin-top:.15rem;font-weight:300}
label.lbl{display:block;font-size:.75rem;font-weight:600;color:var(--brand-dark);letter-spacing:.04em;text-transform:uppercase;margin-bottom:.4rem}
/* Drop zone */
.drop-zone{border:2px dashed var(--border);border-radius:var(--r);padding:1rem;text-align:center;cursor:pointer;transition:all .2s;background:var(--mist);position:relative}
.drop-zone:hover,.drop-zone.drag-over{border-color:var(--brand-mid);background:var(--brand-light)}
.drop-zone input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.drop-icon{font-size:2rem;margin-bottom:.4rem}
.drop-text{font-size:.88rem;color:#666}.drop-text strong{color:var(--brand)}
.drop-meta{font-size:.72rem;color:#bbb;margin-top:.3rem}
.file-chosen{display:none;align-items:center;gap:.7rem;padding:.65rem .9rem;background:#edfaf3;border:1px solid #a3d9b8;border-radius:8px;margin-top:.6rem;font-size:.83rem;color:var(--success);font-weight:500}
.file-chosen.visible{display:flex}
.file-chosen .rm{margin-left:auto;cursor:pointer;font-size:.9rem;color:#999;background:none;border:none;padding:0 .2rem}
/* Report type selector */
.report-types{display:flex;flex-wrap:wrap;gap:.6rem;margin-top:.5rem}
.rtype-section-hd{font-size:.8rem;font-weight:700;color:var(--brand);text-transform:uppercase;letter-spacing:.04em;margin:.3rem 0 .1rem}
.rtype-section-hd.labels{margin-top:1rem;padding-top:.8rem;border-top:1px solid #eee}
.rtype-btn{padding:.55rem 1.1rem;border:1.5px solid var(--border);border-radius:8px;background:#fff;color:#888;font-family:'Roboto Slab',serif;font-size:.78rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase;cursor:pointer;transition:all .15s;white-space:nowrap}
.rtype-btn.active{background:var(--brand);border-color:var(--brand);color:#fff}
.rtype-btn:hover:not(.active){border-color:var(--brand-mid);color:var(--brand-mid)}
/* Run button */
.run-btn{width:100%;padding:.95rem 2rem;background:var(--brand);color:#fff;border:none;border-radius:var(--r);font-family:'Roboto Slab',serif;font-size:1.05rem;font-weight:700;letter-spacing:.02em;text-transform:uppercase;cursor:pointer;display:flex;align-items:center;justify-content:center;gap:.65rem;transition:background .18s,transform .1s,box-shadow .18s;box-shadow:0 4px 14px rgba(109,31,47,.3);margin-top:1.25rem}
.run-btn:hover:not(:disabled){background:var(--brand-dark);box-shadow:0 6px 20px rgba(109,31,47,.4);transform:translateY(-1px)}
.run-btn:disabled{opacity:.55;cursor:not-allowed;transform:none;box-shadow:none}
/* Progress panel */
#prog-panel{display:none;background:#1a1018;border-radius:var(--r);padding:1.1rem 1.4rem;margin-top:1.1rem;border:1px solid #2d1e24}
#prog-panel.visible{display:block}
.prog-hd{display:flex;align-items:center;gap:.65rem;margin-bottom:.75rem;padding-bottom:.65rem;border-bottom:1px solid #2d1e24}
.prog-title{font-size:.82rem;font-weight:600;color:#e0d4d8;letter-spacing:.06em;text-transform:uppercase}
.spinner{width:15px;height:15px;border:2px solid rgba(255,255,255,.15);border-top-color:var(--gold);border-radius:50%;animation:spin .7s linear infinite;flex-shrink:0}
@keyframes spin{to{transform:rotate(360deg)}}
.pbar-wrap{background:rgba(255,255,255,.08);border-radius:4px;height:3px;margin-bottom:.65rem;overflow:hidden}
.pbar{height:100%;background:linear-gradient(90deg,var(--brand-mid),var(--gold));width:0%;transition:width .4s ease}
#log{font-family:monospace;font-size:.76rem;line-height:1.65;color:#c4b5bb;max-height:220px;overflow-y:auto}
#log .ok{color:#6fcf97}#log .warn{color:#f2c94c}#log .err{color:#eb5757}
/* Action bar */
.action-bar{display:flex;gap:.75rem;flex-wrap:wrap;margin-top:1.1rem}
.dl-btn{display:inline-flex;align-items:center;gap:.55rem;padding:.75rem 1.5rem;background:var(--gold);color:#1a1018;border-radius:8px;text-decoration:none;font-weight:700;font-size:.9rem;transition:background .15s,transform .1s;box-shadow:0 3px 10px rgba(201,168,76,.35);border:none;cursor:pointer}
.dl-btn:hover{background:var(--gold-lt);transform:translateY(-1px)}
/* Calendar card */
.cal-list{display:flex;flex-direction:column;gap:0}
.cal-row{display:flex;align-items:center;gap:1rem;padding:.6rem 0;border-bottom:1px solid var(--border)}
.cal-row:last-child{border-bottom:none}
.cal-dot{width:10px;height:10px;border-radius:50%;background:var(--brand);flex-shrink:0}
.cal-week .cal-dot{background:var(--gold)}
.cal-info{display:flex;align-items:baseline;gap:.75rem;flex-wrap:wrap}
.cal-date{font-size:.82rem;font-weight:700;color:var(--brand-dark);min-width:110px}
.cal-event{font-size:.85rem;color:#555}
.cal-week .cal-event{font-style:italic;color:var(--brand)}
/* Error card */
#error-card{display:none;background:#2d0d13;border:1px solid #6d1f2f;border-radius:var(--r);padding:1.1rem 1.4rem;margin-top:1.1rem;color:#f5c2cb;font-size:.85rem}
#error-card.visible{display:block}
#error-card strong{display:block;margin-bottom:.35rem;font-size:.95rem}
#recent-card{margin-top:1.4rem}
#recent-card .recent-hd{font-family:'Roboto Slab',serif;font-size:.85rem;font-weight:700;color:var(--brand);text-transform:uppercase;letter-spacing:.06em;margin-bottom:.7rem}
#recent-list{display:flex;flex-direction:column;gap:.45rem}
.recent-row{display:flex;align-items:center;justify-content:space-between;background:#faf7f7;border:1px solid #ecdcdf;border-radius:8px;padding:.55rem .9rem;gap:.75rem}
.recent-row:hover{background:#f5eeef}
.recent-info{flex:1;min-width:0}
.recent-name{font-size:.85rem;font-weight:600;color:#2d1018;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.recent-time{font-size:.75rem;color:#888;margin-top:.1rem}
.recent-dl{flex-shrink:0;padding:.35rem .85rem;background:var(--brand);color:#fff;border:none;border-radius:6px;font-size:.78rem;font-weight:600;text-decoration:none;cursor:pointer;transition:background .15s}
.recent-dl:hover{background:var(--brand-dark)}
#recent-empty{font-size:.82rem;color:#aaa;text-align:center;padding:.5rem 0}
/* ---- Config tab ---- */
.camp-block{background:#fff;border:1px solid var(--border);border-radius:var(--r);margin-bottom:1rem;overflow:hidden}
.camp-header{display:flex;align-items:center;gap:.75rem;padding:.8rem 1.1rem;background:var(--brand-light);border-bottom:1px solid var(--border)}
.camp-name-input{font-family:'Roboto Slab',serif;font-size:.95rem;font-weight:700;color:var(--brand-dark);border:none;background:transparent;letter-spacing:.02em;text-transform:uppercase;flex:1;outline:none;min-width:0}
.camp-name-input:focus{background:#fff;border-radius:4px;padding:0 .4rem}
.camp-rm{background:none;border:none;cursor:pointer;color:#bbb;font-size:1rem;padding:.2rem;transition:color .15s;flex-shrink:0}
.camp-rm:hover{color:var(--brand)}
.camp-toggle{background:none;border:none;cursor:pointer;color:var(--brand);font-size:.8rem;padding:.1rem .35rem;flex-shrink:0;line-height:1}
.camp-count{font-size:.72rem;color:#999;flex-shrink:0;white-space:nowrap}
.camp-header{cursor:default}
.bunk-table{width:100%;border-collapse:collapse}
.bunk-table th{font-size:.7rem;font-weight:600;color:#999;letter-spacing:.05em;text-transform:uppercase;padding:.5rem .9rem;border-bottom:1px solid var(--border);text-align:left}
.bunk-table td{padding:.45rem .9rem;border-bottom:1px solid #f5f0f1;vertical-align:middle}
.bunk-table tr:last-child td{border-bottom:none}
.bunk-table tr:hover td{background:var(--mist)}
.bunk-input{border:1.5px solid var(--border);border-radius:6px;padding:.38rem .6rem;font-size:.82rem;font-family:'DM Sans',sans-serif;color:var(--ink);background:#fff;transition:border-color .15s;width:100%}
.bunk-input:focus{outline:none;border-color:var(--brand-mid)}
.bunk-num-input{width:70px}
.bunk-rm{background:none;border:none;cursor:pointer;color:#ccc;font-size:.95rem;padding:.2rem;transition:color .15s}
.bunk-rm:hover{color:var(--brand)}
.add-bunk-btn{display:flex;align-items:center;gap:.45rem;padding:.5rem .9rem;background:none;border:1.5px dashed var(--border);border-radius:8px;color:var(--brand-mid);font-size:.8rem;font-weight:600;cursor:pointer;transition:all .15s;margin:.6rem .9rem}
.add-bunk-btn:hover{border-color:var(--brand-mid);background:var(--brand-light)}
.add-camp-btn{display:flex;align-items:center;gap:.5rem;padding:.6rem 1.1rem;background:none;border:1.5px dashed var(--border);border-radius:8px;color:var(--brand-mid);font-size:.83rem;font-weight:600;cursor:pointer;transition:all .15s;width:100%;justify-content:center;margin-bottom:1rem}
.add-camp-btn:hover{border-color:var(--brand-mid);background:var(--brand-light)}
.save-config-btn{width:100%;padding:.85rem 2rem;background:var(--brand);color:#fff;border:none;border-radius:var(--r);font-family:'Roboto Slab',serif;font-size:1rem;font-weight:700;letter-spacing:.02em;text-transform:uppercase;cursor:pointer;display:flex;align-items:center;justify-content:center;gap:.65rem;transition:background .18s,transform .1s,box-shadow .18s;box-shadow:0 4px 14px rgba(109,31,47,.3)}
.save-config-btn:hover{background:var(--brand-dark);box-shadow:0 6px 20px rgba(109,31,47,.4);transform:translateY(-1px)}
#save-msg{display:none;margin-top:.75rem;padding:.75rem 1.25rem;border-radius:8px;font-size:.95rem;font-weight:600;text-align:center;transition:opacity .6s ease}
#save-msg.ok{display:flex;align-items:center;justify-content:center;gap:.5rem;background:#edfaf3;border:1.5px solid #4caf82;color:#1e7d4a}
#save-msg.ok.fade-out{opacity:0}
#save-msg.err{display:block;background:#2d0d13;border:1px solid #6d1f2f;color:#f5c2cb}
/* Misc */
.section-title{font-family:'Roboto Slab',serif;font-size:.85rem;font-weight:700;color:var(--brand-dark);text-transform:uppercase;letter-spacing:.05em;margin-bottom:.65rem}
.empty-state{text-align:center;padding:3rem 2rem;color:#bbb}
.empty-state .empty-icon{font-size:2.5rem;margin-bottom:.75rem}
.empty-state p{font-size:.9rem;line-height:1.6}
/* Weather tile */
.wx-day{flex:1;min-width:80px;background:var(--mist);border:1px solid var(--border);border-radius:10px;padding:.65rem .5rem;text-align:center;display:flex;flex-direction:column;gap:.25rem;align-items:center}
.wx-dow{font-size:.7rem;font-weight:700;color:var(--brand);letter-spacing:.06em;text-transform:uppercase}
.wx-icon{font-size:1.6rem;line-height:1}
.wx-hi{font-size:.95rem;font-weight:700;color:var(--ink)}
.wx-lo{font-size:.78rem;color:#999}
.wx-desc{font-size:.65rem;color:#aaa;margin-top:.1rem}
.wx-pop{font-size:.65rem;color:#1A79BF;font-weight:600;margin-top:.15rem}
.wx-now{font-size:.85rem;color:var(--ink);font-weight:500;margin-bottom:.7rem}
.wx-row{display:flex;gap:.5rem;flex-wrap:wrap}
.wx-note{margin-top:.9rem;padding:.6rem .8rem;background:var(--mist);border:1px solid var(--border);border-radius:8px;font-size:.82rem;color:#555}
/* Week selector buttons */
.week-btn{padding:.55rem 1.1rem;border:1.5px solid var(--border);border-radius:8px;background:#fff;color:#888;font-family:'Roboto Slab',serif;font-size:.78rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase;cursor:pointer;transition:all .15s;white-space:nowrap}
.week-btn.active{background:var(--gold);border-color:var(--gold);color:#1a1018}
.week-btn:hover:not(.active){border-color:var(--gold);color:var(--gold)}
/* Responsive */
@media(max-width:640px){
/* sidebar becomes a horizontal icon rail across the top */
.layout{flex-direction:column}
.sidebar{flex-direction:row;width:auto;min-height:0;border-right:none;border-bottom:2px solid var(--border);position:sticky;top:64px;padding:.4rem .5rem;gap:.3rem;justify-content:center;overflow-x:auto}
.tab{border-left:none;border-bottom:3px solid transparent;border-radius:6px;padding:.55rem .7rem;font-size:1.1rem}
.tab.active{border-left:none;border-bottom-color:var(--brand)}
.tab span:not(.tab-badge){display:none}
.tab-badge{margin-left:.15rem}
header{padding:0 .8rem;gap:.6rem;height:64px}
.h-logo{width:42px;height:42px}
.h-title{font-size:.95rem}
.h-sub{display:none}
/* Pricing/Support collapse to icon-only so the camp name fits on one line */
.h-lbl{display:none}
.h-pricing,.h-support{padding:.45rem .6rem}
.h-badge{display:none}
.container{padding:1rem .75rem 3rem}
.card{padding:1.1rem 1rem}
.rtype-btn{padding:.5rem .8rem;font-size:.72rem}
.run-btn{font-size:.95rem}
.bunk-table{font-size:.78rem}
.bunk-table th,.bunk-table td{padding:.4rem .6rem}
}
/* ---- Password modal ---- */
#pw-overlay{position:fixed;inset:0;background:rgba(20,6,9,.72);backdrop-filter:blur(4px);z-index:9999;display:flex;align-items:center;justify-content:center}
#pw-overlay.hidden{display:none}
#pw-box{background:#fff;border-radius:14px;padding:2.4rem 2.2rem 2rem;max-width:420px;width:90%;box-shadow:0 20px 60px rgba(0,0,0,.35);text-align:center}
#pw-box .pw-logo{width:72px;height:72px;margin:0 auto .9rem}#pw-box .pw-logo img{width:72px;height:72px;object-fit:contain;mix-blend-mode:multiply;display:block}
#pw-box h2{font-family:'Roboto Slab',serif;font-size:1.25rem;color:var(--brand);margin:0 0 .4rem}
#pw-box .pw-sub{font-size:.85rem;color:#555;margin:0 0 1.4rem;line-height:1.55}
#pw-box .pw-sub strong{color:var(--brand-dark)}
#pw-input-wrap{display:flex;gap:.5rem;margin-bottom:.6rem}
#pw-input{flex:1;padding:.7rem 1rem;border:1.5px solid #ddd;border-radius:8px;font-size:.95rem;outline:none;transition:border .18s}
#pw-input:focus{border-color:var(--brand)}
#pw-submit{padding:.7rem 1.2rem;background:var(--brand);color:#fff;border:none;border-radius:8px;font-weight:700;font-size:.95rem;cursor:pointer;transition:background .18s}
#pw-submit:hover{background:var(--brand-dark)}
#pw-error{font-size:.82rem;color:#c0392b;min-height:1.1rem;margin-top:.15rem}
.pw-field{margin-bottom:.6rem}
.pw-field input{width:100%;padding:.7rem 1rem;border:1.5px solid #ddd;border-radius:8px;font-size:.95rem;outline:none;transition:border .18s}
.pw-field input:focus{border-color:var(--brand)}
.pw-go{width:100%;padding:.75rem 1rem;background:var(--brand);color:#fff;border:none;border-radius:8px;font-weight:700;font-size:.95rem;cursor:pointer;transition:background .18s;margin-top:.3rem}
.pw-go:hover{background:var(--brand-dark)}
.pw-toggle{font-size:.82rem;color:#777;margin-top:.9rem}
.pw-toggle a{color:var(--brand);font-weight:600;cursor:pointer;text-decoration:underline}
.h-user{position:relative;display:flex;align-items:center;color:#fff}
.h-user-trigger{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.25);color:#fff;font-size:.9rem;font-weight:700;letter-spacing:.02em;cursor:pointer;display:flex;align-items:center;gap:.4rem;padding:.4rem .8rem;height:34px;box-sizing:border-box;border-radius:6px;transition:background .18s}
.h-user-trigger:hover{background:rgba(255,255,255,.25)}
.h-user-trigger .caret{font-size:.65rem;opacity:.85}
.h-user-menu{position:absolute;top:100%;right:0;margin-top:.35rem;background:#fff;border:1px solid #e2e2e2;border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,.18);min-width:175px;overflow:hidden;z-index:9998}
.h-user-menu.hidden{display:none}
.h-user-menu button{display:flex;align-items:center;gap:.55rem;width:100%;text-align:left;background:none;border:none;padding:.65rem .9rem;font-size:.85rem;color:#333;cursor:pointer}
.h-user-menu button:hover{background:#f4eef0;color:var(--brand)}
.h-user-menu .menu-sep{border-top:1px solid #eee}
</style>
</head>
<body>

<!-- Login / register gate -->
<div id="pw-overlay">
  <div id="pw-box">
    <div class="pw-logo"><img src="/logo.png" alt="Elbow Lane Day Camp"></div>
    <h2>Elbow Lane Reporting Center</h2>

    <div id="login-view">
      <p class="pw-sub">Sign in to continue.</p>
      <div class="pw-field"><input id="login-username" placeholder="Username" autocomplete="username"></div>
      <div class="pw-field"><input id="login-password" type="password" placeholder="Password" autocomplete="current-password"></div>
      <button id="login-btn" class="pw-go">Sign In</button>
      <div class="pw-toggle">No account yet? <a id="show-register">Create one</a></div>
    </div>

    <div id="register-view" style="display:none">
      <p class="pw-sub">Create your account. Your username is shown on reports you upload.</p>
      <div class="pw-field"><input id="reg-username" placeholder="Choose a username (e.g. your name)" autocomplete="username"></div>
      <div class="pw-field"><input id="reg-password" type="password" placeholder="Choose a password" autocomplete="new-password"></div>
      <div class="pw-field"><input id="reg-code" placeholder="Access code" autocomplete="off"></div>
      <button id="reg-btn" class="pw-go">Create Account &amp; Sign In</button>
      <div class="pw-toggle">Already have an account? <a id="show-login">Sign in</a></div>
    </div>

    <div id="pw-error"></div>
  </div>
</div>

<!-- Pricing modal -->
<div id="pricing-overlay" class="hidden">
  <div id="pricing-box">
    <button class="px-close" id="pricing-close">&#x2715;</button>
    <h2>Simple, Transparent Pricing</h2>
    <p class="px-sub">Built for the summer season &mdash; pick the plan that fits your camp. No contracts, cancel anytime.</p>
    <div class="px-grid">
      <div class="px-card">
        <div class="px-tier">Starter</div>
        <div class="px-price">$34.99<span>/mo</span></div>
        <p class="px-desc">Great for smaller camps getting started with digital reporting.</p>
        <ul class="px-features">
          <li>Up to 50 reports per month</li>
          <li>1 staff login</li>
          <li>Custom reports built for your camp: attendance, group rosters, extended hours &amp; transportation</li>
          <li>On-screen camp snapshot: live group totals plus roster</li>
          <li>Upload-once master sheet &amp; season calendar</li>
          <li>Configurable bunks &amp; camps</li>
          <li>Print-ready Excel &amp; Word output</li>
          <li>Email support</li>
        </ul>
        <button class="px-cta" onclick="window.location='mailto:bhimpele@gmail.com?subject=Start%20my%20EL%20Reporting%20Center%20trial%20(Starter)'">Start Free Trial</button>
      </div>
      <div class="px-card featured">
        <div class="px-badge">Most Popular</div>
        <div class="px-tier">Pro</div>
        <div class="px-price">$49.99<span>/mo</span></div>
        <p class="px-desc">For active camps running reports and payroll all season long.</p>
        <ul class="px-features">
          <li class="px-feat-head">Everything in Starter, plus:</li>
          <li>Up to 50 reports per month</li>
          <li>Up to 10 staff logins</li>
          <li>Payroll &amp; staff attendance tracking</li>
          <li>Camper Schedules: set each camper's weekly attendance days</li>
          <li>Families directory: search a family, view &amp; edit contacts and schedules</li>
          <li>Family contacts + custom mailing &amp; transportation labels</li>
          <li>Recent reports history</li>
          <li>Priority email support</li>
        </ul>
        <button class="px-cta" onclick="window.location='mailto:bhimpele@gmail.com?subject=Start%20my%20EL%20Reporting%20Center%20trial%20(Pro)'">Start Free Trial</button>
      </div>
      <div class="px-card">
        <div class="px-tier">Unlimited</div>
        <div class="px-price">$99.99<span>/mo</span></div>
        <p class="px-desc">Full access for large or multi-camp operations.</p>
        <ul class="px-features">
          <li class="px-feat-head">Everything in Pro, plus:</li>
          <li>Unlimited reports</li>
          <li>Unlimited staff logins</li>
          <li>Multi-camp &amp; unlimited bunks</li>
          <li>Onboarding &amp; setup assistance</li>
          <li>Priority support</li>
        </ul>
        <button class="px-cta" onclick="window.location='mailto:bhimpele@gmail.com?subject=Start%20my%20EL%20Reporting%20Center%20trial%20(Unlimited)'">Start Free Trial</button>
      </div>
    </div>
    <p class="px-note">All plans include a 14-day free trial &mdash; no credit card required. No setup fees &middot; your data stays yours.</p>
  </div>
</div>

<!-- First-time "what's new" notice -->
<div id="notice-overlay" class="hidden">
  <div id="notice-box">
    <div class="notice-icon">📋</div>
    <h2>Heads up — new Utilities tab</h2>
    <p>The <strong>Bunks &amp; Camps</strong> tab is now <strong>Utilities</strong>. From now on, upload your <strong>Master Sheet</strong> from the <strong>Utilities</strong> tab (not the Reports tab).</p>
    <p style="color:#777;font-size:.82rem">You can also import &amp; manage <strong>Family Contacts</strong> there. Bunks &amp; Camps settings moved to the bottom of that tab.</p>
    <button id="notice-ok" class="notice-btn">Got it</button>
  </div>
</div>

<!-- Change-my-password modal -->
<div id="cpw-overlay" class="hidden">
  <div id="cpw-box">
    <h2>Change Password</h2>
    <p style="font-size:.85rem;color:#666;margin:0 0 1rem">Update the password for <strong id="cpw-who"></strong>.</p>
    <div class="pw-field"><input id="cpw-current" type="password" placeholder="Current password" autocomplete="current-password"></div>
    <div class="pw-field"><input id="cpw-new" type="password" placeholder="New password" autocomplete="new-password"></div>
    <div class="pw-field"><input id="cpw-confirm" type="password" placeholder="Confirm new password" autocomplete="new-password"></div>
    <div id="cpw-error" style="font-size:.82rem;color:#c0392b;min-height:1.1rem;margin-bottom:.4rem"></div>
    <div style="display:flex;gap:.5rem;justify-content:flex-end">
      <button id="cpw-cancel" class="pr-period-btn pr-sm">Cancel</button>
      <button id="cpw-save" class="pw-go" style="width:auto;margin:0;padding:.55rem 1.2rem">Save</button>
    </div>
  </div>
</div>

<header>
  <div class="h-logo" role="img" aria-label="Elbow Lane Day Camp"></div>
  <div class="h-titlewrap">
    <div class="h-title">Elbow Lane Day Camp</div>
    <div class="h-sub">Reporting Center</div>
  </div>
  <div class="h-nav">
    <span class="h-user" id="h-user" style="display:none">
      <button class="h-user-trigger" id="h-user-btn" title="Account"><span id="h-user-name"></span><span class="caret">▾</span></button>
      <div class="h-user-menu hidden" id="h-user-menu">
        <button id="menu-reset">🔑 Reset Password</button>
        <button id="menu-logout" class="menu-sep">↩ Sign Out</button>
      </div>
    </span>
    <button class="h-pricing" id="pricing-btn">$ <span class="h-lbl">Pricing</span></button>
    <a class="h-support" href="mailto:bhimpele@gmail.com?subject=EL%20Reporting%20Center%20Support">✉ <span class="h-lbl">Support</span></a>
  </div>
</header>

<div class="layout">

<nav class="sidebar">
  <div class="tab active" data-tab="upload">📂 <span>Reports</span></div>
  <div class="tab" data-tab="payroll">🗓️ <span>Payroll</span></div>
  <div class="tab" data-tab="snap" id="tab-snap-nav">📸 <span>Camp Snapshot</span></div>
  <div class="tab" data-tab="families">👪 <span>Families</span></div>
  <div class="tab" data-tab="pricing" id="tab-pricing-nav" style="display:none">💲 <span>Pricing</span><span class="nav-new">NEW</span></div>
  <div class="tab" data-tab="config">⚙️ <span>Utilities</span></div>
  <div class="tab" data-tab="help">❓ <span>FAQs</span></div>
</nav>

<div class="container">

<!-- ===== UPLOAD TAB ===== -->
<div class="tab-panel active" id="tab-upload">

  <div class="card">
    <div class="card-hd">
      <div>
        <div class="card-title">Select Report Type</div>
        <div class="card-hint">Choose the report you want to run</div>
      </div>
    </div>
    <div class="rtype-section-hd">📊 Reports</div>
    <div class="report-types" id="report-types">
        <button class="rtype-btn active" data-rtype="bunk_snapshot">Bunk Snapshot</button>
        <button class="rtype-btn" data-rtype="group_attendance">Group Attendance</button>
        <button class="rtype-btn" data-rtype="am_extend">AM Extend</button>
        <button class="rtype-btn" data-rtype="pm_extend">PM Extend</button>
        <button class="rtype-btn" data-rtype="pm_grp_extend">PM GRP Extend</button>
        <button class="rtype-btn" data-rtype="driver_totals">Driver Totals</button>
    </div>
    <div class="rtype-section-hd labels">🏷️ Labels</div>
    <div class="report-types">
        <button class="rtype-btn" data-rtype="inter_labels">Inter</button>
        <button class="rtype-btn" data-rtype="jr_transport_labels">Junior</button>
        <button class="rtype-btn" data-rtype="upper_labels">Upper</button>
        <button class="rtype-btn" data-rtype="mailing_labels">Mailing</button>
    </div>
  </div>

  <!-- Week selector — only visible when Driver Totals is selected -->
  <div class="card" id="week-card" style="display:none">
    <div class="card-hd">
      <div>
        <div class="card-title">Select Camp Week</div>
        <div class="card-hint">Only campers enrolled in the selected week are included. (The Driver Totals Report instead highlights that week's campers in yellow.)</div>
      </div>
    </div>
    <div style="display:flex;gap:.6rem;flex-wrap:wrap;margin-top:.25rem">
      <button class="week-btn active" data-week="1">Week 1</button>
      <button class="week-btn" data-week="2">Week 2</button>
      <button class="week-btn" data-week="3">Week 3</button>
      <button class="week-btn" data-week="4">Week 4</button>
      <button class="week-btn" data-week="5">Week 5</button>
      <button class="week-btn" data-week="6">Week 6</button>
      <button class="week-btn" data-week="7">Week 7</button>
      <button class="week-btn" data-week="8">Week 8</button>
    </div>
  </div>

  <!-- Saved-master banner: shows which data the report will use -->
  <div id="master-banner" style="display:none;align-items:center;gap:.6rem;padding:.6rem .85rem;background:#eef4fb;border:1px solid #b9d2ec;border-radius:8px;margin:0 0 .8rem;font-size:.83rem;color:#1A79BF;font-weight:500">
    <span>📋</span>
    <span id="master-banner-text" style="flex:1">—</span>
    <a id="master-download" href="/api/master/download" style="cursor:pointer;font-size:.75rem;color:#1A79BF;background:#fff;border:1px solid #b9d2ec;border-radius:6px;padding:.2rem .55rem;text-decoration:none">⬇ Download</a>
    <button id="master-clear" style="cursor:pointer;font-size:.75rem;color:#777;background:#fff;border:1px solid #ccd;border-radius:6px;padding:.2rem .55rem">Clear</button>
  </div>

  <button class="run-btn" id="run-btn" disabled>
    <span id="run-icon">⚙️</span>
    <span id="run-label">Run Report</span>
  </button>

  <div id="prog-panel">
    <div class="prog-hd">
      <div class="spinner" id="spinner"></div>
      <span class="prog-title" id="prog-title">Processing report…</span>
    </div>
    <div class="pbar-wrap"><div class="pbar" id="pbar"></div></div>
    <div id="log"></div>
  </div>

  <div class="action-bar" id="action-bar" style="display:none">
    <a class="dl-btn" id="dl-link" href="#" download>⬇ Download Report</a>
  </div>

  <!-- (Master sheet is now uploaded/managed from the Utilities tab.) -->

  <div id="error-card">
    <strong>⚠ Processing Error</strong>
    <span id="error-msg"></span>
  </div>

  <!-- ===== WEATHER + CALENDAR (two tiles side by side) ===== -->
  <div class="card-grid" style="margin-top:2.5rem;margin-bottom:1.1rem">
  <div class="card" id="weather-card" style="margin-bottom:0">
    <div class="card-hd" style="margin-bottom:.75rem">
      <span class="card-num">🌤</span>
      <div>
        <div class="card-title">5-Day Forecast — Warrington, PA</div>
      </div>
    </div>
    <div id="weather-body">
      <div style="color:#bbb;font-size:.82rem">Loading forecast…</div>
    </div>
  </div>

  <div class="card" id="calendar-card" style="margin-bottom:0">
    <div class="card-hd">
      <span class="card-num">📅</span>
      <div>
        <div class="card-title">Important Dates — Summer 2026</div>
      </div>
    </div>
    <div class="cal-list">
      <div class="cal-row cal-week">
        <div class="cal-dot"></div>
        <div class="cal-info">
          <div class="cal-date">Week of June 15</div>
          <div class="cal-event">Minicamp #1</div>
        </div>
      </div>
      <div class="cal-row">
        <div class="cal-dot"></div>
        <div class="cal-info">
          <div class="cal-date">June 22</div>
          <div class="cal-event">First Day of Camp</div>
        </div>
      </div>
      <div class="cal-row">
        <div class="cal-dot"></div>
        <div class="cal-info">
          <div class="cal-date">July 7</div>
          <div class="cal-event">Camp Pictures</div>
        </div>
      </div>
      <div class="cal-row">
        <div class="cal-dot"></div>
        <div class="cal-info">
          <div class="cal-date">July 20</div>
          <div class="cal-event">Olde Tyme Country Fair</div>
        </div>
      </div>
      <div class="cal-row">
        <div class="cal-dot"></div>
        <div class="cal-info">
          <div class="cal-date">August 4</div>
          <div class="cal-event">Family Fun Night</div>
        </div>
      </div>
      <div class="cal-row">
        <div class="cal-dot"></div>
        <div class="cal-info">
          <div class="cal-date">August 14</div>
          <div class="cal-event">Last Day of Camp</div>
        </div>
      </div>
      <div class="cal-row cal-week">
        <div class="cal-dot"></div>
        <div class="cal-info">
          <div class="cal-date">Week of August 17</div>
          <div class="cal-event">Minicamp #2</div>
        </div>
      </div>
    </div>
  </div>
  </div><!-- /card-grid weather+calendar -->

  <div id="recent-card" class="card">
    <div class="recent-hd">Recent Reports</div>
    <div id="recent-list"><div id="recent-empty">No reports yet.</div></div>
  </div>

</div><!-- /tab-upload -->

<!-- ===== PRICING TAB (admin only) ===== -->
<div class="tab-panel" id="tab-pricing">
  <div class="card">
    <div class="card-hd">
      <div>
        <div class="card-title">Pricing</div>
        <div class="card-hint">Maintain the rate tables, model price changes for an upcoming season, calculate what a family owes, and print a shareable rate sheet. All parts read from the same editable rates.</div>
      </div>
    </div>
    <div class="snap-subtabs">
      <button class="snap-subtab pxtab on" data-px="calc">Calculator</button>
      <button class="snap-subtab pxtab" data-px="explore">Explorer</button>
      <button class="snap-subtab pxtab" data-px="rates">Rate Settings</button>
      <button class="snap-subtab pxtab" data-px="sheet">Rate Sheet</button>
    </div>
    <div class="px-view on" id="px-calc"></div>
    <div class="px-view" id="px-explore"></div>
    <div class="px-view" id="px-rates"></div>
    <div class="px-view" id="px-sheet"></div>
  </div>
</div>

<!-- ===== FAMILIES TAB ===== -->
<div class="tab-panel" id="tab-families">
  <div class="card">
    <div class="card-hd">
      <div>
        <div class="card-title">Families</div>
        <div class="card-hint">Search a camper, parent, or family name to see everything the system has for that family: campers (with bunk, age, grade and weekly schedule), address, and contact info. Sourced from the saved Family Contacts and master sheet.</div>
      </div>
    </div>
    <input type="search" id="fam-dir-search" class="pr-input" placeholder="Search by camper, parent, or family name…" style="width:100%;max-width:460px;font-size:.95rem">
    <div id="fam-dir-results" style="margin-top:1rem"></div>
  </div>
</div>

<!-- ===== CAMP SNAPSHOT TAB ===== -->
<div class="tab-panel" id="tab-snap">
  <div class="card">
    <div class="card-hd">
      <div>
        <div class="card-title">Camp Snapshot</div>
        <div class="card-hint">A live, on-screen version of the Bunk Snapshot report, built from the master sheet currently saved on the server. Switch between camp/bunk totals and the full per-bunk roster.</div>
      </div>
    </div>
    <div class="snap-meta" id="snap-meta"><span>📋</span><span>Loading…</span></div>
    <div class="snap-subtabs">
      <button class="snap-subtab on" data-snap="totals">Totals</button>
      <button class="snap-subtab" data-snap="bunks">Bunks</button>
    </div>
    <div class="snap-view on" id="snap-totals"></div>
    <div class="snap-view" id="snap-bunks"></div>
  </div>
</div>

<!-- ===== UTILITIES TAB ===== -->
<div class="tab-panel" id="tab-config">

  <!-- Master Sheet + Family Contacts (two tiles side by side) -->
  <div class="card-grid">
  <div class="card">
    <div class="card-hd">
      <div>
        <div class="card-title">Master Sheet</div>
        <div class="card-hint">Upload the camper master sheet. It's saved on the server and used for every report — re-upload here whenever the camper data changes.</div>
      </div>
    </div>
    <div class="drop-zone" id="master-drop">
      <input type="file" id="master-file" accept=".csv,.xlsx,.xls">
      <div class="drop-icon">📊</div>
      <div class="drop-text"><strong>Click to choose</strong> or drag &amp; drop the master sheet</div>
      <div class="drop-meta">Accepted formats: .csv, .xlsx, .xls</div>
    </div>
    <div id="master-msg" style="font-size:.82rem;margin-top:.5rem"></div>
    <div id="master-status" style="display:none;align-items:center;gap:.6rem;padding:.6rem .85rem;background:#eef4fb;border:1px solid #b9d2ec;border-radius:8px;margin-top:.6rem;font-size:.83rem;color:#1A79BF;font-weight:500">
      <span>📋</span>
      <span id="master-status-text" style="flex:1">—</span>
      <a id="master-status-dl" href="/api/master/download" style="cursor:pointer;font-size:.75rem;color:#1A79BF;background:#fff;border:1px solid #b9d2ec;border-radius:6px;padding:.2rem .55rem;text-decoration:none">⬇ Download</a>
      <button id="master-status-clear" style="cursor:pointer;font-size:.75rem;color:#777;background:#fff;border:1px solid #ccd;border-radius:6px;padding:.2rem .55rem">Clear</button>
    </div>
  </div>
  <div class="card">
    <div class="card-hd">
      <div>
        <div class="card-title">Family Contacts</div>
        <div class="card-hint">Import the camper family-contact spreadsheet. It's stored on the server and used to source family-contact reports &amp; labels (not displayed here). Re-import to refresh.</div>
      </div>
    </div>
    <div class="drop-zone" id="fam-drop" style="padding:.75rem">
      <input type="file" id="fam-file" accept=".csv,.xlsx,.xls,.tsv">
      <div class="drop-icon" style="font-size:1.4rem">📇</div>
      <div class="drop-text"><strong>Click to choose</strong> or drag &amp; drop a family contact spreadsheet</div>
      <div class="drop-meta">Columns are auto-detected (last/first, bunk, parents, address, pickups)</div>
    </div>
    <div style="display:flex;align-items:center;gap:1rem;margin:.5rem 0 .2rem;font-size:.8rem;color:#666">
      <label><input type="radio" name="fam-import-mode" value="replace" checked> Replace all</label>
      <label><input type="radio" name="fam-import-mode" value="append"> Add to existing</label>
      <span id="fam-msg" style="margin-left:auto"></span>
    </div>
    <div id="fam-status" style="display:none;align-items:center;gap:.6rem;padding:.6rem .85rem;background:#eef4fb;border:1px solid #b9d2ec;border-radius:8px;margin-top:.6rem;font-size:.83rem;color:#1A79BF;font-weight:500">
      <span>📇</span>
      <span id="fam-status-text" style="flex:1">—</span>
    </div>
  </div>
  </div><!-- /card-grid master+family -->

  <!-- Camper Schedules (admins only, sits just below the master sheet) -->
  <div class="card" id="sched-card">
    <div class="card-hd">
      <div>
        <div class="card-title">Camper Schedules</div>
        <div class="card-hint">Find a camper and set which days they attend each week. This overrides the master's default day's for that week (used by Group Attendance, Extend and Bunk Snapshot reports). Pick a camper, set the days, then Save. Saved changes are kept after you upload a new master sheet and reapply automatically.</div>
      </div>
    </div>
    <input type="search" id="sched-search" class="pr-input" placeholder="Search camper name…" style="width:100%;max-width:420px;font-size:.95rem">
    <div id="sched-results" style="margin-top:.8rem"></div>
  </div>

  <!-- Season Calendar + User Accounts (left) | Bunks & Camps (right) -->
  <div class="card-grid" style="align-items:start">
  <div class="util-col">
  <div class="card">
    <div class="card-hd">
      <div>
        <div class="card-title">Season Calendar</div>
        <div class="card-hint">Set the first day of camp (Week 1 Monday). The 8 camp weeks are calculated from it and used for report week #/date ranges and the Payroll day columns.</div>
      </div>
    </div>
    <div style="display:flex;align-items:center;gap:.8rem;flex-wrap:wrap">
      <label style="font-size:.85rem;color:#555">Camp starts:
        <input type="date" id="season-start" style="padding:.4rem .5rem;border:1px solid var(--border);border-radius:6px;font-size:.85rem;margin-left:.4rem">
      </label>
      <button class="pr-period-btn" id="season-save">💾 Save</button>
      <span id="season-msg" style="font-size:.82rem;color:#777"></span>
    </div>
    <div id="season-summary" style="font-size:.82rem;color:#666;margin-top:.6rem"></div>
  </div>
  <!-- User accounts (admins only) -->
  <div class="card" id="users-card" style="display:none">
    <div class="card-hd">
      <div>
        <div class="card-title">User Accounts</div>
        <div class="card-hint">People sign in with their own username &amp; password. New users create an account with the shared access code on the sign-in screen. As an admin you can remove accounts here.</div>
      </div>
    </div>
    <div style="overflow-x:auto">
      <table class="fam-table" id="users-table"></table>
    </div>
    <div style="display:flex;gap:.5rem;flex-wrap:wrap;align-items:center;margin-top:.8rem;padding-top:.8rem;border-top:1px solid #eee">
      <strong style="font-size:.85rem;color:#555">Add user:</strong>
      <input class="pr-input" id="usr-username" placeholder="Username" style="width:140px">
      <input class="pr-input" id="usr-password" placeholder="Password" style="width:120px">
      <input class="pr-input" id="usr-email" placeholder="Their email (optional)" style="width:170px">
      <label style="font-size:.8rem;color:#666"><input type="checkbox" id="usr-admin"> Admin</label>
      <button class="pr-period-btn" id="usr-add">＋ Add User</button>
      <span id="usr-msg" style="font-size:.82rem;color:#777"></span>
    </div>
    <div id="usr-result" style="display:none;margin-top:.7rem;padding:.8rem .9rem;background:#edfaf3;border:1px solid #a3d9b8;border-radius:8px">
      <div style="font-size:.82rem;color:#2e7d32;font-weight:600;margin-bottom:.4rem">✓ Account created — share these credentials:</div>
      <pre id="usr-creds" style="font-size:.82rem;background:#fff;border:1px solid #d7ecdf;border-radius:6px;padding:.6rem .7rem;margin:0 0 .5rem;white-space:pre-wrap;word-break:break-word"></pre>
      <div style="display:flex;gap:.5rem;flex-wrap:wrap">
        <button class="pr-period-btn pr-sm" id="usr-copy">📋 Copy</button>
        <a class="pr-period-btn pr-sm" id="usr-email-link" href="#" style="text-decoration:none">✉ Email it</a>
        <span id="usr-copy-msg" style="font-size:.8rem;color:#777;align-self:center"></span>
      </div>
    </div>
  </div>
  </div><!-- /util-col -->
  <div class="card">
    <div class="card-hd">
      <span class="card-num" style="background:var(--gold);color:#1a1018">★</span>
      <div>
        <div class="card-title">Bunks &amp; Camps</div>
        <div class="card-hint">Manage bunk names, their numbers, and the camp group they belong to. Changes are saved to the server and used when processing all future reports.</div>
      </div>
    </div>

    <div id="config-status" style="display:none;align-items:center;gap:.6rem;padding:.6rem .85rem;background:#eef4fb;border:1px solid #b9d2ec;border-radius:8px;margin-bottom:.9rem;font-size:.83rem;color:#1A79BF;font-weight:500">
      <span>🗂️</span>
      <span id="config-status-text" style="flex:1">—</span>
    </div>
    <div id="config-warn" style="display:none;padding:.6rem .85rem;background:#fdf0e6;border:1px solid #f0c79b;border-radius:8px;margin-bottom:.9rem;font-size:.82rem;color:#9a5b00">
      ⚠ Persistent storage is not configured, so this configuration can reset to the bundled default when the app redeploys or restarts. Set the AWS S3 environment variables on the server to keep it saved.
    </div>

    <div id="camp-list"><!-- rendered by JS --></div>

    <button class="add-camp-btn" id="add-camp-btn">＋ Add Camp Group</button>

    <button class="save-config-btn" id="save-config-btn">💾 Save Configuration</button>
    <div id="save-msg"></div>
  </div>
  </div><!-- /card-grid season+users | bunks -->

</div><!-- /tab-config -->

<!-- ===== PAYROLL TAB ===== -->
<div class="tab-panel" id="tab-payroll">
  <div class="card">
    <div class="card-hd">
      <div>
        <div class="card-title">Staff Attendance</div>
        <div class="card-hint">Check each day a staff member is present. The count on the left totals the checks for the two-week period. Changes save automatically.</div>
      </div>
    </div>

    <div id="payroll-periods" style="display:flex;gap:.5rem;flex-wrap:wrap;margin:.3rem 0 .9rem"></div>

    <div style="display:flex;gap:1rem;flex-wrap:wrap;align-items:center;margin:0 0 .8rem;font-size:.82rem;color:#555">
      <label id="pr-search-wrap">Search:
        <input type="search" id="pr-search" class="pr-input" placeholder="Name or area…" style="width:180px;margin-left:.4rem"
               oninput="prSearch=this.value;renderPayroll()"></label>
      <span id="pr-area-filter" style="display:flex;align-items:center;gap:.45rem">Filter area:
        <span class="pr-multi" id="pr-area-wrap">
          <button type="button" class="pr-input pr-multi-btn" id="pr-area-btn">All areas <span class="caret">▾</span></button>
          <div class="pr-multi-menu hidden" id="pr-area-menu"></div>
        </span></span>
      <label>Sort by:
        <select id="pr-sort" class="pr-input" onchange="renderPayroll()">
          <option value="last">Last name</option>
          <option value="area">Area</option>
          <option value="total">Total (high to low)</option>
        </select></label>
      <label id="pr-ext-period-wrap" style="display:none">Shift:
        <select id="pr-ext-period" class="pr-input" onchange="prExtPeriod=this.value;renderPayroll()">
          <option value="ALL">AM &amp; PM</option>
          <option value="AM">AM only</option>
          <option value="PM">PM only</option>
        </select></label>
    </div>

    <div style="display:flex;gap:.5rem;justify-content:space-between;align-items:center;flex-wrap:wrap;margin:0 0 .3rem">
      <span id="pr-title" style="font-family:'Roboto Slab',serif;font-weight:700;font-size:1rem;color:var(--brand)"></span>
      <div style="display:flex;gap:.5rem;align-items:center">
        <span id="pr-tc-msg" style="font-size:.78rem;color:#777"></span>
        <button id="pr-timecard" class="pr-period-btn pr-sm" title="Import clock-ins from your payroll system">⏱ Import Time Card</button>
        <input type="file" id="pr-timecard-file" accept=".xlsx,.xls,.csv" style="display:none">
        <button id="pr-export" class="pr-period-btn pr-sm">⬇ Excel</button>
        <button id="pr-print" class="pr-period-btn pr-sm">🖨 Print / PDF</button>
        <button id="pr-lock" class="pr-period-btn pr-sm">🔓 Unlocked</button>
      </div>
    </div>
    <div id="pr-filter-note" style="font-size:.85rem;color:#9a5b00;margin:0 0 .5rem"></div>

    <div style="overflow-x:auto">
      <table class="payroll-table" id="payroll-table"></table>
    </div>

    <div style="display:flex;gap:.5rem;flex-wrap:wrap;align-items:center;margin-top:1.1rem;padding-top:.9rem;border-top:1px solid #eee">
      <strong style="font-size:.85rem;color:#555">Add staff:</strong>
      <input class="pr-input" id="pr-last"  placeholder="Last name"  style="width:140px">
      <input class="pr-input" id="pr-first" placeholder="First name" style="width:140px">
      <input class="pr-input" id="pr-area"  placeholder="Area"       style="width:140px">
      <button class="pr-period-btn" id="pr-add">＋ Add Staff</button>
      <span id="pr-msg" style="font-size:.82rem;color:#777"></span>
    </div>
  </div>
</div><!-- /tab-payroll -->

<!-- ===== HELP TAB ===== -->
<div class="tab-panel" id="tab-help">
  <div class="card">
    <div class="card-hd">
      <div>
        <div class="card-title">Frequently Asked Questions</div>
        <div class="card-hint">Quick answers for reports, payroll, schedules, families, contacts, and accounts. Tap a question to expand it.</div>
      </div>
    </div>

    <div class="help-sub">Getting started</div>

    <details class="faq" open>
      <summary>How do I run a report?</summary>
      <div class="faq-body">
        <p>1. In the <strong>Utilities</strong> tab, upload the camper <strong>Master Sheet</strong> (you only do this once — it's saved on the server).</p>
        <p>2. Go to <strong>Reports</strong>, choose a report, and pick a <strong>week</strong> if it asks.</p>
        <p>3. Click <strong>Run Report</strong> — the file builds and downloads automatically.</p>
      </div>
    </details>

    <details class="faq">
      <summary>What is the Master Sheet and how often do I upload it?</summary>
      <div class="faq-body">
        <p>It's the camper data export that feeds <strong>every</strong> report. Upload it in <strong>Utilities → Master Sheet</strong>. Re-upload only when the camper data changes — you don't upload it separately for each report.</p>
      </div>
    </details>

    <details class="faq">
      <summary>First-time setup checklist (admins)</summary>
      <div class="faq-body">
        <ul>
          <li>Create your account on the sign-in screen (you'll be the admin).</li>
          <li>Set the <strong>Season Calendar</strong> start date (Utilities) — it drives report week dates and the payroll columns.</li>
          <li>Upload the <strong>Master Sheet</strong> (Utilities).</li>
          <li>Import <strong>Family Contacts</strong> (Utilities) if you'll use contact reports/labels.</li>
        </ul>
      </div>
    </details>

    <div class="help-sub">Spreadsheet columns</div>

    <details class="faq">
      <summary>Master Sheet — required &amp; optional columns</summary>
      <div class="faq-body">
        <p>Columns are matched by their <strong>header name</strong> (not position), so extra columns and different ordering are fine. A file is recognized as a master when it has <strong>Session</strong>, <strong>Bunk</strong>, and an <strong>Enrollment Extra</strong> column.</p>
        <p class="faq-q">Required</p>
        <ul>
          <li><strong>Last name</strong> &amp; <strong>First name</strong> — camper name</li>
          <li><strong>Bunk</strong> — groups and sorts every report</li>
          <li><strong>Session</strong> — sets which of the 8 weeks each camper is enrolled in (used by week-specific reports)</li>
          <li><strong>Enrollment Extra</strong> — holds the AM "Drop-off" / PM "Pick-up" times and transport notes (2-Way / PM-Only)</li>
        </ul>
        <p class="faq-q">Used by specific reports — include if you run them</p>
        <ul>
          <li><strong>Monday–Friday</strong> (Yes/No) — partial-week schedule on attendance &amp; extend sheets</li>
          <li><strong>Driver</strong> &amp; <strong>Stop #</strong> — Driver Totals and Junior transport labels</li>
          <li><strong>Age</strong> &amp; <strong>Grade</strong> — Bunk Snapshot</li>
          <li><strong>Gender</strong> — splits part-time CIT groups on PM GRP Extend</li>
          <li><strong>CIT Bunk</strong> — a full-time CIT's assigned area on Group Attendance</li>
        </ul>
      </div>
    </details>

    <details class="faq">
      <summary>Family Contacts sheet — columns</summary>
      <div class="faq-body">
        <p>Columns are auto-detected by header, and any column we don't recognize is still stored. Typical headers:</p>
        <ul>
          <li><strong>Last name</strong>, <strong>First name</strong>, <strong>Family last name</strong> (shown on mailing labels), <strong>Bunk name</strong></li>
          <li><strong>P1 / P2 first name, last name, cell phone</strong> — primary &amp; secondary guardians</li>
          <li><strong>Primary family address 1 / 2, city, state, zip</strong></li>
          <li><strong>Authorized Pick-up/Emergency Contact 1–4</strong> — name + authorization</li>
        </ul>
        <p>For <strong>Mailing labels</strong> you need at least the address fields plus a Family last name (it falls back to Last name).</p>
      </div>
    </details>

    <div class="help-sub">Reports</div>

    <details class="faq">
      <summary>What does each report do?</summary>
      <div class="faq-body">
        <ul>
          <li><strong>Bunk Snapshot</strong> — roster by bunk with week/day attendance + a totals tab.</li>
          <li><strong>Group Attendance</strong> — one page per bunk with Mon–Fri sign-in cells. <em>(week-specific)</em></li>
          <li><strong>AM / PM Extend</strong> — extended-hours sign-in / sign-out sheets. <em>(week-specific)</em></li>
          <li><strong>PM GRP Extend</strong> — PM extended care grouped, one group per page. <em>(week-specific)</em></li>
          <li><strong>Driver Totals</strong> — per-driver transportation sheet. <em>(week-specific)</em></li>
          <li><strong>Labels (Word):</strong> Inter, Junior, Upper, and Mailing (one per address). <em>(Inter/Junior/Upper are week-specific)</em></li>
          <li><strong>Upper labels</strong> — one per Upper-camp camper attending the selected week (any number of days), showing the camper name with their bunk (no number) underneath.</li>
        </ul>
        <p>Reports marked <em>week-specific</em> ask you to pick a camp week before running.</p>
      </div>
    </details>

    <details class="faq">
      <summary>My report is empty or missing campers — why?</summary>
      <div class="faq-body">
        <p>Usually the <strong>selected week</strong> (the report only includes campers enrolled that week) or an <strong>out-of-date master sheet</strong>. Re-upload the latest master in Utilities and try again.</p>
      </div>
    </details>

    <details class="faq">
      <summary>How can I see a summary of camp (Camp Snapshot) without generating a whole new report?</summary>
      <div class="faq-body">
        <p>Open the <strong>📸 Camp Snapshot</strong> tab on the left. It shows the same information as the Bunk Snapshot report, right on screen — no file to download.</p>
        <ul>
          <li><strong>Totals</strong> — camper counts by bunk and by camp group, plus the by-week breakdowns.</li>
          <li><strong>Bunks</strong> — the full per-bunk roster (weeks, days, age, grade) with a camper search box.</li>
        </ul>
        <p>The <strong>current camp week</strong> column is shaded in both sub-tabs, and in Bunks the <strong>current weekday</strong> column is shaded too, so today stands out at a glance.</p>
        <p>It always reflects the <strong>master sheet currently saved on the server</strong> — the "Data last updated" line at the top tells you when that was. It loads instantly and only refreshes when a <strong>new master sheet is uploaded</strong>, so you don't wait on it each time.</p>
      </div>
    </details>

    <div class="help-sub">Camper schedules</div>

    <details class="faq">
      <summary>Can I change which days a camper attends without having to upload a new spreadsheet?</summary>
      <div class="faq-body">
        <p>Go to <strong>Utilities → Camper Schedules</strong>. Search a camper by name, click them, and you'll see a row for <strong>each week they're enrolled</strong> with Mon–Fri buttons.</p>
        <p>Each day button is <strong style="color:#2e7d32">green when scheduled</strong> and <strong style="color:#c0392b">red when not</strong> — tap to toggle. When you're done, click <strong>💾 Save schedule</strong>. (If you switch campers or go back with unsaved edits, it saves automatically so nothing is lost.)</p>
      </div>
    </details>

    <details class="faq">
      <summary>How do schedule changes affect reports?</summary>
      <div class="faq-body">
        <p>A saved change <strong>overrides the master sheet's default days</strong> for that camper, in that week only. It's applied to the week-specific reports that show attendance — <strong>Group Attendance</strong>, <strong>AM/PM Extend</strong>, <strong>PM GRP Extend</strong>, and <strong>Bunk Snapshot</strong> — when you run them for that week.</p>
        <p>Saved changes are <strong>kept after you upload a new master sheet</strong> and reapply automatically, so you don't have to redo them each time the master is refreshed.</p>
      </div>
    </details>

    <div class="help-sub">Staff Attendance (Payroll)</div>

    <details class="faq">
      <summary>How do the attendance cells work?</summary>
      <div class="faq-body">
        <p>Click a day cell to cycle <strong>blank → ✓ (present) → ✗</strong>. The count on the left totals the ✓ marks for the two-week block (a ½ counts as half a day). Changes <strong>save automatically</strong> and are shared across devices.</p>
        <p>The <strong>BS</strong> and <strong>SP\MTC</strong> columns appear on the Weeks 1 &amp; 2 block only and cycle ✓ / ✗ / ½ / N/A (they're never counted).</p>
        <p><strong>July 3</strong> has an extra <strong>½</strong> option (blank → ✓ → ½ → ✗), so a half-day holiday can be paid as half a day.</p>
      </div>
    </details>

    <details class="faq">
      <summary>What is the Holiday view?</summary>
      <div class="faq-body">
        <p>The <strong>🎆 Holiday</strong> button (next to Extended Staff) shows every staff member with just the holiday-week columns: <strong>BS</strong>, <strong>SP\MTC</strong>, and the holiday-week days (<strong>Th 7/2</strong>, <strong>Mon 7/6</strong>, <strong>Fri 7/3</strong>). It pulls the same marks you set on the week tabs, and you can edit right here too: any change syncs back to the week tabs automatically (July 3 still offers the ½ option).</p>
      </div>
    </details>

    <details class="faq">
      <summary>Searching, filtering, totals &amp; printing</summary>
      <div class="faq-body">
        <ul>
          <li><strong>Search</strong> matches name or area; <strong>Filter area</strong> lets you pick one or more areas; <strong>Sort</strong> by name/area/total.</li>
          <li><strong>Totals</strong> shows cumulative days across all 8 weeks; <strong>Extended Staff</strong> is a blank AM/PM check-in sheet.</li>
          <li><strong>Print / PDF</strong> and <strong>Excel</strong> both export exactly what's on screen (current search/filter/sort).</li>
          <li><strong>Lock</strong> freezes the sheet so no one can change it accidentally.</li>
        </ul>
      </div>
    </details>

    <details class="faq">
      <summary>How do I find who was missed (or absent/present) on a day?</summary>
      <div class="faq-body">
        <p><strong>Click the date</strong> at the top of any day column to filter the grid by that day. Each click <strong>cycles</strong> through:</p>
        <ul>
          <li><strong>Not yet marked</strong> (blank) — the ones that may have been missed</li>
          <li><strong>Marked ✗</strong> (absent)</li>
          <li><strong>Marked ✓</strong> (present)</li>
          <li><strong>Off</strong> — back to everyone</li>
        </ul>
        <p>The current mode shows above the table, and matching cells are highlighted. It works alongside Search and Filter area.</p>
      </div>
    </details>

    <details class="faq">
      <summary>Can I clear or import a whole day at once?</summary>
      <div class="faq-body">
        <p><strong>Clear a day:</strong> the <strong>✕</strong> at the bottom of each day column wipes every ✓/✗ for that day (after a confirmation).</p>
        <p><strong>Import clock-ins:</strong> click <strong>⏱ Import Time Card</strong> and choose the export from your payroll system (with Last name, First name, and Date columns). Matching staff get a ✓ on each date — even if the file uses nicknames (e.g. "Sam" matches "Samantha"). Any names it can't match are listed so you can mark them by hand. (Unlock the sheet first.)</p>
      </div>
    </details>

    <details class="faq">
      <summary>How do I add, rename, or remove staff?</summary>
      <div class="faq-body">
        <ul>
          <li><strong>Add:</strong> use the <em>Add staff</em> form at the bottom of the sheet.</li>
          <li><strong>Rename:</strong> click a staff <strong>name</strong> to edit it inline ("Last, First").</li>
          <li><strong>Change area:</strong> click the <strong>Area</strong> cell to edit it inline.</li>
          <li><strong>Remove:</strong> the <strong>✕</strong> at the end of each row.</li>
        </ul>
        <p>All of these are disabled while the sheet is <strong>Locked</strong>.</p>
      </div>
    </details>

    <div class="help-sub">Family contacts &amp; labels</div>

    <details class="faq">
      <summary>How do I load family contacts?</summary>
      <div class="faq-body">
        <p>In <strong>Utilities → Family Contacts</strong>, drop in the contact spreadsheet. Columns are auto-detected (name, bunk, parents, address, pickups), and new columns are kept automatically. Use <strong>Replace all</strong> to refresh, or <strong>Add to existing</strong> to append.</p>
        <p><strong>Mailing labels</strong> print one label per <strong>address</strong> (siblings share one), addressed as <strong>"The &lt;Family last name&gt; Family"</strong> over the street address, city, state and zip.</p>
      </div>
    </details>

    <details class="faq">
      <summary>How do I look up everything we have on a family?</summary>
      <div class="faq-body">
        <p>Open the <strong>👪 Families</strong> tab on the left and type a <strong>camper, parent, or family name</strong>. Each match shows a card with everything the system has for that family:</p>
        <ul>
          <li><strong>Camper(s)</strong> — bunk, age, grade, and the weekly schedule (green = attending, red = not), with any Camper Schedules changes already applied. Siblings appear together on one card.</li>
          <li><strong>Address</strong> and <strong>Contacts</strong> — primary &amp; secondary parents with phone numbers, plus authorized pickups.</li>
        </ul>
        <p>The information comes from the saved <strong>Family Contacts</strong> and <strong>master sheet</strong>, so load both in Utilities first. The directory refreshes automatically whenever you upload a new master or re-import contacts.</p>
      </div>
    </details>

    <details class="faq">
      <summary>Can I update a family's contact info or a camper's schedule from here?</summary>
      <div class="faq-body">
        <p>Yes. On any family card, click <strong>✎ Edit</strong> (top-right). You can change:</p>
        <ul>
          <li><strong>Contact info</strong> — address, primary &amp; secondary parent (name, phone, email), and up to four authorized pickups. This is shared by the whole family, so it updates every sibling at once.</li>
          <li><strong>Schedules</strong> — tap the day buttons (green = attending, red = not) for each enrolled week, per camper.</li>
        </ul>
        <p>Click <strong>Save</strong> to keep your changes or <strong>Cancel</strong> to discard them. Schedule changes save just the weeks you touched and behave exactly like the <strong>Camper Schedules</strong> tool — they're applied to the week-specific reports and kept even after a new master sheet is uploaded.</p>
      </div>
    </details>

    <div class="help-sub">Accounts</div>

    <details class="faq">
      <summary>Accounts &amp; passwords</summary>
      <div class="faq-body">
        <ul>
          <li>Everyone signs in with their own <strong>username + password</strong>; new users register with the shared <strong>access code</strong>.</li>
          <li>Change your own password by clicking <strong>your name</strong> in the top-right.</li>
          <li>Admins manage accounts (add / rename / reset password / remove) in <strong>Utilities → User Accounts</strong>.</li>
        </ul>
      </div>
    </details>

    <div class="help-sub">Troubleshooting</div>

    <details class="faq">
      <summary>My Bunks &amp; Camps config reset itself</summary>
      <div class="faq-body">
        <p>That happens when persistent storage isn't configured — the configuration can reset to the default when the app restarts/redeploys. If you see the orange warning on the Bunks &amp; Camps card, the server's <strong>AWS S3 environment variables</strong> need to be set so everything stays saved.</p>
      </div>
    </details>

    <details class="faq">
      <summary>Colors don't print on my PDF</summary>
      <div class="faq-body">
        <p>In the browser's print dialog, turn on <strong>"Background graphics"</strong> (Chrome: under <em>More settings</em>) so the maroon headers and check colors print.</p>
      </div>
    </details>

    <details class="faq">
      <summary>Where do my past reports go?</summary>
      <div class="faq-body">
        <p>The most recent ones are listed under <strong>Recent Reports</strong> on the Reports tab, each with a Download button.</p>
      </div>
    </details>

    <div style="margin-top:1.4rem;padding-top:1rem;border-top:1px solid #eee;font-size:.85rem;color:#666">
      Still stuck? <a class="h-support" style="background:var(--brand);border:none" href="mailto:bhimpele@gmail.com?subject=EL%20Reporting%20Center%20Help">✉ Email Support</a>
    </div>
  </div>
</div><!-- /tab-help -->

<!-- ===== SCHEDULES TAB (admin only) ===== -->
</div><!-- /container -->

</div><!-- /layout -->

<script>
// ─────────────────────────────────────────────
// Tab switching
// ─────────────────────────────────────────────
document.querySelectorAll('.tab').forEach(tab => {
  tab.addEventListener('click', () => {
    document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
    document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
    tab.classList.add('active');
    document.getElementById('tab-' + tab.dataset.tab).classList.add('active');
    if (tab.dataset.tab === 'payroll') renderPayroll();
    if (tab.dataset.tab === 'snap') loadBunkSnapshot();
    if (tab.dataset.tab === 'families') loadFamiliesDir();
    if (tab.dataset.tab === 'pricing') loadPricing();
  });
});

// Bunk Snapshot sub-tabs (Totals / Bunks)
document.querySelectorAll('.snap-subtab').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.snap-subtab').forEach(b => b.classList.remove('on'));
    document.querySelectorAll('.snap-view').forEach(v => v.classList.remove('on'));
    btn.classList.add('on');
    document.getElementById('snap-' + btn.dataset.snap).classList.add('on');
  });
});

// ─────────────────────────────────────────────
// Upload tab state
// ─────────────────────────────────────────────
let selectedReportType = 'bunk_snapshot';
let selectedWeek = 1;
let currentJobId = null;
let pollTimer = null;
let lastLineCount = 0;
let masterLoaded = false;

// Saved master sheet status — reflected on the Run Report banner AND the
// Utilities "Master Sheet" card.
async function loadMaster() {
  try {
    const res = await fetch('/api/master');
    const d = await res.json();
    masterLoaded = !!d.loaded;
    const banner = document.getElementById('master-banner');
    const status = document.getElementById('master-status');
    if (masterLoaded) {
      // If the username is an email, show only the part before the @
      const uploader = (d.uploaded_by || '').replace(/@.*$/, '');
      const by = uploader ? ` by <strong>${uploader}</strong>` : '';
      // Reformat "6/19/2026 4:00 PM EDT" -> "6/19/2026 @ 4:00 PM"
      let when = (d.uploaded_at || '').replace(/\s*[A-Z]{2,4}\s*$/, '')
                                      .replace(/\s+(\d{1,2}:\d{2}\s*[AP]M)/i, ' @ $1');
      const stamp = when ? ` &middot; uploaded on ${when}${by}` : '';
      document.getElementById('master-banner-text').innerHTML =
        `Using saved master: <strong>${d.filename || 'master sheet'}</strong>${stamp}` +
        `. Reports and Labels will use this data until an updated file is uploaded.`;
      banner.style.display = 'flex';
      if (status) {
        document.getElementById('master-status-text').innerHTML =
          `Current master: <strong>${d.filename || 'master sheet'}</strong>${stamp}`;
        status.style.display = 'flex';
      }
    } else {
      // No master saved — prompt the user to upload one in the Utilities tab
      document.getElementById('master-banner-text').innerHTML =
        `No master sheet saved yet. Upload one in the <strong>Utilities</strong> tab to run reports.`;
      document.getElementById('master-download').style.display = 'none';
      document.getElementById('master-clear').style.display = 'none';
      banner.style.display = 'flex';
      if (status) status.style.display = 'none';
    }
    // Restore the banner buttons when a master is present
    if (masterLoaded) {
      document.getElementById('master-download').style.display = '';
      document.getElementById('master-clear').style.display = '';
    }
  } catch(e) { masterLoaded = false; }
  updateRunBtn();
}

async function clearMaster() {
  try { await fetch('/api/master', {method: 'DELETE'}); } catch(e) {}
  loadMaster();
}
document.getElementById('master-clear').addEventListener('click', clearMaster);
document.getElementById('master-status-clear').addEventListener('click', clearMaster);

// Upload a master sheet from the Utilities tab
const masterDrop = document.getElementById('master-drop');
const masterFile = document.getElementById('master-file');
async function uploadMaster(f) {
  const msg = document.getElementById('master-msg');
  if (!f) return;
  msg.style.color = '#666'; msg.textContent = 'Uploading…';
  const fd = new FormData(); fd.append('file', f);
  try {
    const res = await fetch('/api/master', {method: 'POST', body: fd});
    const d = await res.json();
    if (!res.ok || d.error) { msg.style.color = '#c0392b'; msg.textContent = d.error || 'Upload failed.'; return; }
    msg.style.color = '#2e7d32'; msg.textContent = '✓ Master sheet saved.';
    loadMaster();
    snapRendered = false;   // repaint the Bunk Snapshot on next open (timestamp changed → re-fetch)
    famDirLoaded = false;   // master schedules/ages changed → refresh Families on next open
  } catch(e) { msg.style.color = '#c0392b'; msg.textContent = 'Network error: ' + e.message; }
}
masterDrop.addEventListener('dragover', e => { e.preventDefault(); masterDrop.classList.add('drag-over'); });
masterDrop.addEventListener('dragleave', () => masterDrop.classList.remove('drag-over'));
masterDrop.addEventListener('drop', e => {
  e.preventDefault(); masterDrop.classList.remove('drag-over');
  if (e.dataTransfer.files[0]) uploadMaster(e.dataTransfer.files[0]);
});
masterFile.addEventListener('change', e => { if (e.target.files[0]) uploadMaster(e.target.files[0]); });

function updateRunBtn() {
  document.getElementById('run-btn').disabled = !(masterLoaded && selectedReportType);
}

// Reports that use a camp-week selection
const WEEK_AWARE = ['driver_totals','group_attendance','am_extend','pm_extend','pm_grp_extend','inter_labels','jr_transport_labels','upper_labels'];

// Report type buttons
document.querySelectorAll('.rtype-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.rtype-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    selectedReportType = btn.dataset.rtype;
    document.getElementById('week-card').style.display =
      WEEK_AWARE.includes(selectedReportType) ? '' : 'none';
    updateRunBtn();
  });
});

// Week selector buttons
document.querySelectorAll('.week-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.week-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    selectedWeek = parseInt(btn.dataset.week, 10);
  });
});


// Run button
document.getElementById('run-btn').addEventListener('click', async () => {
  if (!masterLoaded || !selectedReportType) return;
  startProcessing();

  const fd = new FormData();   // always runs from the saved master
  fd.append('report_type', selectedReportType);
  if (WEEK_AWARE.includes(selectedReportType)) fd.append('week_num', selectedWeek);

  try {
    const res  = await fetch('/api/process', {method: 'POST', body: fd});
    const data = await res.json();
    if (!res.ok || data.error) { showError(data.error || 'Server error'); return; }
    currentJobId  = data.job_id;
    lastLineCount = 0;
    pollTimer     = setInterval(pollStatus, 1200);
  } catch(err) {
    showError('Network error: ' + err.message);
  }
});

function startProcessing() {
  document.getElementById('run-btn').disabled = true;
  document.getElementById('run-label').textContent = 'Processing…';
  document.getElementById('run-icon').textContent = '⏳';
  document.getElementById('prog-panel').classList.add('visible');
  document.getElementById('action-bar').style.display = 'none';
  document.getElementById('error-card').classList.remove('visible');
  document.getElementById('log').innerHTML = '';
  document.getElementById('pbar').style.width = '10%';
  document.getElementById('prog-title').textContent = 'Processing report…';
  document.getElementById('spinner').style.display = '';
}

async function pollStatus() {
  try {
    const res  = await fetch(`/api/status/${currentJobId}`);
    const data = await res.json();

    // Append new log lines
    const lines = data.progress || [];
    for (let i = lastLineCount; i < lines.length; i++) {
      const entry = lines[i];
      const div   = document.createElement('div');
      div.className = entry.level === 'ok' ? 'ok' : entry.level === 'err' ? 'err' : entry.level === 'warn' ? 'warn' : '';
      div.textContent = entry.msg;
      document.getElementById('log').appendChild(div);
    }
    lastLineCount = lines.length;
    document.getElementById('log').scrollTop = 999999;

    // Progress bar heuristic
    const pct = Math.min(10 + lastLineCount * 25, 90);
    document.getElementById('pbar').style.width = pct + '%';

    if (data.status === 'done') {
      clearInterval(pollTimer);
      document.getElementById('pbar').style.width = '100%';
      document.getElementById('spinner').style.animation = 'none';
      document.getElementById('spinner').style.borderTopColor = '#6fcf97';
      document.getElementById('prog-title').textContent = 'Complete! Downloading…';
      document.getElementById('run-btn').disabled = false;
      document.getElementById('run-label').textContent = 'Run Report';
      document.getElementById('run-icon').textContent = '⚙️';

      const dlLink = document.getElementById('dl-link');
      dlLink.href  = `/api/download/${currentJobId}`;
      document.getElementById('action-bar').style.display = 'flex';
      // Auto-download (button stays as a fallback if the browser blocks it)
      try { dlLink.click(); } catch(e) {}
      loadRecent();
      loadMaster();   // a freshly uploaded master is now saved for reuse
    }

    if (data.status === 'error') {
      clearInterval(pollTimer);
      showError(data.error || 'Unknown error');
    }
  } catch(err) {
    clearInterval(pollTimer);
    showError('Network error while polling: ' + err.message);
  }
}

function showError(msg) {
  document.getElementById('error-msg').textContent = msg;
  document.getElementById('error-card').classList.add('visible');
  document.getElementById('prog-panel').classList.remove('visible');
  document.getElementById('run-btn').disabled = false;
  document.getElementById('run-label').textContent = 'Run Report';
  document.getElementById('run-icon').textContent = '⚙️';
}

// ─────────────────────────────────────────────
// Config tab
// ─────────────────────────────────────────────
let campConfig = {camps: []};
let collapsedCamps = new Set();   // camp indices currently collapsed

async function loadConfig() {
  try {
    const res  = await fetch('/api/config');
    const data = await res.json();
    if (data.error) throw new Error(data.error);
    campConfig = data;
    collapsedCamps = new Set(campConfig.camps.map((_, i) => i));   // start collapsed
    renderCamps();
    loadConfigMeta();
  } catch(e) {
    document.getElementById('camp-list').innerHTML =
      `<div style="padding:1rem;color:#c0392b;font-size:.85rem">⚠ Could not load configuration: ${e.message}</div>`;
  }
}

async function loadConfigMeta() {
  try {
    const res = await fetch('/api/config/meta');
    const m = await res.json();
    const box = document.getElementById('config-status');
    const txt = document.getElementById('config-status-text');
    if (m.saved_at) {
      const by = (m.saved_by || '').replace(/@.*$/, '');
      const when = m.saved_at.replace(/\s*[A-Z]{2,4}\s*$/, '').replace(/\s+(\d{1,2}:\d{2}\s*[AP]M)/i, ' @ $1');
      txt.innerHTML = `Last saved on <strong>${when}</strong>` + (by ? ` by <strong>${by}</strong>` : '') +
                      '. The configuration only changes when someone clicks Save below.';
    } else {
      txt.innerHTML = 'No save recorded yet — this is the current configuration.';
    }
    box.style.display = 'flex';
    document.getElementById('config-warn').style.display = m.persistent ? 'none' : 'block';
  } catch(e) {}
}

function renderCamps() {
  const list = document.getElementById('camp-list');
  list.innerHTML = '';
  // Sort camps by their lowest bunk number; camps with no bunks go to the end
  const sorted = [...campConfig.camps].map((camp, ci) => ({ camp, ci }))
    .sort((a, b) => {
      const minA = a.camp.bunks.length ? Math.min(...a.camp.bunks.map(b => b.number)) : Infinity;
      const minB = b.camp.bunks.length ? Math.min(...b.camp.bunks.map(b => b.number)) : Infinity;
      return minA - minB;
    });
  sorted.forEach(({ camp, ci }) => {
    const collapsed = collapsedCamps.has(ci);
    const block = document.createElement('div');
    block.className = 'camp-block';
    block.innerHTML = `
      <div class="camp-header">
        <button class="camp-toggle" id="camp-caret-${ci}" title="Expand/collapse" onclick="toggleCamp(${ci})">${collapsed ? '▸' : '▾'}</button>
        <input class="camp-name-input" value="${escHtml(camp.name)}" placeholder="Camp Name"
          oninput="campConfig.camps[${ci}].name = this.value">
        <span class="camp-count">${camp.bunks.length} bunk${camp.bunks.length === 1 ? '' : 's'}</span>
        <button class="camp-rm" title="Remove camp" onclick="removeCamp(${ci})">✕</button>
      </div>
      <div class="camp-body" id="camp-body-wrap-${ci}" style="display:${collapsed ? 'none' : 'block'}">
        <table class="bunk-table">
          <thead>
            <tr>
              <th>Bunk Name</th>
              <th style="width:70px">Number</th>
              <th style="width:100px">Grp</th>
              <th style="width:36px"></th>
            </tr>
          </thead>
          <tbody id="bunk-body-${ci}">
            ${[...camp.bunks].sort((a,b) => a.number - b.number).map((b, bi) => bunkRow(ci, camp.bunks.indexOf(b), b)).join('')}
          </tbody>
        </table>
        <button class="add-bunk-btn" onclick="addBunk(${ci})">＋ Add Bunk</button>
      </div>
    `;
    list.appendChild(block);
  });
}

function toggleCamp(ci) {
  const wrap  = document.getElementById('camp-body-wrap-' + ci);
  const caret = document.getElementById('camp-caret-' + ci);
  if (collapsedCamps.has(ci)) {
    collapsedCamps.delete(ci); wrap.style.display = 'block'; caret.textContent = '▾';
  } else {
    collapsedCamps.add(ci); wrap.style.display = 'none'; caret.textContent = '▸';
  }
}

function bunkRow(ci, bi, b) {
  return `<tr id="bunk-${ci}-${bi}">
    <td><input class="bunk-input" value="${escHtml(b.name)}" placeholder="Bunk name"
      oninput="campConfig.camps[${ci}].bunks[${bi}].name = this.value"></td>
    <td><input class="bunk-input bunk-num-input" type="number" min="0" value="${b.number}"
      oninput="campConfig.camps[${ci}].bunks[${bi}].number = parseInt(this.value)||0"></td>
    <td><input class="bunk-input" value="${escHtml(b.grp||'')}" placeholder="Grp"
      oninput="campConfig.camps[${ci}].bunks[${bi}].grp = this.value"></td>
    <td><button class="bunk-rm" title="Remove bunk" onclick="removeBunk(${ci},${bi})">✕</button></td>
  </tr>`;
}

function addCamp() {
  campConfig.camps.push({name: 'New Camp', bunks: []});
  renderCamps();
}

function removeCamp(ci) {
  campConfig.camps.splice(ci, 1);
  // indices shifted — recollapse all so states stay aligned
  collapsedCamps = new Set(campConfig.camps.map((_, i) => i));
  renderCamps();
}

function addBunk(ci) {
  campConfig.camps[ci].bunks.push({name: '', number: 0, grp: ''});
  renderCamps();
  // Focus the new bunk name input
  const rows = document.querySelectorAll(`#bunk-body-${ci} tr`);
  if (rows.length) rows[rows.length-1].querySelector('input')?.focus();
}

function removeBunk(ci, bi) {
  campConfig.camps[ci].bunks.splice(bi, 1);
  renderCamps();
}

document.getElementById('add-camp-btn').addEventListener('click', addCamp);

document.getElementById('save-config-btn').addEventListener('click', async () => {
  const msg = document.getElementById('save-msg');
  if (!confirm('Save this camp / bunk configuration? This overwrites the saved version used by all reports.')) return;
  msg.className = '';
  msg.style.display = 'none';
  try {
    const res  = await fetch('/api/config', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(campConfig),
    });
    const data = await res.json();
    if (data.ok) {
      msg.innerHTML   = '<span style="font-size:1.2rem">✔</span> Configuration saved successfully.';
      msg.className   = 'ok';
      msg.style.display = '';
      msg.style.opacity = '1';
      loadConfigMeta();   // refresh the "last saved" audit box
      clearTimeout(msg._fadeTimer);
      msg._fadeTimer = setTimeout(() => {
        msg.classList.add('fade-out');
        setTimeout(() => { msg.style.display = 'none'; msg.className = ''; }, 650);
      }, 3000);
    } else {
      msg.textContent = '⚠ ' + (data.error || 'Save failed.');
      msg.className   = 'err';
    }
  } catch(e) {
    msg.textContent = '⚠ Network error: ' + e.message;
    msg.className   = 'err';
  }
});

function escHtml(s) {
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

// ---- Recent Reports ----
async function loadRecent() {
  const list = document.getElementById('recent-list');
  try {
    const res   = await fetch('/api/recent');
    const files = await res.json();
    if (!files.length) {
      list.innerHTML = '<div id="recent-empty">No reports yet.</div>';
      return;
    }
    list.innerHTML = files.map(f => {
      const d   = new Date(f.mtime * 1000);
      const fmt = d.toLocaleDateString('en-US', {month:'short',day:'numeric',year:'numeric'})
                + ' ' + d.toLocaleTimeString('en-US', {hour:'numeric',minute:'2-digit'});
      const name = f.name.replace(/\.xlsx$/i, '');
      return `<div class="recent-row">
        <div class="recent-info">
          <div class="recent-name" title="${escHtml(f.name)}">${escHtml(name)}</div>
          <div class="recent-time">${fmt}</div>
        </div>
        <a class="recent-dl" href="${escHtml(f.url)}" download="${escHtml(f.name)}">⬇ Download</a>
      </div>`;
    }).join('');
  } catch(e) {
    list.innerHTML = '<div id="recent-empty">Could not load recent reports.</div>';
  }
}

// ─────────────────────────────────────────────
// Weather tile
// ─────────────────────────────────────────────
const WX_ICONS = {
  0:'☀️', 1:'🌤️', 2:'⛅', 3:'☁️',
  45:'🌫️', 48:'🌫️',
  51:'🌦️', 53:'🌦️', 55:'🌧️',
  61:'🌧️', 63:'🌧️', 65:'🌧️',
  71:'🌨️', 73:'🌨️', 75:'❄️',
  77:'🌨️',
  80:'🌦️', 81:'🌦️', 82:'🌧️',
  85:'🌨️', 86:'❄️',
  95:'⛈️', 96:'⛈️', 99:'⛈️',
};
const WX_DESC = {
  0:'Clear', 1:'Mostly clear', 2:'Partly cloudy', 3:'Overcast',
  45:'Fog', 48:'Icy fog',
  51:'Light drizzle', 53:'Drizzle', 55:'Heavy drizzle',
  61:'Light rain', 63:'Rain', 65:'Heavy rain',
  71:'Light snow', 73:'Snow', 75:'Heavy snow', 77:'Snow grains',
  80:'Showers', 81:'Showers', 82:'Heavy showers',
  85:'Snow showers', 86:'Heavy snow showers',
  95:'Thunderstorm', 96:'T-storm / hail', 99:'T-storm / hail',
};
const DAYS_SHORT = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat'];

async function loadWeather() {
  const body = document.getElementById('weather-body');
  try {
    const res  = await fetch('/api/weather');
    const data = await res.json();
    if (data.error) throw new Error(data.error);

    const daysHtml = data.days.map(d => {
      const dt   = new Date(d.date + 'T12:00:00');
      const dow  = DAYS_SHORT[dt.getDay()];
      const icon = WX_ICONS[d.code] || '🌡️';
      const desc = WX_DESC[d.code]  || '';
      const pop  = (d.pop != null) ? `<div class="wx-pop">💧 ${d.pop}%</div>` : '';
      return `<div class="wx-day">
        <div class="wx-dow">${dow}</div>
        <div class="wx-icon">${icon}</div>
        <div class="wx-hi">${d.high}°</div>
        <div class="wx-lo">${d.low}°</div>
        <div class="wx-desc">${desc}</div>
        ${pop}
      </div>`;
    }).join('');

    // "Right now" line
    let nowHtml = '';
    if (data.current) {
      nowHtml = `<div class="wx-now">Right now: <strong>${data.current.temp}°</strong> ` +
                `${WX_ICONS[data.current.code] || ''} ${WX_DESC[data.current.code] || ''}</div>`;
    }

    // Camp planning note based on rain chance
    const RAINY = [51,53,55,56,57,61,63,65,66,67,80,81,82,95,96,99];
    const wet = data.days.filter(d => (d.pop != null && d.pop >= 50) || RAINY.includes(d.code));
    let note;
    if (!wet.length) {
      note = '☀️ Looks dry all week — great for outdoor activities.';
    } else {
      const names = wet.map(d => DAYS_SHORT[new Date(d.date + 'T12:00:00').getDay()]);
      note = `🌂 Rain likely ${names.join(', ')} — plan indoor backups.`;
    }

    body.innerHTML = nowHtml +
      `<div class="wx-row">${daysHtml}</div>` +
      `<div class="wx-note">${note}</div>`;
  } catch(e) {
    body.innerHTML = `<div style="color:#aaa;font-size:.82rem">Forecast unavailable</div>`;
  }
}

// ─────────────────────────────────────────────
// Payroll tab
// ─────────────────────────────────────────────
let payroll = {staff: [], checks: {}, days: []};
let prPeriod = 0;   // 0..3 → weeks 1&2, 3&4, 5&6, 7&8
let prTotals = false;   // when true, show the cumulative Totals view
let prExt = false;      // when true, show the blank Extended Staff sheet
let prHoliday = false;  // when true, show the Holiday week view
let prExtPeriod = 'ALL';// Extended Staff AM/PM filter: ALL | AM | PM
let prAreas = [];       // selected areas to filter by ([] = all areas)
let prSearch = '';      // free-text search across name + area
let prFilterDay = '';   // iso date being filtered by a day-header click
let prFilterMode = '';  // '' | 'blank' | 'x' | 'check' — cycled by repeat clicks
const PR_DAY_CYCLE = ['blank', 'x', 'check', ''];   // click order
const PR_MODE_LABEL = {blank: 'not yet marked', x: 'marked ✗ (absent)', check: 'marked ✓ (present)'};

// True if a staff member passes the current area filter
function prAreaMatch(s) { return prAreas.length === 0 || prAreas.includes(s.area || ''); }

// True if a staff member matches the search box (name, area, bunk, title)
function prSearchMatch(s) {
  const q = prSearch.trim().toLowerCase();
  if (!q) return true;
  return [s.last, s.first, s.area, s.bunk, s.title].filter(Boolean).join(' ').toLowerCase().includes(q);
}

// Build the multi-select area dropdown (button label + checkbox menu)
function renderAreaFilter(areas) {
  prAreas = prAreas.filter(a => areas.includes(a));   // drop areas that no longer exist
  const btn = document.getElementById('pr-area-btn');
  btn.firstChild.textContent =
    (prAreas.length === 0 ? 'All areas'
     : prAreas.length === 1 ? prAreas[0]
     : prAreas.length + ' areas') + ' ';
  const menu = document.getElementById('pr-area-menu');
  menu.innerHTML =
    `<label><input type="checkbox" id="pr-area-all" ${prAreas.length === 0 ? 'checked' : ''}> All areas</label>` +
    `<div class="pr-multi-sep"></div>` +
    areas.map(a => `<label><input type="checkbox" class="pr-area-cb" value="${a}" ${prAreas.includes(a) ? 'checked' : ''}> ${a}</label>`).join('');
  menu.querySelector('#pr-area-all').addEventListener('change', () => { prAreas = []; renderPayroll(); });
  menu.querySelectorAll('.pr-area-cb').forEach(cb => {
    cb.addEventListener('change', () => {
      prAreas = Array.from(menu.querySelectorAll('.pr-area-cb:checked')).map(x => x.value);
      renderPayroll();
    });
  });
}

// Open/close the area dropdown (set up once)
(function() {
  const btn = document.getElementById('pr-area-btn');
  const menu = document.getElementById('pr-area-menu');
  if (!btn || !menu) return;
  btn.addEventListener('click', e => { e.stopPropagation(); menu.classList.toggle('hidden'); });
  menu.addEventListener('click', e => e.stopPropagation());
  document.addEventListener('click', () => menu.classList.add('hidden'));
})();

async function loadPayroll() {
  try {
    const res = await fetch('/api/payroll');
    payroll = await res.json();
    renderPayroll();
  } catch(e) { /* ignore */ }
}

function prPeriodDays() {
  return payroll.days.slice(prPeriod * 10, prPeriod * 10 + 10);
}

function cellState(id, iso) {
  const v = (payroll.checks[id] || {})[iso];
  if (v === true || v === 'check') return 'check';   // legacy true == check
  if (v === 'x') return 'x';
  if (v === 'half') return 'half';
  return '';
}

// Days that may be marked ½ (half-day pay), matched by m/d label
const PR_HALF_DAYS = ['7/3'];
function prHalfDay(iso) {
  const d = (payroll.days || []).find(x => x.iso === iso);
  return !!(d && PR_HALF_DAYS.includes(d.md));
}

function prDayValue(st) { return st === 'check' ? 1 : st === 'half' ? 0.5 : 0; }

function prCount(id) {
  // Checkmarks count as a full day, ½ as half; X / blank do not count
  return prPeriodDays().reduce((n, d) => n + prDayValue(cellState(id, d.iso)), 0);
}

function xtraState(id, key) {
  const v = (payroll.checks[id] || {})[key];
  return ['check','x','half','na'].includes(v) ? v : '';
}

function symFor(st) {
  return st === 'check' ? '✓' : st === 'x' ? '✗' : st === 'half' ? '½' : st === 'na' ? 'N/A' : '';
}

// Day cell click — blank -> ✓ -> ✗ -> blank, or blank -> ✓ -> ½ -> ✗ -> blank on half-day dates
async function prClickDayCell(cell) {
  if (payroll.locked) return;
  const id = cell.dataset.id, dt = cell.dataset.date;
  const cur = cellState(id, dt);
  const next = prHalfDay(dt)
    ? (cur === '' ? 'check' : cur === 'check' ? 'half' : cur === 'half' ? 'x' : '')
    : (cur === '' ? 'check' : cur === 'check' ? 'x' : '');
  payroll.checks[id] = payroll.checks[id] || {};
  if (next) payroll.checks[id][dt] = next; else delete payroll.checks[id][dt];
  cell.textContent = symFor(next);
  cell.classList.remove('st-check','st-x','st-half','st-none');
  cell.classList.add('st-' + (next || 'none'));
  const cnt = document.getElementById('cnt-' + id); if (cnt) cnt.textContent = prCount(id);
  try { await fetch('/api/payroll/check', {method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({id, date: dt, value: next})}); } catch(e) {}
}

// Extra column click — blank -> ✓ -> ✗ -> ½ -> N/A -> blank (never counted)
async function prClickXCell(cell) {
  if (payroll.locked) return;
  const id = cell.dataset.id, key = cell.dataset.key;
  const xorder = ['', 'check', 'x', 'half', 'na'];
  const next = xorder[(xorder.indexOf(xtraState(id, key)) + 1) % xorder.length];
  payroll.checks[id] = payroll.checks[id] || {};
  if (next) payroll.checks[id][key] = next; else delete payroll.checks[id][key];
  cell.textContent = symFor(next);
  cell.classList.remove('st-check','st-x','st-half','st-na','st-none');
  cell.classList.add('st-' + (next || 'none'));
  try { await fetch('/api/payroll/check', {method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({id, date: key, value: next})}); } catch(e) {}
}

function renderPayroll() {
  // period buttons
  const pb = document.getElementById('payroll-periods');
  pb.innerHTML = '';
  for (let p = 0; p < 4; p++) {
    const b = document.createElement('button');
    b.className = 'pr-period-btn' + ((!prTotals && !prExt && !prHoliday && p === prPeriod) ? ' active' : '');
    b.textContent = `Weeks ${p*2+1} & ${p*2+2}`;
    b.onclick = () => { prPeriod = p; prTotals = false; prExt = false; prHoliday = false; renderPayroll(); };
    pb.appendChild(b);
  }
  const tb = document.createElement('button');   // Totals view, slightly separated
  tb.className = 'pr-period-btn' + (prTotals ? ' active' : '');
  tb.textContent = '🧮 Totals';
  tb.style.marginLeft = '1.4rem';
  tb.onclick = () => { prTotals = true; prExt = false; prHoliday = false; renderPayroll(); };
  pb.appendChild(tb);
  const eb = document.createElement('button');   // Extended Staff blank sheet
  eb.className = 'pr-period-btn' + (prExt ? ' active' : '');
  eb.textContent = '👤 Extended Staff';
  eb.style.marginLeft = '.5rem';
  eb.onclick = () => { prExt = true; prTotals = false; prHoliday = false; renderPayroll(); };
  pb.appendChild(eb);
  const hb = document.createElement('button');   // Holiday week view
  hb.className = 'pr-period-btn' + (prHoliday ? ' active' : '');
  hb.textContent = '🎆 Holiday';
  hb.style.marginLeft = '.5rem';
  hb.onclick = () => { prHoliday = true; prTotals = false; prExt = false; renderPayroll(); };
  pb.appendChild(hb);
  // area filter (multi-select dropdown)
  const areas = [...new Set(payroll.staff.map(s => s.area).filter(Boolean))].sort();
  renderAreaFilter(areas);
  const sortKey = document.getElementById('pr-sort').value;

  // Lock button + add-staff controls reflect lock state (both views)
  const lockBtn = document.getElementById('pr-lock');
  lockBtn.textContent = payroll.locked ? '🔒 Locked' : '🔓 Unlocked';
  lockBtn.classList.toggle('pr-lock-on', payroll.locked);
  lockBtn.classList.toggle('pr-lock-off', !payroll.locked);
  ['pr-last','pr-first','pr-area','pr-add'].forEach(id => {
    const el = document.getElementById(id); if (el) el.disabled = payroll.locked;
  });

  // AM/PM shift selector is only relevant to the Extended Staff sheet;
  // the area filter is not used there, so hide it.
  const extWrap = document.getElementById('pr-ext-period-wrap');
  if (extWrap) extWrap.style.display = prExt ? '' : 'none';
  document.getElementById('pr-ext-period').value = prExtPeriod;
  const areaWrap = document.getElementById('pr-area-filter');
  if (areaWrap) areaWrap.style.display = (prExt || prHoliday) ? 'none' : 'flex';
  const sortWrap = document.getElementById('pr-sort').closest('label');
  if (sortWrap) sortWrap.style.display = (prExt || prTotals || prHoliday) ? 'none' : '';
  const searchWrap = document.getElementById('pr-search-wrap');
  if (searchWrap) searchWrap.style.display = (prExt || prHoliday) ? 'none' : '';

  if (prExt) { renderExtTable('ALL', prExtPeriod); return; }
  if (prTotals) { renderTotalsTable('last'); return; }
  if (prHoliday) { renderHolidayTable(); return; }

  // table
  const days = prPeriodDays();
  // Day-header filter — only valid within the current block
  if (prFilterDay && !days.some(d => d.iso === prFilterDay)) { prFilterDay = ''; prFilterMode = ''; }
  const dayStateOk = (id) => {
    if (!prFilterDay || !prFilterMode) return true;
    const st = cellState(id, prFilterDay);
    return prFilterMode === 'blank' ? st === '' : st === prFilterMode;
  };
  let staff = payroll.staff.filter(s => prAreaMatch(s) && prSearchMatch(s) && dayStateOk(s.id));
  staff.sort((a,b) => {
    if (sortKey === 'total') {
      const d = prCount(b.id) - prCount(a.id);
      if (d) return d;
    } else if (sortKey === 'area') {
      const c = (a.area||'').toLowerCase().localeCompare((b.area||'').toLowerCase());
      if (c) return c;
    }
    return (a.last+a.first).toLowerCase().localeCompare((b.last+b.first).toLowerCase());
  });
  const showExtra = prPeriod === 0;   // BS / SP\\MTC columns only on the Weeks 1 & 2 block
  let note = '';
  if (prFilterDay && prFilterMode) {
    const md = days.find(d => d.iso === prFilterDay);
    note = `Showing staff <strong>${PR_MODE_LABEL[prFilterMode]}</strong> on <strong>${md.dow} ${md.md}</strong> — click the date to cycle, or again to clear.`;
  }
  prSetHeader(payrollTitle(), note);
  let html = `<caption>${payrollTitle()}</caption><thead><tr><th class="pr-hnum">#</th><th class="pr-hstaff">Staff</th><th class="pr-harea">Area</th>`;
  days.forEach((d,i) => {
    const active = (d.iso === prFilterDay && prFilterMode);
    const cls = 'pr-day pr-day-click' + (i === 5 ? ' pr-week-sep' : '') + (active ? ' pr-day-active' : '');
    html += `<th class="${cls}" data-iso="${d.iso}" title="Click to filter this day: not marked → ✗ → ✓ → off">${d.dow}<br>${d.md}</th>`;
  });
  if (showExtra) html += '<th class="pr-extra pr-xsep">BS</th><th class="pr-extra">SP\\MTC</th>';
  html += '<th class="pr-delcol"></th></tr></thead><tbody>';
  staff.forEach(s => {
    const c = payroll.checks[s.id] || {};
    html += `<tr data-id="${s.id}">`;
    html += `<td class="pr-count" id="cnt-${s.id}">${prCount(s.id)}</td>`;
    html += `<td class="pr-name pr-name-edit" data-id="${s.id}" title="Click to edit name">${s.last}, ${s.first}</td>`;
    const areaTxt = (s.area === 'Support' && s.title) ? s.title : (s.area || '');
    const bunkLine = s.bunk ? `<br><small style="color:#888;font-weight:400">${s.bunk}</small>` : '';
    html += `<td class="pr-area pr-area-edit" data-id="${s.id}" title="Click to edit area">${areaTxt}${bunkLine}</td>`;
    days.forEach((d,i) => {
      const st = cellState(s.id, d.iso);
      const sym = symFor(st);
      const hl = (d.iso === prFilterDay && prFilterMode &&
                  (prFilterMode === 'blank' ? st === '' : st === prFilterMode)) ? ' pr-missed-hl' : '';
      const cls = 'pr-cell st-' + (st || 'none') + (i === 5 ? ' pr-week-sep' : '') + hl;
      html += `<td class="${cls}" data-id="${s.id}" data-date="${d.iso}">${sym}</td>`;
    });
    if (showExtra) {
      for (let cc = 1; cc <= 2; cc++) {
        const key = `xtra:${prPeriod}:${cc}`;
        const xs = xtraState(s.id, key);
        const xcls = 'pr-xcell st-' + (xs || 'none') + (cc === 1 ? ' pr-xsep' : '');
        html += `<td class="${xcls}" data-id="${s.id}" data-key="${key}">${symFor(xs)}</td>`;
      }
    }
    html += `<td class="pr-delcol"><button class="pr-del" data-id="${s.id}" title="Remove">✕</button></td>`;
    html += '</tr>';
  });
  html += '</tbody>';
  // Footer: a clear-✕ under each day to wipe that whole day's marks
  html += '<tfoot><tr class="pr-foot"><td></td><td></td><td></td>';
  days.forEach((d,i) => {
    html += `<td${i === 5 ? ' class="pr-week-sep"' : ''}><button class="pr-dayclear" data-iso="${d.iso}" title="Clear all marks for ${d.dow} ${d.md}">✕</button></td>`;
  });
  if (showExtra) html += '<td></td><td></td>';
  html += '<td class="pr-delcol"></td></tr></tfoot>';
  const tbl = document.getElementById('payroll-table');
  tbl.innerHTML = html;
  tbl.className = 'payroll-table pr-weeks' + (payroll.locked ? ' pr-locked' : '');

  // Click a day's date to cycle its filter: not-marked → ✗ → ✓ → off
  tbl.querySelectorAll('th.pr-day-click').forEach(th => {
    th.addEventListener('click', () => {
      const iso = th.dataset.iso;
      if (prFilterDay !== iso) { prFilterDay = iso; prFilterMode = 'blank'; }
      else {
        const next = PR_DAY_CYCLE[(PR_DAY_CYCLE.indexOf(prFilterMode) + 1) % PR_DAY_CYCLE.length];
        prFilterMode = next;
        if (!next) prFilterDay = '';
      }
      renderPayroll();
    });
  });

  // Footer ✕: clear every mark for a whole day
  tbl.querySelectorAll('.pr-dayclear').forEach(btn => {
    btn.addEventListener('click', async () => {
      if (payroll.locked) return;
      const iso = btn.dataset.iso;
      const d = prPeriodDays().find(x => x.iso === iso);
      if (!confirm(`Clear ALL attendance marks for ${d ? d.dow + ' ' + d.md : 'this day'}? This removes every ✓ / ✗ entered that day for all staff.`)) return;
      Object.keys(payroll.checks).forEach(id => { if (payroll.checks[id]) delete payroll.checks[id][iso]; });
      renderPayroll();
      try { await fetch('/api/payroll/clearday', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({date: iso})}); } catch(e) {}
    });
  });

  // Day + extra cells (shared with the Holiday view)
  tbl.querySelectorAll('td.pr-cell').forEach(cell => cell.addEventListener('click', () => prClickDayCell(cell)));
  tbl.querySelectorAll('td.pr-xcell').forEach(cell => cell.addEventListener('click', () => prClickXCell(cell)));

  tbl.querySelectorAll('.pr-del').forEach(btn => {
    btn.addEventListener('click', async () => {
      if (payroll.locked) return;
      const id = btn.dataset.id;
      const s = payroll.staff.find(x => x.id === id);
      if (!confirm(`Remove ${s ? s.last + ', ' + s.first : 'this staff member'}?`)) return;
      payroll.staff = payroll.staff.filter(x => x.id !== id);
      delete payroll.checks[id];
      renderPayroll();
      try { await fetch('/api/payroll/staff/' + id, {method:'DELETE'}); } catch(e) {}
    });
  });

  // Click an Area cell to edit it inline
  // Click a name to edit it inline ("Last, First")
  tbl.querySelectorAll('td.pr-name-edit').forEach(td => {
    td.addEventListener('click', () => {
      if (payroll.locked || td.querySelector('input')) return;
      const id = td.dataset.id;
      const s = payroll.staff.find(x => x.id === id);
      if (!s) return;
      const orig = `${s.last || ''}, ${s.first || ''}`.replace(/^,\s*|,\s*$/g, '');
      td.innerHTML = `<input class="pr-area-input" style="width:96%;text-align:left" value="${orig.replace(/"/g,'&quot;')}">`;
      const inp = td.querySelector('input');
      inp.focus(); inp.select();
      let done = false;
      const commit = async (save) => {
        if (done) return; done = true;
        const val = inp.value.trim();
        if (save && val !== orig) {
          const ci = val.indexOf(',');
          const last = (ci >= 0 ? val.slice(0, ci) : val).trim();
          const first = (ci >= 0 ? val.slice(ci + 1) : '').trim();
          s.last = last; s.first = first;
          try { await fetch('/api/payroll/staff/' + id, {method:'PATCH',
                headers:{'Content-Type':'application/json'}, body: JSON.stringify({last, first})}); } catch(e) {}
        }
        renderPayroll();
      };
      inp.addEventListener('keydown', e => {
        if (e.key === 'Enter')  { e.preventDefault(); commit(true);  }
        if (e.key === 'Escape') { e.preventDefault(); commit(false); }
      });
      inp.addEventListener('blur', () => commit(true));
    });
  });

  tbl.querySelectorAll('td.pr-area-edit').forEach(td => {
    td.addEventListener('click', () => {
      if (payroll.locked || td.querySelector('input')) return;
      const id = td.dataset.id;
      const s = payroll.staff.find(x => x.id === id);
      if (!s) return;
      const orig = s.area || '';
      td.innerHTML = `<input class="pr-area-input" value="${orig.replace(/"/g,'&quot;')}">`;
      const inp = td.querySelector('input');
      inp.focus(); inp.select();
      let done = false;
      const commit = async (save) => {
        if (done) return; done = true;
        const val = inp.value.trim();
        if (save && val !== orig) {
          s.area = val;
          try { await fetch('/api/payroll/staff/' + id, {method:'PATCH',
                headers:{'Content-Type':'application/json'}, body: JSON.stringify({area: val})}); } catch(e) {}
        }
        renderPayroll();
      };
      inp.addEventListener('keydown', e => {
        if (e.key === 'Enter')  { e.preventDefault(); commit(true);  }
        if (e.key === 'Escape') { e.preventDefault(); commit(false); }
      });
      inp.addEventListener('blur', () => commit(true));
    });
  });
}

// Lock / unlock toggle
document.getElementById('pr-lock').addEventListener('click', async () => {
  const next = !payroll.locked;
  payroll.locked = next;
  renderPayroll();
  try { await fetch('/api/payroll/lock', {method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({locked: next})}); } catch(e) {}
});

// ---- Totals tab (cumulative checks across all weeks) ----
function totalChecks(id) {
  const c = payroll.checks[id] || {};
  return payroll.days.reduce((n, d) => n + (c[d.iso] === 'half' ? 0.5 : (c[d.iso] === 'check' || c[d.iso] === true) ? 1 : 0), 0);
}
function isJC(s) { return (s.title || '').toLowerCase().includes('junior'); }

// Extended Staff — blank printable check-in sheet (only AM/PM-extended staff)
function renderExtTable(filterArea, extPeriod) {
  extPeriod = extPeriod || 'ALL';
  const matchShift = e => extPeriod === 'ALL'
    || (extPeriod === 'AM' && /AM/i.test(e))
    || (extPeriod === 'PM' && /PM/i.test(e));
  const staff = payroll.staff
    .filter(s => s.ext && matchShift(s.ext) && (filterArea === 'ALL' || s.area === filterArea))
    .sort((a,b) => (a.last+a.first).toLowerCase().localeCompare((b.last+b.first).toLowerCase()));
  const shiftLbl = extPeriod === 'AM' ? 'AM' : extPeriod === 'PM' ? 'PM' : 'AM & PM';
  const extTitle = `Extended Staff (${shiftLbl}) — daily check-in (${staff.length})`;
  prSetHeader(extTitle, '');
  let html =
    `<caption>${extTitle}</caption><thead><tr><th>Staff</th>` +
    ['MON','TUES','WED','THURS','FRI'].map(d => `<th class="pr-extday">${d}</th>`).join('') +
    '</tr></thead><tbody>';
  staff.forEach(s => {
    html += `<tr><td class="pr-name">${s.last}, ${s.first}</td>` +
            '<td></td><td></td><td></td><td></td><td></td></tr>';
  });
  html += '</tbody>';
  const tbl = document.getElementById('payroll-table');
  tbl.innerHTML = html; tbl.className = 'payroll-table pr-ext';
}

// Holiday week — all staff with their BS / SP\\MTC plus the holiday-week days,
// pulling the same marks shown on the week tabs (editable; July 3 allows ½).
function renderHolidayTable() {
  const staff = payroll.staff.slice()
    .sort((a,b) => (a.last+a.first).toLowerCase().localeCompare((b.last+b.first).toLowerCase()));
  const byMd = md => (payroll.days || []).find(d => d.md === md) || null;
  // [m/d, dark line before this column?]
  const dayCols = [['7/2', true], ['7/6', false], ['7/3', true]]
    .map(([md, sep]) => ({d: byMd(md), sep})).filter(x => x.d);
  const title = 'Holiday Week Attendance';
  prSetHeader(title, dayCols.length ? '' : 'No holiday dates (7/2, 7/6, 7/3) were found in this season&rsquo;s calendar.');
  let html = `<caption>${title}</caption><thead><tr><th class="pr-hstaff">Staff</th><th class="pr-harea">Area</th>` +
    '<th class="pr-extra">BS</th><th class="pr-extra">SP\\MTC</th>';
  dayCols.forEach(c => { html += `<th class="pr-day${c.sep ? ' pr-week-sep' : ''}">${c.d.dow}<br>${c.d.md}</th>`; });
  html += '</tr></thead><tbody>';
  staff.forEach(s => {
    const areaTxt = (s.area === 'Support' && s.title) ? s.title : (s.area || '');
    const bunkLine = s.bunk ? `<br><small style="color:#888;font-weight:400">${s.bunk}</small>` : '';
    html += `<tr data-id="${s.id}"><td class="pr-name">${s.last}, ${s.first}</td>` +
            `<td class="pr-area">${areaTxt}${bunkLine}</td>`;
    ['xtra:0:1','xtra:0:2'].forEach(key => {
      const xs = xtraState(s.id, key);
      html += `<td class="pr-xcell st-${xs||'none'}" data-id="${s.id}" data-key="${key}">${symFor(xs)}</td>`;
    });
    dayCols.forEach(c => {
      const st = cellState(s.id, c.d.iso);
      html += `<td class="pr-cell st-${st||'none'}${c.sep ? ' pr-week-sep' : ''}" data-id="${s.id}" data-date="${c.d.iso}">${symFor(st)}</td>`;
    });
    html += '</tr>';
  });
  html += '</tbody>';
  const tbl = document.getElementById('payroll-table');
  tbl.innerHTML = html;
  tbl.className = 'payroll-table pr-weeks' + (payroll.locked ? ' pr-locked' : '');
  tbl.querySelectorAll('td.pr-cell').forEach(c => c.addEventListener('click', () => prClickDayCell(c)));
  tbl.querySelectorAll('td.pr-xcell').forEach(c => c.addEventListener('click', () => prClickXCell(c)));
}

// Totals view (rendered into the same Payroll table when the Totals button is on)
function renderTotalsTable(sortKey) {
  let staff = payroll.staff.filter(s => prAreaMatch(s) && prSearchMatch(s))
                           .map(s => ({...s, _total: totalChecks(s.id)}));
  staff.sort((a,b) => {
    if (sortKey === 'total') return b._total - a._total ||
      (a.last+a.first).toLowerCase().localeCompare((b.last+b.first).toLowerCase());
    if (sortKey === 'area') {
      const c = (a.area||'').toLowerCase().localeCompare((b.area||'').toLowerCase());
      if (c) return c;
    }
    return (a.last+a.first).toLowerCase().localeCompare((b.last+b.first).toLowerCase());
  });

  prSetHeader(payrollTitle(), '');
  let html = `<caption>${payrollTitle()}</caption><thead><tr><th>Staff</th><th>Area</th><th>Total Days<br><small style="font-weight:400">(all 8 weeks)</small></th></tr></thead><tbody>`;
  staff.forEach(s => {
    const jc = isJC(s) ? ' <small style="color:#1A79BF;font-weight:700">JC</small>' : '';
    const areaTxt = (s.area === 'Support' && s.title) ? s.title : (s.area || '');
    const bunkLine = s.bunk ? `<br><small style="color:#888;font-weight:400">${s.bunk}</small>` : '';
    html += `<tr><td class="pr-name">${s.last}, ${s.first}${jc}</td>` +
            `<td class="pr-area">${areaTxt}${bunkLine}</td>` +
            `<td class="pr-count">${s._total}</td></tr>`;
  });
  html += '</tbody>';
  const tbl = document.getElementById('payroll-table');
  tbl.innerHTML = html;
  tbl.className = 'payroll-table';
}

document.getElementById('pr-add').addEventListener('click', async () => {
  const last = document.getElementById('pr-last').value.trim();
  const first = document.getElementById('pr-first').value.trim();
  const area = document.getElementById('pr-area').value.trim();
  const msg = document.getElementById('pr-msg');
  if (!last && !first) { msg.textContent = 'Enter a name.'; return; }
  msg.textContent = 'Adding…';
  try {
    const res = await fetch('/api/payroll/staff', {method:'POST', headers:{'Content-Type':'application/json'},
          body: JSON.stringify({last, first, area})});
    const entry = await res.json();
    payroll.staff.push(entry);
    document.getElementById('pr-last').value = '';
    document.getElementById('pr-first').value = '';
    document.getElementById('pr-area').value = '';
    msg.textContent = '';
    renderPayroll();
  } catch(e) { msg.textContent = 'Error adding staff.'; }
});

function payrollTitle() {
  let t = prTotals ? 'Staff Attendance Totals — All 8 Weeks'
                   : `Staff Attendance — Weeks ${prPeriod*2+1} & ${prPeriod*2+2}`;
  if (prAreas.length) t += '  —  ' + prAreas.join(', ');
  return t;
}

// Set the section title (left of the action buttons) + the filter note line below
function prSetHeader(title, note) {
  const t = document.getElementById('pr-title');
  const n = document.getElementById('pr-filter-note');
  if (t) t.textContent = title;
  if (n) { n.innerHTML = note || ''; n.style.display = note ? '' : 'none'; }
}

// Print / save-as-PDF (prints exactly what's on screen, filtered/sorted).
// Narrow views (Extended Staff, Totals) print portrait; the wide grid lands­cape.
document.getElementById('pr-print').addEventListener('click', () => {
  // All payroll views print portrait — the grid only needs ~8" of width, so
  // landscape just wastes space and fits fewer rows per page.
  let st = document.getElementById('pr-print-orient');
  if (!st) { st = document.createElement('style'); st.id = 'pr-print-orient'; document.head.appendChild(st); }
  st.textContent = `@media print{@page{size:portrait;margin:.4in}}`;
  window.print();
});

// Export the current (filtered/sorted) view to a real .xlsx (server-built)
document.getElementById('pr-export').addEventListener('click', () => {
  const view = prTotals ? 'totals' : prExt ? 'ext' : prHoliday ? 'holiday' : 'weeks';
  const areas = encodeURIComponent(prExt ? '' : prAreas.join(','));
  const sort = document.getElementById('pr-sort').value;
  const q = encodeURIComponent(prExt ? '' : prSearch.trim());
  window.location = `/api/payroll/export?view=${view}&period=${prPeriod}&areas=${areas}&sort=${sort}&extp=${prExtPeriod}&q=${q}`;
});

// Import time card (clock-ins) → mark ✓ for matched staff on each date
const prTcFile = document.getElementById('pr-timecard-file');
document.getElementById('pr-timecard').addEventListener('click', () => {
  if (payroll.locked) { alert('Unlock the sheet before importing.'); return; }
  prTcFile.click();
});
prTcFile.addEventListener('change', async e => {
  const f = e.target.files[0];
  e.target.value = '';
  if (!f) return;
  const msg = document.getElementById('pr-tc-msg');
  msg.style.color = '#666'; msg.textContent = 'Importing time card…';
  const fd = new FormData(); fd.append('file', f);
  try {
    const res = await fetch('/api/payroll/import-timecard', {method:'POST', body: fd});
    const d = await res.json();
    if (!res.ok || d.error) { msg.style.color = '#c0392b'; msg.textContent = d.error || 'Import failed.'; return; }
    await loadPayroll();
    msg.style.color = '#2e7d32';
    msg.textContent = `✓ Marked ${d.checks_set} check-in(s) for ${d.staff_matched} staff across ${d.dates.length} day(s).`;
    if (d.unmatched && d.unmatched.length) {
      msg.style.color = '#9a5b00';
      msg.textContent += `  ${d.unmatched.length} not matched.`;
      alert('These names from the time card were not matched to a staff member (mark them manually if needed):\n\n' + d.unmatched.join('\n'));
    }
  } catch(err) { msg.style.color = '#c0392b'; msg.textContent = 'Network error: ' + err.message; }
});

// ---- Family contacts ----
let families = [];
const famEsc = s => String(s ?? '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');

async function loadFamilies() {
  let d = {};
  try {
    const res = await fetch('/api/families');
    d = await res.json();
    families = d.families || [];
  } catch(e) { families = []; }
  const box = document.getElementById('fam-status');
  const txt = document.getElementById('fam-status-text');
  if (!box) return;
  if (families.length) {
    const uploader = (d.uploaded_by || '').replace(/@.*$/, '');
    let when = (d.uploaded_at || '').replace(/\s*[A-Z]{2,4}\s*$/, '')
                                    .replace(/\s+(\d{1,2}:\d{2}\s*[AP]M)/i, ' @ $1');
    txt.innerHTML = `<strong>${families.length}</strong> family records stored` +
      (when ? ` &middot; uploaded on ${when}` : '') +
      (uploader ? ` by <strong>${uploader}</strong>` : '') + '.';
    box.style.display = 'flex';
  } else {
    box.style.display = 'none';
  }
}

const famDrop = document.getElementById('fam-drop');
const famFile = document.getElementById('fam-file');
async function importFamilies(f) {
  const msg = document.getElementById('fam-msg');
  if (!f) return;
  const mode = (document.querySelector('input[name="fam-import-mode"]:checked') || {}).value || 'replace';
  msg.style.color = '#666'; msg.textContent = 'Importing…';
  const fd = new FormData(); fd.append('file', f); fd.append('mode', mode);
  try {
    const res = await fetch('/api/families/import', {method:'POST', body: fd});
    const d = await res.json();
    if (!res.ok || d.error) { msg.style.color = '#c0392b'; msg.textContent = d.error || 'Import failed.'; return; }
    msg.style.color = '#2e7d32'; msg.textContent = `✓ Imported ${d.count} (${d.total} total).`;
    loadFamilies();
    famDirLoaded = false;   // refresh the Families directory on next open
  } catch(e) { msg.style.color = '#c0392b'; msg.textContent = 'Network error: ' + e.message; }
}
famDrop.addEventListener('dragover', e => { e.preventDefault(); famDrop.classList.add('drag-over'); });
famDrop.addEventListener('dragleave', () => famDrop.classList.remove('drag-over'));
famDrop.addEventListener('drop', e => {
  e.preventDefault(); famDrop.classList.remove('drag-over');
  if (e.dataTransfer.files[0]) importFamilies(e.dataTransfer.files[0]);
});
famFile.addEventListener('change', e => { if (e.target.files[0]) { importFamilies(e.target.files[0]); e.target.value=''; } });

// ---- Season calendar (single start date) ----
let seasonStart = '';

async function loadSeason() {
  try {
    const res = await fetch('/api/season');
    const d = await res.json();
    seasonStart = d.start || '';
    const inp = document.getElementById('season-start');
    if (inp) inp.value = seasonStart;
    renderSeasonSummary(d);
  } catch(e) {}
}

function renderSeasonSummary(d) {
  const el = document.getElementById('season-summary');
  if (!el) return;
  const w1 = (d.weeks && d.weeks[0] && d.weeks[0].range) || '';
  el.innerHTML = (w1 && d.end)
    ? `8 weeks: <strong>Week 1</strong> ${w1} &hellip; through <strong>${d.end}</strong>.`
    : '';
}

document.getElementById('season-save').addEventListener('click', async () => {
  const msg = document.getElementById('season-msg');
  const start = document.getElementById('season-start').value;
  if (!start) { msg.style.color = '#c0392b'; msg.textContent = 'Pick the first day of camp.'; return; }
  // Warn only if the start actually changed — Payroll checks are keyed to dates
  if (seasonStart && start !== seasonStart && !confirm(
      'Changing the start date will re-date the Payroll day columns. Any attendance ' +
      'checks already entered are tied to the old dates and will NOT carry over to the ' +
      'new ones.\n\nThis is safe before the season starts. Continue?')) {
    return;
  }
  msg.style.color = '#666'; msg.textContent = 'Saving…';
  try {
    const res = await fetch('/api/season', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({start})});
    const d = await res.json();
    if (!res.ok || d.error) { msg.style.color = '#c0392b'; msg.textContent = d.error || 'Could not save.'; return; }
    seasonStart = d.start || start;
    renderSeasonSummary(d);
    msg.style.color = '#2e7d32'; msg.textContent = '✓ Saved. Reports & Payroll now use these dates.';
    loadPayroll();   // refresh payroll day columns
  } catch(e) { msg.style.color = '#c0392b'; msg.textContent = 'Network error: ' + e.message; }
});

// Load all data for the app (only after the user is signed in)
function loadAllData() {
  loadConfig();
  loadRecent();
  loadWeather();
  loadMaster();
  loadPayroll();
  loadFamilies();
  loadUsers();
  loadSeason();
  loadSchedules();
}

// ---- Camper Schedules (admin) ----
let schedCampers = [], schedOverrides = {}, schedWeeks = [];
let schedCurrentKey = null, schedDirty = false;   // explicit-save state
const SCHED_DAYS = ['M','T','W','R','F'];
const SCHED_DAY_LABEL = {M:'M', T:'T', W:'W', R:'Th', F:'F'};

async function loadSchedules() {
  if (!currentUser) return;
  schedCurrentKey = null; schedDirty = false;
  try {
    const res = await fetch('/api/schedules');
    if (!res.ok) return;
    const d = await res.json();
    schedCampers = d.campers || [];
    schedOverrides = d.overrides || {};
    schedWeeks = d.weeks || [];
  } catch(e) {}
  renderSchedResults();
}

function renderSchedResults() {
  const box = document.getElementById('sched-results');
  if (!box) return;
  const q = (document.getElementById('sched-search').value || '').trim().toLowerCase();
  if (!q) { box.innerHTML = '<div style="color:#aaa;font-size:.85rem">Start typing a camper name…</div>'; return; }
  const hits = schedCampers.filter(c => (c.name || '').toLowerCase().includes(q)).slice(0, 50);
  if (!hits.length) { box.innerHTML = '<div style="color:#aaa;font-size:.85rem">No matches.</div>'; return; }
  box.innerHTML = hits.map(c =>
    `<div class="sched-hit" data-key="${famEsc(c.key)}"><span>${famEsc(c.name)}</span><span class="sh-bunk">${famEsc(c.bunk)}</span></div>`).join('');
  box.querySelectorAll('.sched-hit').forEach(el => el.addEventListener('click', () => openSchedEditor(el.dataset.key)));
}

// Enter the editor for a camper (save any pending edits on the previous camper first)
async function openSchedEditor(key) {
  if (schedDirty && schedCurrentKey && schedCurrentKey !== key) await saveSchedule(schedCurrentKey);
  schedCurrentKey = key;
  schedDirty = false;
  renderSchedEditor(key);
}

function renderSchedEditor(key) {
  const c = schedCampers.find(x => x.key === key);
  const box = document.getElementById('sched-results');
  if (!c) { renderSchedResults(); return; }
  const def = c.days || 'MTWRF';
  const ov = schedOverrides[key] || {};
  const enrolled = schedWeeks.filter(w => c.weeks && c.weeks[w.n - 1]);
  let h = `<button class="sched-back" id="sched-back">← Back to search</button>`;
  h += `<div style="font-weight:700;color:var(--brand-dark);font-size:1rem">${famEsc(c.name)}</div>`;
  h += `<div style="font-size:.82rem;color:#888;margin-bottom:.7rem">${famEsc(c.bunk)}</div>`;
  if (!enrolled.length) h += '<div style="color:#aaa;font-size:.85rem">This camper is not enrolled in any week.</div>';
  enrolled.forEach(w => {
    const isOv = ov[String(w.n)] != null;
    const cur = isOv ? ov[String(w.n)] : def;
    h += `<div class="sched-wk">` +
      `<span class="sw-label">Week ${w.n} <span style="color:#999;font-weight:400">${w.range || ''}</span>${isOv ? '<span class="sched-ov">(custom)</span>' : ''}</span>` +
      SCHED_DAYS.map(L => `<button class="sched-day${cur.includes(L) ? ' on' : ''}" data-wk="${w.n}" data-day="${L}">${SCHED_DAY_LABEL[L]}</button>`).join('') +
      (isOv ? `<button class="sched-back" style="margin:0" data-reset="${w.n}">reset</button>` : '') +
      `</div>`;
  });
  if (enrolled.length) {
    h += `<div style="margin-top:.9rem;display:flex;align-items:center;gap:.7rem">` +
      `<button class="pr-period-btn" id="sched-save"${schedDirty ? '' : ' disabled style=\"opacity:.55;cursor:default\"'}>💾 Save schedule</button>` +
      `<span id="sched-savemsg" style="font-size:.82rem;color:${schedDirty ? '#b26a00' : '#777'}">${schedDirty ? 'Unsaved changes' : 'All changes saved'}</span>` +
      `</div>`;
  }
  box.innerHTML = h;
  document.getElementById('sched-back').addEventListener('click', schedBack);
  box.querySelectorAll('.sched-day').forEach(btn =>
    btn.addEventListener('click', () => toggleSchedDay(key, parseInt(btn.dataset.wk, 10), btn.dataset.day)));
  box.querySelectorAll('[data-reset]').forEach(btn =>
    btn.addEventListener('click', () => resetSchedWeek(key, parseInt(btn.dataset.reset, 10))));
  const saveBtn = document.getElementById('sched-save');
  if (saveBtn) saveBtn.addEventListener('click', () => saveSchedule(key, true));
}

async function schedBack() {
  if (schedDirty && schedCurrentKey) await saveSchedule(schedCurrentKey);
  schedCurrentKey = null; schedDirty = false;
  renderSchedResults();
}

function toggleSchedDay(key, wk, day) {
  const c = schedCampers.find(x => x.key === key);
  const ovc = schedOverrides[key] || (schedOverrides[key] = {});
  const cur = ovc[String(wk)] != null ? ovc[String(wk)] : (c.days || 'MTWRF');
  const set = new Set(cur.split(''));
  set.has(day) ? set.delete(day) : set.add(day);
  ovc[String(wk)] = SCHED_DAYS.filter(L => set.has(L)).join('');
  schedDirty = true;
  renderSchedEditor(key);
}

function resetSchedWeek(key, wk) {
  if (schedOverrides[key]) { delete schedOverrides[key][String(wk)]; if (!Object.keys(schedOverrides[key]).length) delete schedOverrides[key]; }
  schedDirty = true;
  renderSchedEditor(key);
}

async function saveSchedule(key, fromButton) {
  key = key || schedCurrentKey;
  if (!key) return;
  const msg = fromButton ? document.getElementById('sched-savemsg') : null;
  if (msg) { msg.textContent = 'Saving…'; msg.style.color = '#777'; }
  try {
    const res = await fetch('/api/schedules', {method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({key, replace: schedOverrides[key] || {}})});
    if (!res.ok) throw new Error('save failed');
    schedDirty = false;
    if (fromButton && schedCurrentKey === key) renderSchedEditor(key);   // refresh button/state
  } catch(e) {
    if (msg) { msg.textContent = 'Save failed — try again'; msg.style.color = '#c0392b'; }
  }
}

(function(){ const s = document.getElementById('sched-search'); if (s) s.addEventListener('input', renderSchedResults); })();

// ---- Bunk Snapshot viewer (admin) ----
// Data only changes when a new master sheet is uploaded, so we cache the last
// snapshot in localStorage and reuse it unless the master's upload timestamp
// has changed. That means no spinner / reload on normal opens.
const SNAP_CACHE_KEY = 'el_snap_cache_v2';
let snapRendered = false;
let snapCurWeek = null, snapCurDay = null;   // today's camp week (1-8) / weekday (0-4) for column highlight

function _snapPaint(d) {
  renderSnapMeta(d.meta || {});
  renderSnapTotals(d.totals || {});
  renderSnapBunks(d.report || []);
  snapRendered = true;
}

async function loadBunkSnapshot(force) {
  if (!currentUser) return;
  const metaEl = document.getElementById('snap-meta');

  // 1) Paint instantly from cache (even across page reloads) if we have one.
  let cache = null;
  try { cache = JSON.parse(localStorage.getItem(SNAP_CACHE_KEY) || 'null'); } catch(e) {}
  if (!snapRendered) {
    if (cache && cache.report) _snapPaint(cache);
    else metaEl.innerHTML = '<span>📋</span><span>Loading…</span>';
  }

  // 2) Cheap check: did the master sheet change since our cached copy?
  //    Also grab today's week/day (changes daily, independent of the master).
  let curStamp = '', hasMaster = true;
  try {
    const mr = await fetch('/api/master');
    if (mr.ok) {
      const md = await mr.json();
      hasMaster = !!md.loaded; curStamp = md.uploaded_at || '';
      snapCurWeek = (md.current_week ?? null); snapCurDay = (md.current_day ?? null);
    }
  } catch(e) {}
  if (!force && cache && cache.report && hasMaster && curStamp && cache.uploaded_at === curStamp) {
    if (cache && cache.report) _snapPaint(cache);   // repaint so today's column highlight is applied
    return;
  }

  // 3) Master changed (or no cache / forced) — fetch fresh and re-cache.
  try {
    const res = await fetch('/api/bunk-snapshot');
    if (!res.ok) { if (!snapRendered) metaEl.innerHTML = '<span>⚠️</span><span>Could not load snapshot.</span>'; return; }
    const d = await res.json();
    if (!d.has_master) {
      metaEl.innerHTML = '<span>📋</span><span>No master sheet uploaded yet. Upload one in Utilities to see the snapshot.</span>';
      document.getElementById('snap-totals').innerHTML = '';
      document.getElementById('snap-bunks').innerHTML = '';
      try { localStorage.removeItem(SNAP_CACHE_KEY); } catch(e) {}
      snapRendered = true; return;
    }
    if (d.error) { if (!snapRendered) metaEl.innerHTML = '<span>⚠️</span><span>' + famEsc(d.error) + '</span>'; return; }
    snapCurWeek = (d.current_week ?? snapCurWeek); snapCurDay = (d.current_day ?? snapCurDay);
    _snapPaint(d);
    try {
      localStorage.setItem(SNAP_CACHE_KEY, JSON.stringify({
        uploaded_at: (d.meta && d.meta.uploaded_at) || curStamp || '',
        meta: d.meta, totals: d.totals, report: d.report }));
    } catch(e) {}
  } catch(e) { if (!snapRendered) metaEl.innerHTML = '<span>⚠️</span><span>Could not load snapshot.</span>'; }
}

function renderSnapMeta(m) {
  const who = (m.uploaded_by || '').replace(/@.*$/, '');
  let when = (m.uploaded_at || '').replace(/\s*[A-Z]{2,4}\s*$/, '');
  let txt = 'Data last updated';
  if (when) txt += ': ' + when;
  if (who) txt += ' by ' + who;
  if (m.filename) txt += ' (' + m.filename + ')';
  document.getElementById('snap-meta').innerHTML = '<span>📋</span><span>' + famEsc(txt) + '</span>';
}

function _snapTbl(headLeft, rows, opts) {
  // rows: array of {cells:[...], cls:''}; first cell is left-aligned label
  opts = opts || {};
  const wk = opts.weeks;         // true => header is #1..#8
  const hl = opts.hlCol ?? -1;   // column index to highlight (label=0, week n = n)
  let ci = 0;
  let h = '<table class="snap-tbl"><thead><tr>';
  if (Array.isArray(headLeft)) headLeft.forEach(hh => { h += `<th class="${ci===0?'snap-l':''}${ci===hl?' snap-hl':''}">${famEsc(hh)}</th>`; ci++; });
  if (wk) for (let i=1;i<=8;i++) { h += `<th class="${ci===hl?'snap-hl':''}">#${i}</th>`; ci++; }
  h += '</tr></thead><tbody>';
  rows.forEach(r => {
    h += `<tr class="${r.cls||''}">`;
    r.cells.forEach((c,i) => h += `<td class="${i===0?'snap-l':''}${i===hl?' snap-hl':''}">${c===''||c==null?'':famEsc(String(c))}</td>`);
    h += '</tr>';
  });
  return h + '</tbody></table>';
}

function renderSnapTotals(t) {
  // Bunk Totals (Camp | Bunk | Total)
  const bunkRows = (t.bunk_totals||[]).map((b,i)=>({cls:i%2?'snap-alt':'',cells:[b.camp,b.bunk,b.total]}));
  bunkRows.push({cls:'snap-total',cells:['TOTAL','',t.bunk_grand]});
  // Group Totals (Camp | Total)
  const grpRows = (t.group_totals||[]).map((g,i)=>({cls:i%2?'snap-alt':'',cells:[g.label,g.total]}));
  grpRows.push({cls:'snap-total',cells:['Total',t.group_grand]});
  // Group Totals by Week
  const gwRows = (t.group_by_week||[]).map((g,i)=>({cls:i%2?'snap-alt':'',cells:[g.label,...g.weeks]}));
  gwRows.push({cls:'snap-total',cells:['Total',...(t.week_total||[])]});
  // Bunk Totals by Week
  const bwRows = (t.bunk_by_week||[]).map((b,i)=>({cls:i%2?'snap-alt':'',cells:[b.bunk,...b.weeks]}));
  bwRows.push({cls:'snap-total',cells:['Total',...(t.week_total||[])]});

  document.getElementById('snap-totals').innerHTML =
    '<div class="snap-grids">' +
      '<div>' +
        '<div class="snap-sec-title">Bunk Totals</div>' + _snapTbl(['Camp','Bunk','Total'], bunkRows) +
        '<div class="snap-sec-title" style="margin-top:1.2rem">Group Totals</div>' + _snapTbl(['Camp','Total'], grpRows) +
      '</div>' +
      '<div>' +
        '<div class="snap-sec-title">Group Totals by Week</div>' + _snapTbl(['Group'], gwRows, {weeks:true, hlCol:snapCurWeek}) +
        '<div class="snap-sec-title" style="margin-top:1.2rem">Bunk Totals by Week</div>' + _snapTbl(['Bunk'], bwRows, {weeks:true, hlCol:snapCurWeek}) +
      '</div>' +
    '</div>';
}

function renderSnapBunks(report) {
  let h = '<input type="search" id="snap-bunk-search" class="pr-input snap-search" placeholder="Search camper name…">';
  report.forEach(b => {
    h += '<div class="snap-bunk-block" data-block>';
    h += `<div class="snap-bunk-name">${famEsc(b.bunk)}</div>`;
    const wkCls = n => (n === snapCurWeek ? ' snap-hl' : '');
    const dayCls = di => (di === 0 ? 'snap-sep' : '') + (di === snapCurDay ? ' snap-hl' : '');
    h += '<table class="snap-tbl"><thead><tr>' +
         '<th class="snap-l">Child</th>';
    for (let n = 1; n <= 8; n++) h += `<th class="${wkCls(n).trim()}">#${n}</th>`;
    ['M','T','W','R','F'].forEach((L,di) => h += `<th class="${dayCls(di).trim()}">${L}</th>`);
    h += '<th class="snap-sep">Age</th><th>Grade</th></tr></thead><tbody>';
    b.campers.forEach((c,i) => {
      h += `<tr class="${i%2?'snap-alt':''}" data-n="${famEsc((c.name||'').toLowerCase())}">`;
      h += `<td class="snap-l">${famEsc(c.name)}</td>`;
      c.weeks.forEach((w,wi) => h += `<td class="${wkCls(wi+1).trim()}">${w?w:''}</td>`);
      c.days.forEach((d,di) => h += `<td class="${dayCls(di).trim()}">${famEsc(d||'')}</td>`);
      h += `<td class="snap-sep">${c.age===''?'':famEsc(String(c.age))}</td><td>${famEsc(String(c.grade||''))}</td>`;
      h += '</tr>';
    });
    // Total row
    h += '<tr class="snap-total"><td class="snap-l">Total: ' + b.total + '</td>';
    b.week_sums.forEach((s,wi) => h += `<td class="${wkCls(wi+1).trim()}">${s}</td>`);
    ['M','T','W','R','F'].forEach((L,di) => h += `<td class="${dayCls(di).trim()}"></td>`);
    h += '<td class="snap-sep"></td><td></td></tr>';
    h += '</tbody></table></div>';
  });
  const box = document.getElementById('snap-bunks');
  box.innerHTML = h;
  const search = document.getElementById('snap-bunk-search');
  if (search) search.addEventListener('input', () => {
    const q = (search.value||'').trim().toLowerCase();
    box.querySelectorAll('[data-block]').forEach(blk => {
      let shown = 0;
      blk.querySelectorAll('tr[data-n]').forEach(tr => {
        const hit = !q || tr.dataset.n.includes(q);
        tr.style.display = hit ? '' : 'none';
        if (hit) shown++;
      });
      const totalRow = blk.querySelector('tr.snap-total');
      if (totalRow) totalRow.style.display = q ? 'none' : '';
      blk.style.display = (q && shown === 0) ? 'none' : '';
    });
  });
}

// ---- Families directory ----
let famDir = [], famDirWeeks = [], famDirLoaded = false, famDirState = {has_families:true, has_master:true};
let famEditKey = null, famEditSched = {};   // which card is being edited + working schedule copy
const FAM_DAY_LBL = {M:'M', T:'T', W:'W', R:'Th', F:'F'};

async function loadFamiliesDir(force) {
  if (!currentUser) return;
  if (famDirLoaded && !force) return;
  try {
    const res = await fetch('/api/families/full');
    if (!res.ok) return;
    const d = await res.json();
    famDir = d.families || [];
    famDirWeeks = d.weeks || [];
    famDirState = {has_families: !!d.has_families, has_master: !!d.has_master};
    famDirLoaded = true;
  } catch(e) {}
  renderFamDir();
}

function renderFamDir() {
  const box = document.getElementById('fam-dir-results');
  if (!box) return;
  if (!famDirState.has_families) {
    box.innerHTML = '<div style="color:#888;font-size:.9rem">No family contacts loaded yet. Import a Family Contacts spreadsheet in <strong>Utilities</strong> first.</div>';
    return;
  }
  const q = (document.getElementById('fam-dir-search').value || '').trim().toLowerCase();
  if (!q) { box.innerHTML = '<div style="color:#aaa;font-size:.85rem">Start typing a camper, parent, or family name…</div>'; return; }
  const hits = famDir.filter(f => f.search.includes(q)).slice(0, 25);
  if (!hits.length) { box.innerHTML = '<div style="color:#aaa;font-size:.85rem">No matching families.</div>'; return; }
  box.innerHTML = '<div style="font-size:.8rem;color:#888;margin-bottom:.6rem">' + hits.length + (famDir.filter(f=>f.search.includes(q)).length>25?'+ ':' ') + 'famil' + (hits.length===1?'y':'ies') + ' found</div>' +
    '<div class="fam-cards">' + hits.map(famCardHTML).join('') + '</div>';
  wireFamDir();
}

function wireFamDir() {
  const box = document.getElementById('fam-dir-results');
  if (!box) return;
  box.querySelectorAll('[data-fedit]').forEach(btn =>
    btn.addEventListener('click', () => enterFamEdit(btn.dataset.fedit)));
  if (!famEditKey) return;
  box.querySelectorAll('.fam-eday').forEach(btn => btn.addEventListener('click', () => {
    const ci = btn.dataset.ci, wk = btn.dataset.wk, day = btn.dataset.day;
    const cur = (famEditSched[ci] && famEditSched[ci][wk]) || '';
    const set = new Set(cur.split(''));
    set.has(day) ? set.delete(day) : set.add(day);
    famEditSched[ci] = famEditSched[ci] || {};
    famEditSched[ci][wk] = ['M','T','W','R','F'].filter(L => set.has(L)).join('');
    btn.classList.toggle('on');
  }));
  const sv = box.querySelector('.fam-save');   if (sv) sv.addEventListener('click', saveFamEdit);
  const cn = box.querySelector('.fam-cancel'); if (cn) cn.addEventListener('click', () => { famEditKey = null; renderFamDir(); });
}

function enterFamEdit(key) {
  famEditKey = key;
  const f = famDir.find(x => x.key === key);
  famEditSched = {};
  if (f) f.campers.forEach((c, ci) => {
    famEditSched[ci] = {};
    (c.weeks_detail || []).forEach(w => { famEditSched[ci][String(w.n)] = w.days || ''; });
  });
  renderFamDir();
}

function famEditFormHTML(f) {
  const F = f.fields || {};
  const inp = (field, ph) => `<input class="fam-inp" data-field="${field}" value="${famEsc(F[field] || '')}" placeholder="${famEsc(ph || '')}">`;
  let h = `<div class="fam-card" data-fkey="${famEsc(f.key)}"><div class="fam-card-hd"><span>The ${famEsc(f.name)} Family</span><span style="font-size:.78rem;opacity:.85">Editing</span></div><div class="fam-card-bd">`;

  // Per-camper schedules
  h += '<div class="fam-sec"><div class="fam-sec-h">Schedule' + (f.campers.length>1?'s':'') + '</div>';
  f.campers.forEach((c, ci) => {
    h += '<div class="fam-camper"><div class="fam-camper-name">' + famEsc(c.name) + '</div>';
    if (!c.in_master) h += '<div class="fam-note">Not in the current master sheet.</div>';
    else if (!c.weeks_detail.length) h += '<div class="fam-note">Not enrolled in any week.</div>';
    else c.weeks_detail.forEach(w => {
      const cur = (famEditSched[ci] && famEditSched[ci][String(w.n)]) || '';
      h += '<div class="fam-wk-row"><span class="fam-wk-lbl">Week ' + w.n + (w.range ? ' · ' + famEsc(w.range) : '') + '</span><span>' +
        ['M','T','W','R','F'].map(L => `<span class="fam-day fam-eday${cur.includes(L) ? ' on' : ''}" data-ci="${ci}" data-wk="${w.n}" data-day="${L}">${FAM_DAY_LBL[L]}</span>`).join('') +
        '</span></div>';
    });
    h += '</div>';
  });
  h += '</div>';

  // Contact form (shared by the whole family)
  h += '<div class="fam-sec"><div class="fam-sec-h">Contacts</div>';
  h += '<div class="fam-flbl">Address</div>' + inp('address','Street address') + inp('address2','Apt / unit (optional)');
  h += '<div class="fam-inp-row">' + inp('city','City') + inp('state','State') + inp('zip','Zip') + '</div>';
  h += '<div class="fam-flbl">Primary parent</div>';
  h += '<div class="fam-inp-row">' + inp('primary_first','First') + inp('primary_last','Last') + '</div>';
  h += '<div class="fam-inp-row">' + inp('primary_phone','Phone') + inp('primary_email','Email') + '</div>';
  h += '<div class="fam-flbl">Secondary parent</div>';
  h += '<div class="fam-inp-row">' + inp('secondary_first','First') + inp('secondary_last','Last') + '</div>';
  h += '<div class="fam-inp-row">' + inp('secondary_phone','Phone') + inp('secondary_email','Email') + '</div>';
  h += '<div class="fam-flbl">Authorized pickups</div>';
  for (let i = 1; i <= 4; i++) h += '<div class="fam-inp-row">' + inp('pu'+i+'_name','Pickup '+i+' name') + inp('pu'+i+'_auth','Authorization') + '</div>';
  h += '</div>';

  h += '<div class="fam-edit-actions"><button class="pr-period-btn fam-save">💾 Save</button><button class="sched-back fam-cancel" style="margin:0">Cancel</button><span class="fam-edit-msg"></span></div>';
  return h + '</div></div>';
}

async function saveFamEdit() {
  const f = famDir.find(x => x.key === famEditKey);
  if (!f) return;
  const box = document.getElementById('fam-dir-results');
  const msg = box.querySelector('.fam-edit-msg');
  const fields = {};
  box.querySelectorAll('.fam-inp').forEach(i => { fields[i.dataset.field] = i.value.trim(); });
  if (msg) { msg.textContent = 'Saving…'; msg.style.color = '#777'; }
  try {
    // Contact info applies to every record in the family (siblings share it)
    for (const id of (f.ids || [])) {
      await fetch('/api/families/' + encodeURIComponent(id), {method:'PATCH', headers:{'Content-Type':'application/json'}, body: JSON.stringify(fields)});
    }
    // Schedules: only the weeks the user actually changed, per camper
    for (let ci = 0; ci < f.campers.length; ci++) {
      const c = f.campers[ci];
      if (!c.in_master || !c.sched_key) continue;
      const orig = {}; (c.weeks_detail || []).forEach(w => { orig[String(w.n)] = w.days || ''; });
      const cur = famEditSched[ci] || {};
      for (const wk of Object.keys(cur)) {
        if ((cur[wk] || '') !== (orig[wk] || '')) {
          await fetch('/api/schedules', {method:'POST', headers:{'Content-Type':'application/json'},
            body: JSON.stringify({key: c.sched_key, week: parseInt(wk, 10), days: cur[wk]})});
        }
      }
    }
    famEditKey = null;
    famDirLoaded = false;
    snapRendered = false;   // contact/schedule changes ripple into other views
    await loadFamiliesDir(true);
  } catch(e) {
    if (msg) { msg.textContent = 'Save failed — try again'; msg.style.color = '#c0392b'; }
  }
}

function famDays(days) {
  return ['M','T','W','R','F'].map(L =>
    `<span class="fam-day${(days||'').includes(L) ? ' on' : ''}">${FAM_DAY_LBL[L]}</span>`).join('');
}

function famCardHTML(f) {
  if (f.key === famEditKey) return famEditFormHTML(f);
  let h = `<div class="fam-card" data-fkey="${famEsc(f.key)}"><div class="fam-card-hd"><span>The ${famEsc(f.name)} Family</span><button class="fam-edit-btn" data-fedit="${famEsc(f.key)}">✎ Edit</button></div><div class="fam-card-bd">`;

  // Campers
  h += '<div class="fam-sec"><div class="fam-sec-h">Camper' + (f.campers.length>1?'s':'') + '</div>';
  f.campers.forEach(c => {
    const meta = [c.bunk ? 'Bunk: ' + c.bunk : '', (c.age!==''&&c.age!=null) ? 'Age: ' + c.age : '', c.grade ? 'Grade: ' + c.grade : ''].filter(Boolean).join(' · ');
    h += '<div class="fam-camper"><div class="fam-camper-name">' + famEsc(c.name) + '</div>';
    if (meta) h += '<div class="fam-camper-meta">' + famEsc(meta) + '</div>';
    if (!c.in_master) {
      h += '<div class="fam-note">Not found in the current master sheet (no schedule).</div>';
    } else if (!c.weeks_detail.length) {
      h += '<div class="fam-note">Not enrolled in any week.</div>';
    } else {
      c.weeks_detail.forEach(w => {
        h += '<div class="fam-wk-row"><span class="fam-wk-lbl">Week ' + w.n + (w.range ? ' · ' + famEsc(w.range) : '') + '</span><span>' + famDays(w.days) + '</span></div>';
      });
    }
    h += '</div>';
  });
  h += '</div>';

  // Address
  const a = f.address || {};
  const line2 = [a.city, a.state].filter(Boolean).join(', ') + (a.zip ? ' ' + a.zip : '');
  if (a.address || line2.trim()) {
    h += '<div class="fam-sec"><div class="fam-sec-h">Address</div>';
    if (a.address) h += '<div class="fam-row">' + famEsc(a.address) + '</div>';
    if (a.address2) h += '<div class="fam-row">' + famEsc(a.address2) + '</div>';
    if (line2.trim()) h += '<div class="fam-row">' + famEsc(line2.trim()) + '</div>';
    h += '</div>';
  }

  // Contacts
  const ct = f.contacts || {};
  const hasC = (ct.primary && (ct.primary.name || ct.primary.phone || ct.primary.email)) || (ct.secondary && (ct.secondary.name || ct.secondary.phone || ct.secondary.email)) || (ct.emails_other && ct.emails_other.length) || (ct.pickups && ct.pickups.length);
  if (hasC) {
    h += '<div class="fam-sec"><div class="fam-sec-h">Contacts</div>';
    [['Primary', ct.primary], ['Secondary', ct.secondary]].forEach(([lbl, p]) => {
      if (p && (p.name || p.phone || p.email)) {
        const bits = [p.phone, p.email].filter(Boolean).map(famEsc).join(' · ');
        h += '<div class="fam-row"><span class="fam-lbl">' + lbl + ':</span>' + famEsc(p.name || '—') + (bits ? ' · ' + bits : '') + '</div>';
      }
    });
    (ct.emails_other || []).forEach(em => {
      h += '<div class="fam-row"><span class="fam-lbl">Email:</span>' + famEsc(em) + '</div>';
    });
    (ct.pickups || []).forEach(pu => {
      h += '<div class="fam-pickup"><span class="fam-lbl">Pickup:</span>' + famEsc(pu.name) + (pu.auth ? ' <span style="color:#888">(' + famEsc(pu.auth) + ')</span>' : '') + '</div>';
    });
    h += '</div>';
  }

  return h + '</div></div>';
}

(function(){ const s = document.getElementById('fam-dir-search'); if (s) s.addEventListener('input', renderFamDir); })();

// ---- Pricing module (admin) ----
let pricing = null, pxLoaded = false;
let pxTier = 'ES';
let pxCampers = [{weeks:'8', days:'5', transport:'none'}];
let pxCC = {on:false, days:'5', kids:1, weeks:0};
let pxExp = {pct:3, round:50};
const pxMoney = n => '$' + (Math.round(Number(n)||0)).toLocaleString('en-US');

async function loadPricing(force) {
  if (!currentUser || !currentUser.is_admin) return;
  if (pxLoaded && !force) return;
  try {
    const r = await fetch('/api/pricing');
    if (!r.ok) return;
    pricing = await r.json();
    pxLoaded = true;
  } catch(e) { return; }
  renderPxCalc(); renderPxExplore(); renderPxRates(); renderPxSheet();
}

// Sub-tab switching
function pxShowSub(name) {
  document.querySelectorAll('.pxtab').forEach(b => b.classList.toggle('on', b.dataset.px === name));
  document.querySelectorAll('.px-view').forEach(v => v.classList.toggle('on', v.id === 'px-' + name));
}
document.querySelectorAll('.pxtab').forEach(btn => btn.addEventListener('click', () => pxShowSub(btn.dataset.px)));

// ---------- Calculator ----------
function renderPxCalc() {
  if (!pricing) return;
  const box = document.getElementById('px-calc');
  const weekOpts = pricing.camp.week_order || ['8','7','6','5','4','Mini'];
  const trLabel = {none:'None', '1way':'1-way', '2way':'2-way'};
  let h = '<div class="px-sec"><div class="px-sec-title">Summer Camp</div>';
  h += '<div class="px-controls"><label class="px-field">Rate<select id="px-tier">' +
    `<option value="ES"${pxTier==='ES'?' selected':''}>Early Signup (fall)</option>` +
    `<option value="Final"${pxTier==='Final'?' selected':''}>Regular</option></select></label></div>`;
  pxCampers.forEach((c,i) => {
    const opt = (v,cur) => `<option value="${v}"${cur===v?' selected':''}>${v}</option>`;
    h += '<div class="px-camper-row" data-i="'+i+'">' +
      '<span style="font-weight:700;color:#555">Camper '+(i+1)+'</span>' +
      '<label class="px-field">Weeks<select data-f="weeks">'+weekOpts.map(w=>opt(w,c.weeks)).join('')+'</select></label>' +
      '<label class="px-field">Days/wk<select data-f="days">'+['5','4','3'].map(d=>opt(d,c.days)).join('')+'</select></label>' +
      '<label class="px-field">Transport<select data-f="transport">'+['none','1way','2way'].map(t=>`<option value="${t}"${c.transport===t?' selected':''}>${trLabel[t]}</option>`).join('')+'</select></label>' +
      (pxCampers.length>1?'<button class="px-btn ghost px-rm" data-i="'+i+'" style="padding:.35rem .6rem">✕</button>':'') +
      '</div>';
  });
  h += '<button class="px-btn ghost" id="px-add-camper">＋ Add camper</button></div>';
  h += '<div class="px-sec"><div class="px-sec-title">Childcare / School <span style="font-weight:400;color:#999;font-size:.8rem">(separate from camp)</span></div><div class="px-controls">' +
    '<label class="px-field">Include<select id="px-cc-on"><option value="no"'+(pxCC.on?'':' selected')+'>No</option><option value="yes"'+(pxCC.on?' selected':'')+'>Yes</option></select></label>' +
    '<label class="px-field">Days/wk<select id="px-cc-days">'+['5','4','3'].map(d=>`<option value="${d}"${pxCC.days===d?' selected':''}>${d}</option>`).join('')+'</select></label>' +
    '<label class="px-field">Children<select id="px-cc-kids"><option value="1"'+(pxCC.kids===1?' selected':'')+'>1</option><option value="2"'+(pxCC.kids===2?' selected':'')+'>2 (siblings)</option></select></label>' +
    '<label class="px-field">Weeks<input type="number" id="px-cc-weeks" min="0" value="'+(pxCC.weeks||'')+'" placeholder="#"></label>' +
    '</div></div><div id="px-calc-total"></div>';
  box.innerHTML = h;
  document.getElementById('px-tier').addEventListener('change', e => { pxTier = e.target.value; renderPxCalc(); });
  box.querySelectorAll('.px-camper-row').forEach(row => {
    const i = +row.dataset.i;
    row.querySelectorAll('select[data-f]').forEach(sel => sel.addEventListener('change', () => { pxCampers[i][sel.dataset.f] = sel.value; renderPxCalc(); }));
  });
  box.querySelectorAll('.px-rm').forEach(b => b.addEventListener('click', () => { pxCampers.splice(+b.dataset.i,1); renderPxCalc(); }));
  document.getElementById('px-add-camper').addEventListener('click', () => { pxCampers.push({weeks:'8',days:'5',transport:'none'}); renderPxCalc(); });
  ['px-cc-on','px-cc-days','px-cc-kids','px-cc-weeks'].forEach(id => {
    document.getElementById(id).addEventListener('change', () => {
      pxCC.on = document.getElementById('px-cc-on').value === 'yes';
      pxCC.days = document.getElementById('px-cc-days').value;
      pxCC.kids = +document.getElementById('px-cc-kids').value;
      pxCC.weeks = +document.getElementById('px-cc-weeks').value || 0;
      renderPxCalc();
    });
  });
  renderPxCalcTotal();
}

function renderPxCalcTotal() {
  const tiers = pricing.camp.tiers[pxTier] || {}, dm = pricing.camp.day_mult || {};
  let lines = [], campGrand = 0;
  pxCampers.forEach((c,i) => {
    const base = Number(tiers[c.weeks]||0);
    const mult = Number(dm[c.days]!=null ? dm[c.days] : 1);
    const tuition = base*mult; let sub = tuition;
    let detail = pxMoney(tuition)+' tuition ('+(c.weeks==='Mini'?'Mini':c.weeks+' wks')+(mult!==1?' ×'+mult:'')+')';
    const wksNum = parseInt(c.weeks,10)||0;
    if (c.transport!=='none' && wksNum) {
      const wkr = Number((pricing.transport[c.transport]||{})[c.days]||0);
      const tr = wkr*wksNum; sub += tr;
      detail += ' + '+pxMoney(tr)+' transport ('+pxMoney(wkr)+'/wk × '+wksNum+')';
    }
    campGrand += sub;
    lines.push('<div class="px-line"><span>Camper '+(i+1)+': '+detail+'</span><span>'+pxMoney(sub)+'</span></div>');
  });
  const tierLbl = pxTier==='ES' ? 'Early Signup' : 'Regular';
  let html = '<div class="px-total-box">'+lines.join('')+'<div class="px-line px-sub"><span>Camp total ('+tierLbl+')</span><span class="px-grand">'+pxMoney(campGrand)+'</span></div></div>';
  if (pxCC.on) {
    const row = pricing.childcare[pxCC.days] || {};
    const weekly = pxCC.kids===2 ? Number(row.sibling2||0) : Number(row.base||0);
    const wks = pxCC.weeks||0; const ccTotal = wks ? weekly*wks : weekly;
    const lbl = pxCC.kids===2 ? '2 siblings' : '1 child';
    html += '<div class="px-total-box" style="margin-top:.8rem"><div class="px-line"><span>Childcare ('+pxCC.days+' days/wk, '+lbl+')'+(wks?': '+pxMoney(weekly)+'/wk × '+wks+' wks':'')+'</span><span>'+pxMoney(ccTotal)+'</span></div>' +
      '<div class="px-line px-sub"><span>Childcare total'+(wks?'':' (weekly rate)')+'</span><span class="px-grand">'+pxMoney(ccTotal)+'</span></div></div>';
  }
  document.getElementById('px-calc-total').innerHTML = html;
}

// ---------- Explorer ----------
function pxRound(x) { const s = pxExp.round; return s ? Math.round(x/s)*s : Math.round(x); }
function pxExpTable(title, table, order, rowLabel) {
  let h = '<div class="px-sec"><div class="px-sec-title">'+title+'</div><table class="px-tbl"><thead><tr>' +
    '<th class="px-l">'+rowLabel+'</th><th>Current</th><th>+'+pxExp.pct+'%</th><th>$ diff</th><th>% diff</th></tr></thead><tbody>';
  order.forEach(k => {
    const base = Number(table[k]||0);
    const nw = pxRound(base*(1+pxExp.pct/100));
    const diff = nw-base, pct = base ? (diff/base*100) : 0;
    const cls = diff ? 'px-diff-up' : 'px-diff-flat';
    h += '<tr><td class="px-l">'+k+'</td><td>'+pxMoney(base)+'</td><td>'+pxMoney(nw)+'</td>' +
      '<td class="'+cls+'">'+(diff>0?'+':'')+pxMoney(diff)+'</td><td class="'+cls+'">'+pct.toFixed(2)+'%</td></tr>';
  });
  return h+'</tbody></table></div>';
}
function renderPxExplore() {
  if (!pricing) return;
  const box = document.getElementById('px-explore');
  let h = '<div class="px-controls">' +
    '<label class="px-field">% increase<input type="number" id="px-exp-pct" step="0.1" value="'+pxExp.pct+'"></label>' +
    '<label class="px-field">Round to<select id="px-exp-round">'+[['0','$1'],['25','$25'],['50','$50'],['100','$100']].map(([v,l])=>`<option value="${v}"${pxExp.round==+v?' selected':''}>${l}</option>`).join('')+'</select></label>' +
    '<button class="px-btn" id="px-apply" style="align-self:center">Apply to Rate Settings</button>' +
    '</div>' +
    '<div style="font-size:.8rem;color:#888;margin-bottom:1rem">This is a preview. Use <strong>Apply to Rate Settings</strong> to copy these camp tuition and childcare figures into the editable rates, where you can review and Save them. Transportation is not changed by the explorer.</div>';
  const ccBase = {}, ccOrder = ['5','4','3'];
  ccOrder.forEach(d => ccBase[d] = pricing.childcare[d].base);
  h += pxExpTable('Summer Camp — Early Signup (ES)', pricing.camp.tiers.ES, pricing.camp.week_order, 'Weeks');
  h += pxExpTable('Summer Camp — Regular (Final)', pricing.camp.tiers.Final, pricing.camp.week_order, 'Weeks');
  h += pxExpTable('Childcare — 1 child (weekly)', ccBase, ccOrder, 'Days/wk');
  box.innerHTML = h;
  document.getElementById('px-exp-pct').addEventListener('input', e => { pxExp.pct = parseFloat(e.target.value)||0; renderPxExplore(); });
  document.getElementById('px-exp-round').addEventListener('change', e => { pxExp.round = parseInt(e.target.value,10); renderPxExplore(); });
  document.getElementById('px-apply').addEventListener('click', pxApplyExplore);
}

function pxApplyExplore() {
  const roundLbl = pxExp.round ? ('$' + pxExp.round) : '$1';
  if (!confirm('Apply a ' + pxExp.pct + '% increase (rounded to ' + roundLbl + ') to the camp tuition (ES and Regular) and childcare rates? The new figures load into Rate Settings for you to review, and nothing is saved until you click Save there.')) return;
  const bump = obj => Object.keys(obj).forEach(k => { obj[k] = pxRound(Number(obj[k]||0) * (1 + pxExp.pct/100)); });
  bump(pricing.camp.tiers.ES);
  bump(pricing.camp.tiers.Final);
  ['5','4','3'].forEach(d => {
    const r = pricing.childcare[d];
    r.base = pxRound(Number(r.base||0) * (1 + pxExp.pct/100));
    r.sibling2 = pxRound(Number(r.sibling2||0) * (1 + pxExp.pct/100));
  });
  renderPxRates();
  pxShowSub('rates');
  const msg = document.getElementById('px-save-msg');
  if (msg) { msg.textContent = 'New rates loaded. Review, then click Save rates to keep them.'; msg.style.color = '#b26a00'; }
}

// ---------- Rate Settings ----------
function renderPxRates() {
  if (!pricing) return;
  const box = document.getElementById('px-rates');
  const wk = pricing.camp.week_order;
  const inp = (id,v) => `<input class="px-rate-inp" id="${id}" value="${v==null?'':v}">`;
  let h = '<div class="px-controls"><label class="px-field">Season label<input id="px-season-label" style="min-width:110px;text-align:left" value="'+famEsc(pricing.season_label||'')+'"></label></div>';
  h += '<div class="px-sec"><div class="px-sec-title">Summer Camp tuition</div><table class="px-tbl"><thead><tr><th class="px-l">Weeks</th><th>Early Signup (ES)</th><th>Regular (Final)</th></tr></thead><tbody>';
  wk.forEach(w => h += '<tr><td class="px-l">'+w+'</td><td>'+inp('px-camp-ES-'+w, pricing.camp.tiers.ES[w])+'</td><td>'+inp('px-camp-Final-'+w, pricing.camp.tiers.Final[w])+'</td></tr>');
  h += '</tbody></table>';
  h += '<div style="font-size:.8rem;color:#888;margin:.6rem 0 .3rem"><strong>Day rate factor:</strong> the percentage of the full 5-day tuition charged to a 4-day or 3-day camper, where 100% is the full rate (for example, 90% applies a 10% reduction). Each value is currently 100%, so the number of days attended does not yet affect camp tuition.</div>';
  const pct = v => parseFloat((Number(v||0) * 100).toFixed(2));
  const pctInp = (id,v) => inp(id, pct(v)) + ' %';
  h += '<table class="px-tbl"><thead><tr><th>5-day</th><th>4-day</th><th>3-day</th></tr></thead><tbody><tr>'+
    '<td style="color:#888">100% <span style="font-size:.75rem">(base)</span></td><td>'+pctInp('px-dm-4', pricing.camp.day_mult['4'])+'</td><td>'+pctInp('px-dm-3', pricing.camp.day_mult['3'])+'</td></tr></tbody></table></div>';
  h += '<div class="px-sec"><div class="px-sec-title">Transportation (weekly)</div><table class="px-tbl"><thead><tr><th class="px-l">Type</th><th>5-day</th><th>4-day</th><th>3-day</th></tr></thead><tbody>';
  [['2way','2-way'],['1way','1-way']].forEach(([k,l]) => h += '<tr><td class="px-l">'+l+'</td><td>'+inp('px-tr-'+k+'-5', pricing.transport[k]['5'])+'</td><td>'+inp('px-tr-'+k+'-4', pricing.transport[k]['4'])+'</td><td>'+inp('px-tr-'+k+'-3', pricing.transport[k]['3'])+'</td></tr>');
  h += '</tbody></table></div>';
  h += '<div class="px-sec"><div class="px-sec-title">Childcare / School (weekly)</div><table class="px-tbl"><thead><tr><th class="px-l">Days/wk</th><th>1 child</th><th>2 siblings (combined)</th></tr></thead><tbody>';
  ['5','4','3'].forEach(d => h += '<tr><td class="px-l">'+d+'</td><td>'+inp('px-cc-'+d+'-base', pricing.childcare[d].base)+'</td><td>'+inp('px-cc-'+d+'-sib', pricing.childcare[d].sibling2)+'</td></tr>');
  h += '</tbody></table></div>';
  h += '<div style="margin-top:.5rem"><button class="px-btn" id="px-save">💾 Save rates</button><span class="px-msg" id="px-save-msg"></span></div>';
  box.innerHTML = h;
  document.getElementById('px-save').addEventListener('click', savePxRates);
}

async function savePxRates() {
  const num = id => { const el = document.getElementById(id); const v = parseFloat((el && el.value || '').replace(/[^0-9.]/g,'')); return isNaN(v) ? 0 : v; };
  const lblEl = document.getElementById('px-season-label');
  if (lblEl) pricing.season_label = lblEl.value.trim();
  pricing.camp.week_order.forEach(w => { pricing.camp.tiers.ES[w] = num('px-camp-ES-'+w); pricing.camp.tiers.Final[w] = num('px-camp-Final-'+w); });
  ['5','4','3'].forEach(d => { pricing.childcare[d].base = num('px-cc-'+d+'-base'); pricing.childcare[d].sibling2 = num('px-cc-'+d+'-sib'); });
  pricing.camp.day_mult['5'] = 1;   // 5-day is the base (always 100%)
  ['4','3'].forEach(d => { pricing.camp.day_mult[d] = num('px-dm-'+d) / 100; });
  ['2way','1way'].forEach(k => ['5','4','3'].forEach(d => pricing.transport[k][d] = num('px-tr-'+k+'-'+d)));
  const msg = document.getElementById('px-save-msg'); msg.textContent = 'Saving…'; msg.style.color = '#777';
  try {
    const r = await fetch('/api/pricing', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(pricing)});
    if (!r.ok) throw 0;
    msg.textContent = '✓ Saved'; msg.style.color = '#2e7d32';
    renderPxCalc(); renderPxExplore(); renderPxSheet();
  } catch(e) { msg.textContent = 'Save failed — try again'; msg.style.color = '#c0392b'; }
}

// ---------- Rate Sheet (printable one-pager) ----------
function renderPxSheet() {
  if (!pricing) return;
  const box = document.getElementById('px-sheet');
  const wk = pricing.camp.week_order;
  let h = '<div style="margin-bottom:.8rem" class="px-noprint"><button class="px-btn" id="px-print">🖨 Print / Save PDF</button></div>';
  h += '<div class="px-sheet-head"><h2>Elbow Lane Day Camp</h2><div style="color:#666;font-weight:600">'+famEsc(pricing.season_label||'')+' Rates</div></div>';
  h += '<div class="px-sheet-grid">';
  // Camp tuition
  h += '<div><div class="px-sec-title">Summer Camp Tuition</div><table class="px-tbl"><thead><tr><th class="px-l">Weeks</th><th>Early Signup</th><th>Regular</th></tr></thead><tbody>';
  wk.forEach(w => h += '<tr><td class="px-l">'+w+'</td><td>'+pxMoney(pricing.camp.tiers.ES[w])+'</td><td>'+pxMoney(pricing.camp.tiers.Final[w])+'</td></tr>');
  h += '</tbody></table></div>';
  // Transport
  h += '<div><div class="px-sec-title">Transportation (weekly)</div><table class="px-tbl"><thead><tr><th class="px-l">Type</th><th>5-day</th><th>4-day</th><th>3-day</th></tr></thead><tbody>';
  [['2way','2-way'],['1way','1-way']].forEach(([k,l]) => h += '<tr><td class="px-l">'+l+'</td><td>'+pxMoney(pricing.transport[k]['5'])+'</td><td>'+pxMoney(pricing.transport[k]['4'])+'</td><td>'+pxMoney(pricing.transport[k]['3'])+'</td></tr>');
  h += '</tbody></table></div>';
  // Childcare
  h += '<div><div class="px-sec-title">Childcare / School (weekly)</div><table class="px-tbl"><thead><tr><th class="px-l">Days/wk</th><th>1 child</th><th>2 siblings</th></tr></thead><tbody>';
  ['5','4','3'].forEach(d => h += '<tr><td class="px-l">'+d+'</td><td>'+pxMoney(pricing.childcare[d].base)+'</td><td>'+pxMoney(pricing.childcare[d].sibling2)+'</td></tr>');
  h += '</tbody></table></div>';
  h += '</div>';
  box.innerHTML = h;
  document.getElementById('px-print').addEventListener('click', () => {
    let st = document.getElementById('px-print-style');
    if (!st) { st = document.createElement('style'); st.id = 'px-print-style'; document.head.appendChild(st); }
    st.textContent = '@media print{ body *{visibility:hidden!important} #px-sheet,#px-sheet *{visibility:visible!important} #px-sheet{position:absolute;left:0;top:0;width:100%} .px-noprint{display:none!important} @page{margin:.5in} }';
    window.print();
  });
}

// ---- First-time "Utilities" notice (shows once per browser, after sign-in) ----
function maybeShowNotice() {
  const KEY = 'el_seen_utilities_notice_v1';
  const overlay = document.getElementById('notice-overlay');
  const close = () => { overlay.classList.add('hidden'); try { localStorage.setItem(KEY, '1'); } catch(e) {} };
  try {
    if (localStorage.getItem(KEY)) return;
  } catch(e) {}
  overlay.classList.remove('hidden');
  document.getElementById('notice-ok').onclick = close;
  overlay.addEventListener('click', e => { if (e.target === overlay) close(); });
}

// ---- Pricing modal ----
(function() {
  const overlay = document.getElementById('pricing-overlay');
  document.getElementById('pricing-btn').addEventListener('click', () => overlay.classList.remove('hidden'));
  document.getElementById('pricing-close').addEventListener('click', () => overlay.classList.add('hidden'));
  overlay.addEventListener('click', e => { if (e.target === overlay) overlay.classList.add('hidden'); });
})();

// ---- Authentication (per-user login accounts) ----
let currentUser = null;

async function loadUsers() {
  // Admin-only; the endpoint 403s for non-admins, in which case we hide the card.
  const card = document.getElementById('users-card');
  if (!currentUser || !currentUser.is_admin) { if (card) card.style.display = 'none'; return; }
  try {
    const res = await fetch('/api/users');
    if (!res.ok) { card.style.display = 'none'; return; }
    const d = await res.json();
    const tbl = document.getElementById('users-table');
    let h = '<thead><tr><th>Username</th><th>Role</th><th></th></tr></thead><tbody>';
    (d.users || []).forEach(u => {
      const isMe = currentUser && u.username.toLowerCase() === currentUser.username.toLowerCase();
      h += `<tr><td>${famEsc(u.username)}${isMe ? ' (you)' : ''}</td>` +
           `<td>${u.is_admin ? 'Admin' : 'User'}</td>` +
           `<td style="white-space:nowrap"><button class="usr-ic usr-rename" data-u="${famEsc(u.username)}" title="Rename">✎</button> ` +
           `<button class="pr-period-btn pr-sm usr-pw" data-u="${famEsc(u.username)}">Reset PW</button> ` +
           `${isMe ? '' : `<button class="pr-del usr-del" data-u="${famEsc(u.username)}" title="Remove">✕</button>`}</td></tr>`;
    });
    h += '</tbody>';
    tbl.innerHTML = h;
    card.style.display = '';
    tbl.querySelectorAll('.usr-del').forEach(btn => {
      btn.addEventListener('click', async () => {
        const un = btn.dataset.u;
        if (!confirm(`Remove user "${un}"?`)) return;
        try { await fetch('/api/users/' + encodeURIComponent(un), {method:'DELETE'}); } catch(e) {}
        loadUsers();
      });
    });
    tbl.querySelectorAll('.usr-pw').forEach(btn => {
      btn.addEventListener('click', async () => {
        const un = btn.dataset.u;
        const pw = prompt(`New password for "${un}" (min 4 characters):`);
        if (!pw) return;
        const r = await fetch('/api/users/' + encodeURIComponent(un), {method:'PATCH',
          headers:{'Content-Type':'application/json'}, body: JSON.stringify({password: pw})});
        const dd = await r.json();
        alert(r.ok && !dd.error ? `Password reset for ${un}.` : (dd.error || 'Could not reset password.'));
      });
    });
    tbl.querySelectorAll('.usr-rename').forEach(btn => {
      btn.addEventListener('click', async () => {
        const un = btn.dataset.u;
        const nu = prompt(`New username for "${un}":`, un);
        if (!nu || nu.trim() === un) return;
        const r = await fetch('/api/users/' + encodeURIComponent(un), {method:'PATCH',
          headers:{'Content-Type':'application/json'}, body: JSON.stringify({username: nu.trim()})});
        const dd = await r.json();
        if (!r.ok || dd.error) { alert(dd.error || 'Could not rename.'); return; }
        // If we renamed ourselves, update the header chip
        if (currentUser && currentUser.username.toLowerCase() === un.toLowerCase()) {
          currentUser.username = dd.username;
          document.getElementById('h-user-name').textContent = dd.username;
        }
        loadUsers();
      });
    });
  } catch(e) { card.style.display = 'none'; }
}

document.getElementById('usr-add').addEventListener('click', async () => {
  const msg = document.getElementById('usr-msg');
  const username = document.getElementById('usr-username').value.trim();
  const password = document.getElementById('usr-password').value;
  const email    = document.getElementById('usr-email').value.trim();
  const body = { username, password, is_admin: document.getElementById('usr-admin').checked };
  if (!username || !password) { msg.style.color = '#c0392b'; msg.textContent = 'Username and password required.'; return; }
  try {
    const res = await fetch('/api/users', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
    const d = await res.json();
    if (!res.ok || d.error) { msg.style.color = '#c0392b'; msg.textContent = d.error || 'Could not add user.'; return; }
    msg.style.color = '#2e7d32'; msg.textContent = `✓ Added ${d.username}.`;

    // Build a shareable credentials message + Copy / Email actions
    const url = window.location.origin;
    const message =
      `Elbow Lane Reporting Center — your login\n` +
      `Site: ${url}\n` +
      `Username: ${username}\n` +
      `Password: ${password}\n\n` +
      `Sign in at the link above. Keep this private.`;
    document.getElementById('usr-creds').textContent = message;
    const subject = 'Your Elbow Lane Reporting Center login';
    document.getElementById('usr-email-link').href =
      `mailto:${encodeURIComponent(email)}?subject=${encodeURIComponent(subject)}&body=${encodeURIComponent(message)}`;
    document.getElementById('usr-copy-msg').textContent = '';
    document.getElementById('usr-result').style.display = '';

    ['usr-username','usr-password','usr-email'].forEach(id => document.getElementById(id).value = '');
    document.getElementById('usr-admin').checked = false;
    loadUsers();
  } catch(e) { msg.style.color = '#c0392b'; msg.textContent = 'Network error: ' + e.message; }
});

document.getElementById('usr-copy').addEventListener('click', async () => {
  const text = document.getElementById('usr-creds').textContent;
  const cm = document.getElementById('usr-copy-msg');
  try {
    await navigator.clipboard.writeText(text);
    cm.textContent = 'Copied!';
  } catch(e) {
    // Fallback: select the text for manual copy
    const r = document.createRange(); r.selectNodeContents(document.getElementById('usr-creds'));
    const sel = window.getSelection(); sel.removeAllRanges(); sel.addRange(r);
    cm.textContent = 'Press Ctrl/Cmd+C to copy.';
  }
  setTimeout(() => { cm.textContent = ''; }, 4000);
});

(function() {
  const overlay  = document.getElementById('pw-overlay');
  const errEl    = document.getElementById('pw-error');
  const loginView = document.getElementById('login-view');
  const regView   = document.getElementById('register-view');

  function showApp(user) {
    currentUser = user;
    overlay.classList.add('hidden');
    document.getElementById('h-user').style.display = 'flex';
    document.getElementById('h-user-name').textContent = user.name || user.username;
    const pxNav = document.getElementById('tab-pricing-nav');   // admin-only
    if (pxNav) pxNav.style.display = user.is_admin ? '' : 'none';
    loadAllData();
    maybeShowNotice();
  }

  document.getElementById('show-register').addEventListener('click', () => {
    loginView.style.display = 'none'; regView.style.display = ''; errEl.textContent = '';
  });
  document.getElementById('show-login').addEventListener('click', () => {
    regView.style.display = 'none'; loginView.style.display = ''; errEl.textContent = '';
  });

  async function doLogin() {
    errEl.textContent = '';
    const body = {
      username: document.getElementById('login-username').value.trim(),
      password: document.getElementById('login-password').value,
    };
    if (!body.username || !body.password) { errEl.textContent = 'Enter your username and password.'; return; }
    try {
      const res = await fetch('/api/login', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
      const d = await res.json();
      if (!res.ok || d.error) { errEl.textContent = d.error || 'Sign-in failed.'; return; }
      showApp(d);
    } catch(e) { errEl.textContent = 'Network error: ' + e.message; }
  }

  async function doRegister() {
    errEl.textContent = '';
    const body = {
      username: document.getElementById('reg-username').value.trim(),
      password: document.getElementById('reg-password').value,
      code:     document.getElementById('reg-code').value.trim(),
    };
    if (!body.username || !body.password) { errEl.textContent = 'Choose a username and password.'; return; }
    try {
      const res = await fetch('/api/register', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
      const d = await res.json();
      if (!res.ok || d.error) { errEl.textContent = d.error || 'Could not create account.'; return; }
      showApp(d);
    } catch(e) { errEl.textContent = 'Network error: ' + e.message; }
  }

  document.getElementById('login-btn').addEventListener('click', doLogin);
  document.getElementById('reg-btn').addEventListener('click', doRegister);
  document.getElementById('login-password').addEventListener('keydown', e => { if (e.key === 'Enter') doLogin(); });
  document.getElementById('reg-code').addEventListener('keydown', e => { if (e.key === 'Enter') doRegister(); });

  // Account dropdown (click your name)
  const userMenu = document.getElementById('h-user-menu');
  function closeMenu() { userMenu.classList.add('hidden'); }
  document.getElementById('h-user-btn').addEventListener('click', e => {
    e.stopPropagation();
    userMenu.classList.toggle('hidden');
  });
  document.addEventListener('click', () => closeMenu());
  document.getElementById('menu-logout').addEventListener('click', async () => {
    closeMenu();
    try { await fetch('/api/logout', {method:'POST'}); } catch(e) {}
    location.reload();
  });

  // Reset Password menu item → change-password dialog
  const cpw = document.getElementById('cpw-overlay');
  const cpwErr = document.getElementById('cpw-error');
  function openCpw() {
    if (!currentUser) return;
    document.getElementById('cpw-who').textContent = currentUser.username;
    ['cpw-current','cpw-new','cpw-confirm'].forEach(id => document.getElementById(id).value = '');
    cpwErr.textContent = '';
    cpw.classList.remove('hidden');
    document.getElementById('cpw-current').focus();
  }
  function closeCpw() { cpw.classList.add('hidden'); }
  document.getElementById('menu-reset').addEventListener('click', () => { closeMenu(); openCpw(); });
  document.getElementById('cpw-cancel').addEventListener('click', closeCpw);
  cpw.addEventListener('click', e => { if (e.target === cpw) closeCpw(); });
  document.getElementById('cpw-save').addEventListener('click', async () => {
    const current = document.getElementById('cpw-current').value;
    const nw = document.getElementById('cpw-new').value;
    const conf = document.getElementById('cpw-confirm').value;
    cpwErr.textContent = '';
    if (!current || !nw) { cpwErr.textContent = 'Fill in all fields.'; return; }
    if (nw.length < 4) { cpwErr.textContent = 'New password must be at least 4 characters.'; return; }
    if (nw !== conf) { cpwErr.textContent = 'New passwords do not match.'; return; }
    try {
      const res = await fetch('/api/account/password', {method:'POST',
        headers:{'Content-Type':'application/json'}, body: JSON.stringify({current, new: nw})});
      const d = await res.json();
      if (!res.ok || d.error) { cpwErr.textContent = d.error || 'Could not change password.'; return; }
      closeCpw();
      alert('Password changed.');
    } catch(e) { cpwErr.textContent = 'Network error: ' + e.message; }
  });

  // On load: if already signed in, go straight in; otherwise show the gate.
  (async () => {
    try {
      const res = await fetch('/api/me');
      const d = await res.json();
      if (d.authenticated) { showApp(d); return; }
      // No accounts yet → default to the create-account view
      if (!d.has_users) { loginView.style.display = 'none'; regView.style.display = ''; }
    } catch(e) {}
    document.getElementById('login-username').focus();
  })();
})();
</script>
</body>
</html>
"""

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5001)
